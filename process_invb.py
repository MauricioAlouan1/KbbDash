"""
This script, `process_inv.py`, is designed to handle inventory data files, primarily for the KBB MF project. 
Its primary goals are as follows:
1. Define and validate the base directory for accessing inventory data files.
2. Set up date range variables to specify the start and end periods for processing.
3. Implement functions to:
   - Process and stack inventory files for specific months and years.
   - Format data for consistent and streamlined reporting.
4. Organize files within a directory structure (e.g., `/clean/YYYY_MM/`) for efficient data retrieval and processing.

Key Features:
- Dynamic handling of inventory files based on specified year and month.
- Robust error handling to ensure smooth execution even if files or directories are missing.
- Integration with other scripts and workflows in the KBB MF project.

Prerequisites:
- Ensure that the base directory paths specified in `path_options` exist and contain the necessary inventory files.
- Verify that the `/clean/YYYY_MM/` directory structure is consistent with the expected format.

This script is integral to maintaining the accuracy and efficiency of inventory management workflows.
"""


import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook, load_workbook
import sys

# Define the base directory as before, now adding the /clean part
path_options = [
    '/Users/mauricioalouan/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data/',
    '/Users/simon/Library/CloudStorage/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data'
]
for path in path_options:
    if os.path.exists(path):
        base_dir = path
        break
else:
    print("None of the specified directories exist.")
    base_dir = None

# Define the date range variables
# ---- minimal prompt block (like Conc_Estoque.py) ----
def _prompt_year_month():
    import argparse
    from datetime import datetime
    ap = argparse.ArgumentParser(add_help=False)
    ap.add_argument("--year", "-y", type=int)
    ap.add_argument("--month", "-m", type=int)
    args, _ = ap.parse_known_args()
    if args.year is not None and args.month is not None:
        return args.year, args.month
    # If missing, ask interactively; defaults mimic Conc_Estoque.py (month defaults to previous)
    now = datetime.now()
    default_year = now.year if now.month > 1 else now.year - 1
    default_month = now.month - 1 if now.month > 1 else 12
    print("Year and/or month not provided.")
    year = int(input(f"Enter year (default {default_year}): ") or default_year)
    month = int(input(f"Enter month [1-12] (default {default_month}): ") or default_month)
    return year, month

start_year, start_month = _prompt_year_month()
end_year, end_month = start_year, start_month
# ---- end minimal prompt block ----

# Function to process inventory files for a given month and year
def process_inventory_files(year, month):
    """Process and stack inventory files for a given year and month."""         
    try:
        # Format the month to always be two digits (e.g., 01, 02, ..., 12)
        month_str = f'{month:02d}'
        
        # The files for each month are inside the /clean/YYYY_MM/ folder
        clean_folder = os.path.join(base_dir, f'clean/{year}_{month_str}')

        # Define all file types with their corresponding 'Local' values
        file_configs = {
            f'B_Estoq_{year}_{month_str}_clean.xlsx': 'Bling',
            f'T_EstTrans_{year}_{month_str}_clean.xlsx': 'Transito',
            f'O_Estoq_{year}_{month_str}_clean.xlsx': None,  # Special case with its own column
            f'B_EFullAj_{year}_{month_str}_clean.xlsx': 'Ajuste',
            f'B_EFullAm_{year}_{month_str}_clean.xlsx': 'Amazon Full',
            f'B_EFullMg_{year}_{month_str}_clean.xlsx': 'Magalu Full',
            f'B_EFullML_{year}_{month_str}_clean.xlsx': 'ML Full'
        }

        combined_dfs = []

        # Process each file
        for file_name, local_value in file_configs.items():
            file_path = os.path.join(clean_folder, file_name)
            try:
                if os.path.exists(file_path):
                    if 'O_Estoq' in file_name:
                        # Special handling for O_Estoq
                        df = pd.read_excel(file_path, usecols=['Código do Produto', 'Quantidade', 'Local de Estoque (Código)'])
                        df.rename(columns={
                              'Código do Produto': 'Codigo',
                            'Quantidade': 'Quantidade',
                            'Local de Estoque (Código)': 'Local'
                        }, inplace=True)
                    elif 'T_EstTrans' in file_name:
                        # Special handling for T_EstTrans
                        df = pd.read_excel(file_path, usecols=['CodProd', 'Qt'])
                        df.rename(columns={'CodProd': 'Codigo', 'Qt': 'Quantidade'}, inplace=True)
                        df['Local'] = 'Transito'
                    else:
                        # General handling
                        df = pd.read_excel(file_path, usecols=['Código', 'Quantidade'])
                        df.rename(columns={'Código': 'Codigo', 'Quantidade': 'Quantidade'}, inplace=True)
                        if local_value:
                            df['Local'] = local_value
                    combined_dfs.append(df)
                else:
                    print(f"File not found: {file_name}. Skipping this file.")
            except Exception as e:
                print(f"Error processing inventory files for {year}-{month_str}, file prefix: {file_name}: {e}")
                continue  # Skip this file and proceed with the next

        # Combine all dataframes
        if combined_dfs:
            combined_df = pd.concat(combined_dfs, ignore_index=True)
        else:
            print(f"No files found for {year}-{month_str}. Returning an empty DataFrame.")
            combined_df = pd.DataFrame(columns=['Codigo', 'Quantidade', 'Local'])

        return combined_df

    except Exception as e:
        print(f"Error processing inventory files for {year}-{month_str}: {e}")
        return None



# Function to lookup CU values and additional columns
# Function to lookup CU values and additional columns
def lookup_cu_values(inventory_df, cutoff_date):

    try:
        entradas_df = pd.read_excel(
            os.path.join(base_dir, 'Tables', 'T_Entradas.xlsx'),
            dtype={'Pai': str, 'Filho': str}
        )

        prodf_df = pd.read_excel(
            os.path.join(base_dir, 'Tables', 'T_ProdF.xlsx'),
            dtype={'CodPF': str, 'CodPP': str}
        )

        # ➤ Converte a data da última entrada
        # ➤ Remove rows with missing Pai and convert AnoMes
        entradas_df = entradas_df[entradas_df['Pai'].notna() & (entradas_df['Pai'] != '')].copy()
        #entradas_df['AnoMes'] = entradas_df['AnoMes'].astype(str).str.zfill(4).astype(int)
        cutoff_anomes = (cutoff_date.year - 2000) * 100 + cutoff_date.month + 0

        # ➤ Filter only up to cutoff
        entradas_cutoff = entradas_df[entradas_df['AnoMes'] <= cutoff_anomes].copy()

        # ➤ Split entries into two categories
        # 1. Before current month → keep latest row per Pai (CUF already computed)
        latest_before = (
            entradas_cutoff[entradas_cutoff['AnoMes'] < cutoff_anomes]
            .sort_values(['Pai', 'AnoMes'])
            .drop_duplicates(subset='Pai', keep='last')[['Pai', 'CUF']]
            .rename(columns={'CUF': 'UCP'})
        )

        # 2. Current month → compute CUEm weighted avg per Pai
        current_month_df = entradas_cutoff[entradas_cutoff['AnoMes'] == cutoff_anomes].copy()
        current_month_df['WeightedCU'] = (current_month_df['Ult CU R$'] + current_month_df['AddR']) * current_month_df['Qt']
        cuem = (
            current_month_df.groupby('Pai')
            .agg({'WeightedCU': 'sum', 'Qt': 'sum'})
            .query('Qt > 0')
            .assign(CUEm=lambda x: x['WeightedCU'] / x['Qt'])[['CUEm']]
            .reset_index()
        )

        # ➤ Merge both UCP (CUF) and CUEm
        cu_final = pd.merge(latest_before, cuem, on='Pai', how='outer')


        inventory_df['Codigo'] = inventory_df['Codigo'].astype(str)
        inventory_df.rename(columns={'Quantidade': 'Quantidade_Inv', 'Codigo': 'Codigo_Inv'}, inplace=True)

        prodf_df.rename(columns={'CodPF': 'CodPF_Prod', 'CodPP': 'CodPP_Prod'}, inplace=True)

        inventory_df = pd.merge(inventory_df, prodf_df[['CodPF_Prod', 'CodPP_Prod']], 
                                left_on='Codigo_Inv', right_on='CodPF_Prod', how='left')

        inventory_df = pd.merge(inventory_df, cu_final,
                                left_on='CodPP_Prod', right_on='Pai', how='left')

        inventory_df['Quantidade_Inv'] = pd.to_numeric(inventory_df['Quantidade_Inv'], errors='coerce').fillna(0)
        inventory_df['UCP'] = pd.to_numeric(inventory_df['UCP'], errors='coerce').fillna(0)

        # Use CUF if available; otherwise use CUEm
        inventory_df['UCU'] = inventory_df['UCP'].where(inventory_df['UCP'] > 0, inventory_df['CUEm'])
        inventory_df['UCT'] = inventory_df['UCU'] * inventory_df['Quantidade_Inv']

        return inventory_df

    except Exception as e:
        print(f"Error looking up CU values: {e}")
        return None


# Main function to handle the process for all months within the date range
def process_all_months():
    # Loop through each year and month in the specified range
    for year in range(start_year, end_year + 1):
        for month in range(1, 13):
            if year == start_year and month < start_month:
                continue
            if year == end_year and month > end_month:
                break

            print(f"Processing data for year {year}, month {month:02d}")

            # Step 1: Process and stack inventory data for the given year and month
            inventory_df = process_inventory_files(year, month)
            if inventory_df is None:
                continue

            # Step 2: Lookup CU values and calculate UCU and UCT
            cutoff_date = pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0)
            final_df = lookup_cu_values(inventory_df, cutoff_date)

            if final_df is None:
                continue

            # Add AnoMes
            final_df['AnoMes'] = (year % 100) * 100 + month
            # Step 3: Save the resulting dataframe to a new Excel file
            output_filepath = os.path.join(base_dir, 'clean',f'{year}_{month:02d}', f'R_Estoq_fdm_{year}_{month:02d}.xlsx')
            final_df.to_excel(output_filepath, index=False, sheet_name='Data')
            print(f"Saved combined inventory data for {year}-{month:02d} to {output_filepath}")
            format_and_add_pivot(output_filepath, final_df, year,month)
            print(f"Added Formating and Pivots for {year}-{month:02d} to {output_filepath}")

# Format and add pivot tables using openpyxl
def format_and_add_pivot(output_filepath, df, year, month):
    # Load the workbook
    wb = load_workbook(output_filepath)
    ws = wb['Data']
    
    # Apply number format to specified columns
    number_format = '#,##0.00'
    columns_to_format = ['UCP', 'UCF', 'UCU', 'UCT']
    # Add autofilter to the pivot table
    ws.auto_filter.ref = ws.dimensions

    for col in columns_to_format:
        if col in df.columns:
            col_idx = df.columns.get_loc(col) + 1  # Adjust for Excel's 1-based indexing
            for row in range(2, len(df) + 2):  # Start from the second row (excluding header)
                cell = ws.cell(row=row, column=col_idx)
                cell.number_format = number_format
    
    # Create a pivot table on a new sheet
    pivot_table = df.pivot_table(
       index='Codigo_Inv',
        columns='Local',
        values='Quantidade_Inv',
        aggfunc='sum',
        fill_value=0
    )
############################################
    # Ensure total column is correct
    pivot_table['Total'] = pivot_table.sum(axis=1)

    # Add a total cost column
    total_cost = df.groupby('Codigo_Inv')['UCT'].sum()
    pivot_table['Total Cost'] = total_cost

    # Add a unit cost column
    pivot_table['Unit Cost'] = pivot_table['Total Cost'] / pivot_table['Total']
    # Add AnoMes
    pivot_table['AnoMes'] = (year % 100) * 100 + month

    # Validate totals
    original_total_cost = df['UCT'].sum()
    pivot_total_cost = pivot_table['Total Cost'].sum()

    print(f"Original Total Cost: {original_total_cost}")
    print(f"Pivot Table Total Cost: {pivot_total_cost}")

    # Check for mismatched rows
    unmatched_rows = df[~df['Codigo_Inv'].isin(pivot_table.index)]
    if not unmatched_rows.empty:
        print("Unmatched rows found:")
        print(unmatched_rows)
    else:
        print('### No unmatched rows ###')
        
############################################

    # Reset the index for better readability
    pivot_table = pivot_table.reset_index()

    # Add the pivot table to a new sheet
    pivot_sheet_name = 'PT01'
    if pivot_sheet_name not in wb.sheetnames:
        wb.create_sheet(title=pivot_sheet_name)
    pivot_ws = wb[pivot_sheet_name]
    
    # Write headers
    for col_idx, header in enumerate(pivot_table.columns, start=1):
        pivot_ws.cell(row=1, column=col_idx, value=header)

    # Write the pivot table data
    for r_idx, row in enumerate(pivot_table.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            pivot_ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Add totals row at the bottom
    totals_row_idx = len(pivot_table) + 2
    pivot_ws.cell(row=totals_row_idx, column=1, value="Grand Total").font = Font(bold=True)
    for col_idx, col_name in enumerate(pivot_table.columns[1:], start=2):  # Skip 'CodPF_Prod'
        total_value = pivot_table[col_name].sum()
        cell = pivot_ws.cell(row=totals_row_idx, column=col_idx, value=total_value)
        cell.font = Font(bold=True)
        if col_name in ['Total Cost', 'Unit Cost']:
            cell.number_format = number_format

    # Apply formatting to the last 3 columns (bold + light gray fill)
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for col_name in ['Total', 'Total Cost', 'Unit Cost']:
        col_idx = pivot_table.columns.get_loc(col_name) + 1
        for row in range(2, totals_row_idx + 1):  # Include totals row
            cell = pivot_ws.cell(row=row, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = gray_fill
            cell.number_format = number_format

    # Add autofilter to the pivot table
    pivot_ws.auto_filter.ref = pivot_ws.dimensions

    # Save the workbook
    wb.save(output_filepath)


if __name__ == "__main__":
    process_all_months()
