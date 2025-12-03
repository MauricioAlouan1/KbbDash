import os
import pandas as pd
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────
# 1. Prompt Logic
# ─────────────────────────────────────────────────────────────
def _prompt_year_month(year=None, month=None):
    import argparse
    from datetime import datetime
    
    if year is not None and month is not None:
        return year, month

    ap = argparse.ArgumentParser(add_help=False)
    ap.add_argument("--year", "-y", type=int)
    ap.add_argument("--month", "-m", type=int)
    args, _ = ap.parse_known_args()

    if args.year is not None and args.month is not None:
        return args.year, args.month

    now = datetime.now()
    default_year = now.year if now.month > 1 else now.year - 1
    default_month = now.month - 1 if now.month > 1 else 12
    print("Year and/or month not provided.")
    try:
        year = int(input(f"Enter year (default {default_year}): ") or default_year)
        month = int(input(f"Enter month [1-12] (default {default_month}): ") or default_month)
    except EOFError:
        print("Non-interactive mode detected. Using defaults.")
        year, month = default_year, default_month
        
    return year, month

def resolve_base_dir():
    path_options = [
        '/Users/mauricioalouan/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data/',
        '/Users/simon/Library/CloudStorage/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data'
    ]
    for path in path_options:
        if os.path.exists(path):
            return path
    raise FileNotFoundError("❌ None of the specified directories exist.")

# ─────────────────────────────────────────────────────────────
# 2. Data Loading and Processing
# ─────────────────────────────────────────────────────────────
def load_entrada_df(file_path: str) -> pd.DataFrame:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns)

    if "Pai" in df.columns:
        df["Pai"] = df["Pai"].astype(str).str.strip()
    if "AnoMes" in df.columns:
        df["AnoMes"] = pd.to_numeric(df["AnoMes"], errors="coerce").astype("Int64")
    for col in ["CU_E", "CU_F"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

def apply_prev_month_values(df: pd.DataFrame, base_dir: str, year: int, month: int) -> pd.DataFrame:
    """
    Fills Qt_I and CU_I_new based on previous month's Conc_Estoq_<tag>.xlsx.
    - Qt_I = previous month's Qt_SS
    - CU_I_new = previous month's CU_F
    - If not found: Qt_I = 0, CU_I_new = CU_E
    """
    import numpy as np

    prev_year, prev_month = get_prev_month(year, month)
    tag = f"{prev_year:04d}_{prev_month:02d}"
    file_path = os.path.join(base_dir, "clean", tag, f"Conc_Estoq_{tag}.xlsx")

    if not os.path.exists(file_path):
        print(f"⚠️ Arquivo anterior não encontrado: {file_path}")
        df["Qt_I"] = 0
        df["CU_I_new"] = df["CU_E"]
        return df

    print(f"🔁 Lendo dados do mês anterior: {file_path}")
    prev_df = pd.read_excel(file_path, sheet_name="Conc", dtype={"CODPP": str})
    prev_df["CODPP"] = prev_df["CODPP"].astype(str).str.strip().str.upper()
    prev_df["Qt_SS"] = pd.to_numeric(prev_df["Qt_SS"], errors="coerce").fillna(0)
    prev_df["CU_F"]   = pd.to_numeric(prev_df["CU_F"], errors="coerce").fillna(0)

    df["Pai"] = df["Pai"].astype(str).str.strip().str.upper()

    # Merge both Qt_SS and CU_F
    df = df.merge(
        prev_df[["CODPP", "Qt_SS", "CU_F"]].rename(columns={"CU_F": "CU_F_prev", "Qt_SS": "Qt_SS_prev"}),
        left_on="Pai", right_on="CODPP", how="left"
    )

    # Apply fallback logic for new items
    df["Qt_I"] = np.where(df["Qt_SS_prev"].notna(), df["Qt_SS_prev"], 0)
    df["CU_I_new"] = np.where(df["CU_F_prev"].notna() & (df["CU_F_prev"] > 0), df["CU_F_prev"], df["CU_E"])

    print(f"✅ Aplicados valores do mês anterior para {df['Qt_SS_prev'].notna().sum()} itens existentes.")
    print(f"🆕 Itens novos com Qt_I=0 e CU_I=CU_E: {(df['Qt_SS_prev'].isna()).sum()}")

    return df


# ─────────────────────────────────────────────────────────────
# 3. Excel Writing (with formatting preserved)
# ────────────────────────────────────────────────────────
def write_column_to_excel(df, excel_path, out_path, ano_mes, column_name, values_series):
    wb = load_workbook(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col_map = {str(h): i + 1 for i, h in enumerate(headers)}

    # Add column if missing
    if column_name not in col_map:
        col_idx = len(headers) + 1
        ws.cell(row=1, column=col_idx, value=column_name)
        col_map[column_name] = col_idx

    # Normalize keys
    values_series.index = values_series.index.astype(str).str.strip()
    written = 0
    print(f"\n📝 Linhas escritas em '{column_name}' (AnoMes = {ano_mes}):")

    for row_idx in range(2, ws.max_row + 1):
        cell_pai = str(ws.cell(row=row_idx, column=col_map.get("Pai", 0)).value or "").strip()
        cell_anomes = ws.cell(row=row_idx, column=col_map.get("AnoMes", 0)).value

        if cell_anomes == ano_mes and cell_pai in values_series.index:
            val = values_series.loc[cell_pai]
            if val is not None and (pd.notna(val) or val == 0):
                ws.cell(row=row_idx, column=col_map[column_name], value=float(val))
                written += 1
                print(f"→ Linha {row_idx}: Pai={cell_pai}, {column_name}={val}")

    wb.save(out_path)
    print(f"📤 Returning written={written} for column '{column_name}'")
    return written

def write_column_by_index_to_excel(df, excel_path, out_path, column_name, values_series):
    wb = load_workbook(excel_path)
    ws = wb.active

    # Mapear cabeçalhos
    headers = [cell.value for cell in ws[1]]
    col_map = {str(h): i + 1 for i, h in enumerate(headers)}

    # Criar coluna se não existir
    if column_name not in col_map:
        col_idx = len(headers) + 1
        ws.cell(row=1, column=col_idx, value=column_name)
        col_map[column_name] = col_idx

    # Forçar índices como inteiros (linhas do Excel)
    values_series.index = values_series.index.astype(int)
    written = 0

    print(f"\n📝 Linhas escritas por índice em '{column_name}':")
    for idx, val in values_series.items():
        excel_row = idx + 2  # Excel data starts at row 2

        if pd.notna(val) or val == 0:
            ws.cell(row=excel_row, column=col_map[column_name], value=float(val))
            written += 1
            print(f"→ Linha {excel_row}: index={idx}, {column_name}={val}")

    wb.save(out_path)
    print(f"📤 Returning written={written} for column '{column_name}'")
    return written

def calculate_qtsp_from_invoices(base_dir: str, year: int, month: int) -> pd.DataFrame:
    # Try to locate the 'nfs' directory relative to the dropbox root
    # base_dir is usually .../Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data
    # We want .../Dropbox/nfs
    
    dropbox_root = None
    if "/Dropbox/" in base_dir:
        dropbox_root = base_dir.split("/Dropbox/")[0] + "/Dropbox"
    else:
        # Fallback or assumption
        dropbox_root = os.path.dirname(base_dir) # Go up one level? No, that's too deep.
        # Let's try standard paths
        possible_roots = [
            '/Users/mauricioalouan/Dropbox',
            '/Users/simon/Library/CloudStorage/Dropbox'
        ]
        for p in possible_roots:
            if os.path.exists(p):
                dropbox_root = p
                break
    
    if not dropbox_root:
        print("⚠️ Could not find Dropbox root to locate invoices.")
        return pd.DataFrame(columns=["CODPP", "Qt_S"])

    nfi_path = os.path.join(dropbox_root, "nfs", "Mauricio", "Contabilidade - Tsuriel", f"NFI_{year}_{month:02d}_todos.xlsx")
    
    if not os.path.exists(nfi_path):
        print(f"⚠️ Invoice file not found: {nfi_path}")
        return pd.DataFrame(columns=["CODPP", "Qt_S"])

    print(f"📖 Reading invoices from: {nfi_path}")
    try:
        df = pd.read_excel(nfi_path)
        
        # Filter for Sales if needed? 
        # Assuming all items in NFI are relevant for stock output (Qt_S)
        # Or maybe filter by 'Natureza'?
        # For now, we sum all 'qCom' by 'CProd'
        
        if "CProd" not in df.columns or "qCom" not in df.columns:
            print("❌ Columns CProd or qCom not found in invoice file.")
            return pd.DataFrame(columns=["CODPP", "Qt_S"])

        df["CODPP"] = df["CProd"].astype(str).str.upper().str.strip()
        df["qCom"] = pd.to_numeric(df["qCom"], errors="coerce").fillna(0)
        
        # Group by Product
        df_agg = df.groupby("CODPP", as_index=False)["qCom"].sum()
        df_agg.rename(columns={"qCom": "Qt_S"}, inplace=True)
        
        print(f"✅ {len(df_agg)} products aggregated from invoices.")
        return df_agg

    except Exception as e:
        print(f"❌ Error reading invoices: {e}")
        return pd.DataFrame(columns=["CODPP", "Qt_S"])

def get_prev_month(year: int, month: int) -> tuple[int, int]:
    if month == 1:
        return year - 1, 12
    return year, month - 1


# ─────────────────────────────────────────────────────────────
# 4. Main logic
# ─────────────────────────────────────────────────────────────
def main(year=None, month=None):
    year, month = _prompt_year_month(year, month)
    ano_mes = (year % 100) * 100 + month
    print(f"▶ Target AnoMes = {ano_mes}")

    base_dir = resolve_base_dir()
    tables_dir = os.path.join(base_dir, "Tables")
    file_path = os.path.join(tables_dir, "T_Entradas.xlsx")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"❌ File not found: {file_path}")

    # Ask user whether to overwrite original
    # For automation, we default to overwriting or use a flag. 
    # Here we will default to overwriting if running non-interactively, or we can just set out_path = file_path
    out_path = file_path
    print(f"💾 Will overwrite: {out_path}")

    # ─────────────────────────────────────────────
    # Load and process base table
    # ─────────────────────────────────────────────
    df = load_entrada_df(file_path)
    df_keep = df[df["AnoMes"] == ano_mes][["Pai", "CU_E", "CU_I", "Qt_I", "Qt_S"]].copy()
    df_keep.reset_index(inplace=True)  # inclui o número da linha original como coluna

    print(f"\n🔍 Pai values in T_Entradas for AnoMes BBB {ano_mes}:")
    print(df_keep)

    # Apply previous-month values (Qt_I and CU_I)
    df = apply_prev_month_values(df_keep, base_dir, year, month)
    #df.loc[df["AnoMes"] == ano_mes, "CU_I"] = df.loc[df["AnoMes"] == ano_mes, "CU_I_new"]
    df_keep = df[["index", "Pai", "CU_E", "Qt_I", "CU_I_new"]].copy()
    df_keep = df_keep.rename(columns={"CU_I_new": "CU_I"})
    print(f"\n🔍 Pai values in T_Entradas for AnoMes CCC {ano_mes}:")
    print(df_keep)

    # ─────────────────────────────────────────────
    # Calculate Qt_S from invoices
    # ─────────────────────────────────────────────
    prods_com_saida = calculate_qtsp_from_invoices(base_dir, year, month)

    print("\n Produtos com saidas: DDD")
    print(prods_com_saida)

    # ─────────────────────────────────────────────
    # Merge Qt_S into df (fill with 0 if missing)
    # ─────────────────────────────────────────────
    df = df_keep.merge(prods_com_saida, left_on="Pai", right_on="CODPP", how="left").fillna(0)
    print(f"\n🔍 Pai values in T_Entradas for AnoMes EEE {ano_mes}:")
    print(df)

    # 🔄 Preparar Series com index = linha no Excel
    cu_i_series = df.set_index("index")["CU_I"].round(3)
    qt_i_series = df.set_index("index")["Qt_I"].round(3)
    qt_s_series = df.set_index("index")["Qt_S"].round(3)

    # ─────────────────────────────────────────────
    # Write values to Excel
    # ─────────────────────────────────────────────
    # 📝 Gravação no Excel

    written_cui = write_column_by_index_to_excel(
        df, excel_path=file_path, out_path=out_path,
        column_name="CU_I",
        values_series=cu_i_series
    )
    print(f"✅ Wrote {written_cui} CU_I values for AnoMes {ano_mes}")

    written_qtip = write_column_by_index_to_excel(
        df, excel_path=out_path, out_path=out_path,
        column_name="Qt_I",
        values_series=qt_i_series
    )
    print(f"✅ Wrote {written_qtip} Qt_I values for AnoMes {ano_mes}")

    written_qtsp = write_column_by_index_to_excel(
        df, excel_path=out_path, out_path=out_path,
        column_name="Qt_S",
        values_series=qt_s_series
    )
    print(f"✅ Wrote {written_qtsp} Qt_S values for AnoMes {ano_mes}")

    print(f"💾 Saved to: {out_path}")
# ─────────────────────────────────────────────────────────────
# Entry Point
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    main()
