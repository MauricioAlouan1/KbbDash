"""
Process data in excel files from RAW folder to CLEAN folder.
Only process data that has ot been processed yet.
To re-process, delete file from CLEAN Folder.
"""

import re
import os
import openpyxl
import pandas as pd
import numpy as np
import chardet

# Define the potential base directories
path_options = [
    '/Users/mauricioalouan/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data/',
    '/Users/simon/Library/CloudStorage/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data'
]
# Iterate over the list and set base_dir to the first existing path
for path in path_options:
    if os.path.exists(path):
        base_dir = path
        break
else:
    # If no valid path is found, raise an error or handle it appropriately
    print("None of the specified directories exist.")
    base_dir = None  # Or set a default path if appropriate
#print("Base directory set to:", base_dir)

def _strip_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Trim trailing/leading spaces from all column names."""
    df.columns = df.columns.map(lambda c: c.strip() if isinstance(c, str) else c)
    return df

def _strip_code_values(df: pd.DataFrame) -> pd.DataFrame:
    """
    Trim spaces from common 'code' columns used nos merges (estoque/transfer/ML etc.).
    Não altera dtypes além de garantir string para aplicar strip().
    """
    candidate_cols = [
        "Código", "Codigo", "Código do Produto", "Codigo do Produto",
        "CodProd", "SKU"
    ]
    for col in df.columns:
        if col in candidate_cols:
            df[col] = df[col].astype(str).str.strip()
    return df

def _normalize_basic(df: pd.DataFrame) -> pd.DataFrame:
    """Aplicar normalização básica (nomes + códigos)."""
    df = _strip_column_names(df)
    df = _strip_code_values(df)
    return df

def find_header_row(filepath, header_name):
    """Utility function to find the header row index using pandas."""
    for i, row in pd.read_excel(filepath, header=None).iterrows():
        if header_name in row.values:
            return i
    raise ValueError(f"Header {header_name} not found in the file.")

def process_O_NFCI(data):
    """Inspect and process O_NFCI files: Remove rows where 'Situação' is effectively blank."""
    if not data.empty:
        # Print unique values in 'Situação' to inspect what's being considered as blank
        ### print("Unique values in 'Situação':", data['Situação'].dropna().unique())
        # Remove rows where 'Situação' appears to be blank or any unexpected content
        data = data[data['Situação'].apply(lambda x: x not in [None, '', ' ', np.nan, np.float64])]
    
    # Remove specified columns if they exist
    cols_to_remove = ["Projeto", "CMC Unitário do Movimento", "CMC Total do Movimento"]
    for col in cols_to_remove:
        if col in data.columns:
            data = data.drop(columns=[col])
    
    # Rename columns to standardized names
    rename_map = {
        "Data de Emissão (completa)" : "Data",
        "Código do Produto": "CODPF",
        "Quantidade": "Qt",
        "Preco Calc": "PMerc_U",
        "Total de Mercadoria": "PMerc_T",
        "Valor do ICMS ST": "A_ICMSST_T",
        "Valor do IPI": "A_IPI_T",
        "Total da Nota Fiscal": "PNF_T",
        "Valor do ICMS": "ICMS_T"
    }
    # Only rename columns that exist
    existing_rename_map = {old: new for old, new in rename_map.items() if old in data.columns}
    if existing_rename_map:
        data = data.rename(columns=existing_rename_map)
    
    # AnoMes should already be added by load_and_clean_data, but ensure it exists
    # (The value is set in load_and_clean_data for processors in the auto-add list)
    if 'AnoMes' not in data.columns:
        # Fallback: if somehow AnoMes wasn't added, we can't add it here without filepath
        # This should not happen if process_O_NFCI is in the auto-add list
        pass
    
    return data

def process_O_CC(data):
    """Process O_CC files: additional specific requirements here, if any."""
    # Assuming 'Valor (R$)' is in column F and needs to be non-zero
    data['Valor (R$)'] = pd.to_numeric(data['Valor (R$)'], errors='coerce')  # Ensure numeric
    data = data[data['Valor (R$)'] != 0]  # Remove rows where 'Valor (R$)' is zero
    data = data[data['Situação'] != "Atrasado"]  # Remove rows where 'Situaçao' = 'Atrasado'
    return data

def process_O_CtasAPagar(data):
    """Process O_CtasAPagar files: remove the row immediately below the headers."""
    # Remove the first row below the headers
    if not data.empty:
        data = data.iloc[1:]  # Remove the first row, which could be totals or sub-headers
    return data

def process_O_Estoq(data):
    """Process O_Estoq files: adapt this function to meet specific requirements."""
    # Example: Remove rows where 'Código do Produto' is empty
    data = data[data['Código do Produto'].notna()]
    return data

def process_B_Estoq(data):
    """Process B_Estoq files safely regardless of xlsx or xls format."""
    if not data.empty:

        col = 'Quantidade'

        # Se já for numérico (xls geralmente é), não mexer
        if not pd.api.types.is_numeric_dtype(data[col]):

            # Só aplica substituições se for string
            data[col] = (
                data[col]
                .astype(str)
                .str.replace(r'\.', '', regex=True)
                .str.replace(',', '.', regex=True)
            )

        # Agora converte para float de forma segura
        data[col] = pd.to_numeric(data[col], errors="coerce")

        # Remove quantidade = 0
        data = data[data[col] != 0]

        # Remove last row
        data = data.iloc[:-1]

    return data


def process_T_EstTrans(data):
    """Process O_Estoq files: adapt this function to meet specific requirements."""
    # Example: Remove rows where 'Código do Produto' is empty
    data = data[data['CodProd'].notna()]
    return data
    
def process_L_LPI(data):
    data = data[data["Data"].notna()].copy()
    
    cols_to_delete = ['Preço', 'Preço Total', 'Desconto Pedido', 'Desconto Item', 
                      'Desconto Total', 'Desconto Item Seller', 'Comissão', 'Frete Comprador', 
                      'Acrescimo', 'Recebimento', 'Custo', 'Custo Total', 'Imposto', 
                      'Lucro Bruto', 'Margem de Lucro']  # Make sure the column names match exactly
    data.drop(columns=[col for col in cols_to_delete if col in data.columns], inplace=True)

    """Process L_LPI files: convert formatted currency in specific columns to float."""
    currency_columns = ['Preço Com Desconto', 'Desconto Pedido Seller',
                        'Frete Seller']  # Update if more columns are involved
    for col in currency_columns:
        if col in data.columns:
            data[col] = data[col].apply(convert_currency_to_float)
    
    # Rename columns to standardized names
    rename_map = {
        "Integração": "Integracao",
        "Preço Com Desconto": "PMerc_T",
        "Código Pedido": "CodPed",
        "Status Pedido": "Status",
        "SKU": "CODPF",
        "Vendas": "Qt"
    }
    # Only rename columns that exist
    existing_rename_map = {old: new for old, new in rename_map.items() if old in data.columns}
    if existing_rename_map:
        data = data.rename(columns=existing_rename_map)
    
    # Add calculated column: PMerc_U = PMerc_T/Qt
    if "PMerc_T" in data.columns and "Qt" in data.columns:
        # Ensure numeric types
        data["PMerc_T"] = pd.to_numeric(data["PMerc_T"], errors="coerce")
        data["Qt"] = pd.to_numeric(data["Qt"], errors="coerce")
        # Calculate PMerc_U, handling division by zero
        data["PMerc_U"] = data["PMerc_T"] / data["Qt"].replace(0, pd.NA)
    
    return data

def detect_encoding_and_delimiter(file_path):
    """Detect the file encoding and delimiter automatically."""
    with open(file_path, "rb") as f:
        result = chardet.detect(f.read(50000))  # Analyze first 50,000 bytes
    encoding = result["encoding"]

    # Try detecting the delimiter
    with open(file_path, "r", encoding=encoding) as f:
        first_line = f.readline()
        if "," in first_line:
            delimiter = ","
        elif ";" in first_line:
            delimiter = ";"
        elif "\t" in first_line:
            delimiter = "\t"
        else:
            delimiter = ","

    print(f"✅ Detected encoding: {encoding}, Delimiter: '{delimiter}'")
    return encoding, delimiter

def process_KON_RelGeral(data):
    """
    Pré-processamento do Kon_RelGeral:
    - Normalização
    - Explode de OBS_LANCAMENTO (antes do merge)
    - Linhas SALDO e DIFERENÇA
    - Merge com T_KonCats no final
    """

    # -------- 1) NORMALIZAÇÃO BÁSICA ----------
    data = _normalize_basic(data).copy()

    # Datas → datetime
    for dc in ['DATA_PEDIDO','DATA_NF','DATA_PREVISTA','DATA_REPASSE']:
        if dc in data.columns:
            data[dc] = pd.to_datetime(data[dc], errors='coerce')

    # Campos numéricos
    for mc in ['VALOR_PREVISTO','VALOR_REPASSE','DIFERENCA']:
        if mc in data.columns:
            data[mc] = pd.to_numeric(data[mc], errors='coerce').fillna(0.0)

    # SKU normalize
    if 'SKU' in data.columns:
        data['SKU'] = data['SKU'].astype(str).str.strip()
        data['SKU'] = data['SKU'].replace({'nan': pd.NA, '': pd.NA})
        data['HasSKU'] = data['SKU'].notna()
    else:
        data['HasSKU'] = False

    # OrderKey
    if 'REF_PEDIDO' in data.columns:
        data['OrderKey'] = data['REF_PEDIDO'].astype(str).str.strip()
    else:
        data['OrderKey'] = pd.NA

    # AnoMes (já vem do load_and_clean_data)
    if 'AnoMes' not in data.columns:
        data['AnoMes'] = extract_month_year_from_filename("Kon_RelGeral_XXXX_YY.xlsx")


    # --------------------------------------------------------------------
    # -------- 2) EXPLODE OBS_LANCAMENTO (ANTES DO MERGE) ---------------
    # --------------------------------------------------------------------
    if {"CANAL", "TP_LANCAMENTO", "OBS_LANCAMENTO", "VALOR_REPASSE"}.issubset(data.columns):

        # Só Mercado Livre + DESPESAS DE SERVIÇOS
        mask = (
            data["CANAL"].astype(str).str.upper().eq("MERCADO LIVRE")
            & data["TP_LANCAMENTO"].astype(str).str.upper().eq("DESPESAS DE SERVIÇOS")
        )
        df_target = data[mask].copy()

        import re

        def extract_value(txt: str):
            """Pega o número depois de R$ no texto."""
            m = re.search(r"R\$?\s*([\d\.\,]+)", str(txt))
            if not m:
                return None
            num = m.group(1).replace(".", "").replace(",", ".")
            try:
                return float(num)
            except Exception:
                return None

        def clean_category(txt: str):
            """
            Remove 'despesas referentes a' e o valor,
            deixando só algo como:
              'CAMPANHAS DE PUBLICIDADE - PRODUCT ADS'
              'TARIFA PELO SERVIÇO DE ARMAZENAMENTO FULL'
              'TARIFA DE MANUTENÇÃO DA MINHA PÁGINA'
            """
            txt = str(txt)
            txt = re.sub(r"(?i)despesas referentes a", "", txt)
            txt = re.sub(r"R\$?\s*[\d\.\,]+", "", txt)  # tira o valor
            txt = re.sub(r"\s{2,}", " ", txt)
            return txt.strip(" -")

        new_rows = []
        diff_rows = []

        for idx, row in df_target.iterrows():
            original_total = float(row.get("VALOR_REPASSE", 0) or 0)

            # Se não tiver valor, pula
            if original_total == 0:
                continue

            # Sinal do lançamento (normalmente negativo)
            sign = -1.0 if original_total < 0 else 1.0

            texto = str(row["OBS_LANCAMENTO"] or "")
            parts = [p.strip() for p in texto.split(" + ") if p.strip()]

            soma_partes = 0.0

            for part in parts:
                valor = extract_value(part)
                if valor is None:
                    continue

                categoria = clean_category(part)
                valor_assinado = sign * valor  # mantém o mesmo sinal do total

                soma_partes += valor_assinado

                nrow = row.copy()
                # 🔹 Aqui vai o valor específico daquela parte
                nrow["VALOR_REPASSE"] = valor_assinado
                # 🔹 Categoria nova vai em TP_LANCAMENTO (para merge depois)
                nrow["TP_LANCAMENTO"] = categoria
                # Pode deixar o OBS igual à categoria também
                nrow["OBS_LANCAMENTO"] = categoria

                new_rows.append(nrow)

            # Linha original vira só um SALDO "zerado" (não entra na soma)
            data.loc[idx, "TP_LANCAMENTO"] = "Valor Explodido"

            # Diferença entre o total original e a soma das partes
            diff = original_total - soma_partes
            if abs(diff) > 0.02:
                adj = row.copy()
                adj["VALOR_REPASSE"] = diff
                # 🔹 Categoria da diferença também entra em TP_LANCAMENTO
                adj["TP_LANCAMENTO"] = "AJUSTE AUTOMATICO"
                adj["OBS_LANCAMENTO"] = "DIFERENCA DE SOMA"
                diff_rows.append(adj)

        # Inserir novas linhas explodidas
        if new_rows:
            data = pd.concat([data, pd.DataFrame(new_rows)], ignore_index=True)

        # Inserir linhas de ajuste (quando a soma não bater)
        if diff_rows:
            data = pd.concat([data, pd.DataFrame(diff_rows)], ignore_index=True)



    # --------------------------------------------------------------------
    # -------- 3) MERGE COM T_KonCats  (DEPOIS DO EXPLODE) --------------
    # --------------------------------------------------------------------
    try:
        tpath = os.path.join(base_dir, 'Tables', 'T_KonCats.xlsx')
        if os.path.exists(tpath):
            tcat = pd.read_excel(tpath)
            tcat = _normalize_basic(tcat)

            if 'TP_LANCAMENTO' in data.columns and 'TP_Lancamento' in tcat.columns:
                data = data.merge(
                    tcat[['TP_Lancamento','Kon_Gr','Kon_SGr']],
                    left_on='TP_LANCAMENTO',
                    right_on='TP_Lancamento',
                    how='left'
                )
                data.drop(columns=['TP_Lancamento'], inplace=True)

            for col in ['Kon_Gr','Kon_SGr']:
                if col in data.columns:
                    data[col] = data[col].fillna('ZZZ')

    except Exception as e:
        print(f"⚠️ Could not attach T_KonCats.xlsx: {e}")

    return data


def process_MGK_Pacotes_CSV(file_path):
    """Process MGK_Pacotes CSV files by handling encoding, delimiter, and data formatting."""

    # Detect encoding and delimiter
    encoding, delimiter = detect_encoding_and_delimiter(file_path)

    # Load CSV with detected encoding and delimiter
    data = pd.read_csv(file_path, encoding=encoding, delimiter=delimiter)

    # Convert date column
    if "Data do Pacote" in data.columns:
        data["Data do Pacote"] = pd.to_datetime(data["Data do Pacote"], errors="coerce", dayfirst=True)

    # ✅ Expand list of numeric fields
    numeric_columns = [
        "Valor total do Pacote", "Valor total (Forma de Pagamento 1)", "Valor total (Forma de Pagamento 2)",
        "Valor total dos Produtos do pacote", "Desconto totais do pacote", "Frete total do pacote"
    ]
    
    # ✅ Ensure correct numeric formatting for all specified columns
    for col in numeric_columns:
        if col in data.columns:
            data[col] = data[col].replace(r"[R$\s]", "", regex=True).replace(",", ".", regex=True)
            data[col] = pd.to_numeric(data[col], errors="coerce")  # Convert to float

    # Convert any other numeric-looking columns
    for col in data.columns:
        if data[col].dtype == "object":  # If column is stored as text
            try:
                data[col] = pd.to_numeric(data[col], errors="ignore")  # Convert if possible
            except:
                pass  # Ignore if not convertible

    # Clean text columns
    text_columns = ["Número do pedido", "Número do pacote", "Status pacote no momento que o relatório foi solicitado",
                    "Forma de pagamento 1", "Nome do cliente", "CPF/CNPJ do Cliente", "Cidade", "Estado"]
    for col in text_columns:
        if col in data.columns:
            data[col] = data[col].astype(str).str.strip()

    print("✅ MGK_Pacotes CSV processing completed with all numeric fields correctly formatted.")
    return data

def process_MLK_ExtLib(data: pd.DataFrame) -> pd.DataFrame:
    """Process MLK_ExtLib files (CSV or XLSX) into a standardized format for downstream merges."""

    # --- Normalize column names ---
    rename_map = {
        "RELEASE_DATE": "DATE",   # some exports use RELEASE_DATE
        # add other known variants here if needed
    }
    data.rename(columns=rename_map, inplace=True)

    # ✅ Exclude unwanted rows
    if "RECORD_TYPE" in data.columns:
        data = data[~data["RECORD_TYPE"].isin(["initial_available_balance", "total"])].copy()

    # ✅ Convert DATE
    if "DATE" in data.columns:
        data["DATE"] = pd.to_datetime(data["DATE"], errors="coerce", dayfirst=True)
        data["DATE"] = data["DATE"].dt.tz_localize(None)

    # ✅ Numeric columns
    numeric_columns = [
        "NET_CREDIT_AMOUNT", "NET_DEBIT_AMOUNT", "GROSS_AMOUNT",
        "SELLER_AMOUNT", "MP_FEE_AMOUNT", "FINANCING_FEE_AMOUNT", "SHIPPING_FEE_AMOUNT", "TAXES_AMOUNT",
        "COUPON_AMOUNT", "TAX_AMOUNT_TELCO", "EFFECTIVE_COUPON_AMOUNT"
    ]
    for col in numeric_columns:
        if col in data.columns:
            data[col] = data[col].replace(r"[R$\s]", "", regex=True).replace(",", ".", regex=True)
            data[col] = pd.to_numeric(data[col], errors="coerce")

    # ✅ Force ID fields as strings
    string_columns = ["ORDER_ID", "TRANSACTION_ID", "REFERENCE_NUMBER"]
    for col in string_columns:
        if col in data.columns:
            data[col] = data[col].astype(str).str.strip()

    # ✅ Add computed fields
    if {"NET_CREDIT_AMOUNT", "NET_DEBIT_AMOUNT"}.issubset(data.columns):
        data["NETVALUE"] = data["NET_CREDIT_AMOUNT"] - data["NET_DEBIT_AMOUNT"]
    if "DESCRIPTION" in data.columns:
        data["DESC"] = data["DESCRIPTION"].apply(lambda x: str(x).split("_")[0] if "_" in str(x) else x)

    return data

def process_MLK_ExtLib_CSV(file_path):
    """Process MLK_ExtLib CSV files while preserving long numeric columns, filtering rows, and adding new columns."""
    
    # Detect encoding and delimiter
    encoding, delimiter = detect_encoding_and_delimiter(file_path)

    # Fallback to UTF-8 if ASCII is detected (to avoid errors)
    if encoding and encoding.lower() == "ascii":
        encoding = "utf-8"
    elif encoding is None:
        encoding = "utf-8"  # Fallback seguro

    # Define columns that must be treated as strings to avoid digit loss
    string_columns = ["ORDER_ID", "TRANSACTION_ID", "REFERENCE_NUMBER"]  # Add more if needed

    # Load CSV with proper dtypes
    data = pd.read_csv(file_path, encoding=encoding, delimiter=delimiter, dtype={col: str for col in string_columns})

    # ✅ Exclude unwanted rows where RECORD_TYPE is "initial_available_balance" or "total"
    if "RECORD_TYPE" in data.columns:
        data = data[~data["RECORD_TYPE"].isin(["initial_available_balance", "total"])]

    # ✅ Convert date column & ensure timezone-unaware timestamps
    if "DATE" in data.columns:
        data["DATE"] = pd.to_datetime(data["DATE"], errors="coerce", dayfirst=True)
        data["DATE"] = data["DATE"].dt.tz_localize(None)  # Remove timezone

    # ✅ Expand list of numeric fields
    numeric_columns = [
        "NET_CREDIT_AMOUNT", "NET_DEBIT_AMOUNT", "GROSS_AMOUNT",
        "SELLER_AMOUNT", "MP_FEE_AMOUNT", "FINANCING_FEE_AMOUNT", "SHIPPING_FEE_AMOUNT", "TAXES_AMOUNT",
        "COUPON_AMOUNT", "TAX_AMOUNT_TELCO", "EFFECTIVE_COUPON_AMOUNT"
    ]

    # ✅ Ensure correct numeric formatting for all specified columns
    for col in numeric_columns:
        if col in data.columns:
            data[col] = data[col].replace(r"[R$\s]", "", regex=True).replace(",", ".", regex=True)
            data[col] = pd.to_numeric(data[col], errors="coerce")  # Convert to float

    # ✅ Ensure ORDER_ID and other long numeric IDs remain strings
    for col in string_columns:
        if col in data.columns:
            data[col] = data[col].astype(str)  # Ensure stored as string

    # ✅ Add new column: NETVALUE = NET_CREDIT_AMOUNT - NET_DEBIT_AMOUNT
    if "NET_CREDIT_AMOUNT" in data.columns and "NET_DEBIT_AMOUNT" in data.columns:
        data["NETVALUE"] = data["NET_CREDIT_AMOUNT"] - data["NET_DEBIT_AMOUNT"]
    
    # ✅ Add new column: DESC (Extracts before first "_" if present)
    if "DESCRIPTION" in data.columns:
        data["DESC"] = data["DESCRIPTION"].apply(lambda x: x.split("_")[0] if "_" in str(x) else x)

    print("✅ MLK_ExtLib CSV processing completed successfully. Unwanted rows removed. New columns added.")
    return data

def process_MGK_Extrato(data):
    """Process MGK_Extrato files by removing the last row (totals) while keeping all formatting."""
    
    # Check if there is an empty row before the totals
    empty_row_index = data.index[data.isnull().all(axis=1)]
    if len(empty_row_index) > 0:
        last_data_row = empty_row_index[0]  # First empty row index
        data = data.iloc[:last_data_row]  # Keep everything before the empty row

    print("✅ MGK_Extrato processing completed. Totals row removed.")
    return data

def process_SHK_Extrato(data):
    """Process SHK_Extrato files: adapt this function to meet specific requirements."""
    # Example: Remove rows where 'Código do Produto' is empty
    data = data[data['Data'].notna()]
    return data

def process_MLK_Vendas(data):
    """Process MLK_Vendas files."""
    # Example processing: remove rows where 'N.º de venda' is NaN
    data = process_ml_data(data)
    data = simplify_status(data)
    #data = data[data['N.º de venda'].notna()]
    return data

def rename_repeated_columns(df):
    """Rename repeated columns by appending a number to each repeated column name."""
    columns = df.columns
    new_columns = []
    counts = {}

    for col in columns:
        if col in counts:
            counts[col] += 1
            new_col = f"{col}{counts[col]:02d}"
        else:
            counts[col] = 0
            new_col = col
        
        new_columns.append(new_col)
    
    df.columns = new_columns
    return df

def propagate_package_info(df): 
    # Define the columns to propagate
    columns_to_propagate = [
        'Mês de faturamento das suas tarifas', 'NF-e em anexo', 'Dados pessoais ou da empresa', 'Tipo e número do documento',
        'Endereço', 'Comprador', 'CPF', 'Endereço', 'Cidade', 'Status', 'CEP', 'País',
        'Forma de entrega', 'Data a caminho', 'Data de entrega', 'Motorista', 'Número de rastreamento'
    ]

    # Identify package rows (rows where SKU is NaN)
    package_rows = df[df['SKU'].isna()]

    for idx, package_row in package_rows.iterrows():
        # Get the order ID
        order_id = package_row['N.º de venda_hyperlink']
        
        # Get the SKU rows for this package
        sku_rows = df[(df['N.º de venda_hyperlink'] == order_id) & df['SKU'].notna()]
        
        for col in columns_to_propagate:
            df.loc[sku_rows.index, col] = package_row[col]
    
    return df

def check_columns_and_rename(df, required_dict):
    """
    Verifica se pelo menos uma das colunas alternativas existe para cada campo requerido
    e renomeia para o nome padrão.

    Lança um ValueError se alguma coluna essencial estiver ausente.
    """
    missing = []

    for standard_name, options in required_dict.items():
        found = False
        for opt in options:
            if opt in df.columns:
                df.rename(columns={opt: standard_name}, inplace=True)
                found = True
                break
        if not found:
            missing.append(f"{standard_name} (expected one of: {options})")

    if missing:
        raise ValueError(f"Dataframe does not contain required columns: {missing}")
    
def process_ml_data(df):
    # Ensure the required columns exist before processing
    required_columns = ['N.º de venda', 'SKU', 'Receita por produtos (BRL)', 'Receita por envio (BRL)', 'Tarifa de venda e impostos', 'Tarifas de envio', 'Cancelamentos e reembolsos (BRL)']
    required_columns_alternatives = {
    'N.º de venda': ['N.º de venda'],
    'SKU': ['SKU'],
    'Receita por produtos (BRL)': ['Receita por produtos (BRL)'],
    'Receita por envio (BRL)': ['Receita por envio (BRL)'],
    'Tarifa de venda e impostos': ['Tarifa de venda e impostos', 'Tarifa de venda e impostos (BRL)'],
    'Tarifas de envio': ['Tarifas de envio', 'Tarifas de envio (BRL)'],
    'Cancelamentos e reembolsos (BRL)': ['Cancelamentos e reembolsos (BRL)']}

    check_columns_and_rename(df, required_columns_alternatives)

    if not all(col in df.columns for col in required_columns):
        raise ValueError("Dataframe does not contain all required columns.")

    # Strip any whitespace from column names
    df.columns = df.columns.str.strip()

    # Rename repeated columns
    df = rename_repeated_columns(df)

    df.rename(columns={"Unidades": "Quantidade"}, inplace = True)
    df.rename(columns={"Data de entrega01": "Data de devolucao"}, inplace = True)

    # Convert to numeric, coerce errors to NaN, and then fill NaN with 0
    print ('Convert to numeric')
    df['Quantidade'] = pd.to_numeric(df['Quantidade'], errors='coerce').fillna(0)
    df['Preço unitário de venda do anúncio (BRL)'] = pd.to_numeric(df['Preço unitário de venda do anúncio (BRL)'], errors='coerce').fillna(0)
    df['Receita por envio (BRL)'] = pd.to_numeric(df['Receita por envio (BRL)'], errors='coerce').fillna(0)
    df['Tarifa de venda e impostos'] = pd.to_numeric(df['Tarifa de venda e impostos'], errors='coerce').fillna(0)
    df['Tarifas de envio'] = pd.to_numeric(df['Tarifas de envio'], errors='coerce').fillna(0)
    df['Cancelamentos e reembolsos (BRL)'] = pd.to_numeric(df['Cancelamentos e reembolsos (BRL)'], errors='coerce').fillna(0)
    df['Total (BRL)'] = pd.to_numeric(df['Total (BRL)'], errors='coerce').fillna(0)

    # Step 1: Calculate the number of unique SKUs per order (excluding NaN SKUs)
    # Adjust the SKUs in Order count if it's greater than 1
    df['QtdSKUsPac'] = df[df['SKU'].notna()].groupby('N.º de venda_hyperlink')['SKU'].transform('nunique')
    df['QtdSKUsPac'] = df['QtdSKUsPac'].apply(lambda x: x-1 if x > 1 else x)

    # Step 2: Calculate the total number of items per order
    df['QtdItensPac'] = df.groupby('N.º de venda_hyperlink')['Quantidade'].transform('sum')

    # Calculate VlrTotalpSKU
    df['VlrTotalpSKU'] = df['Preço unitário de venda do anúncio (BRL)'] * df['Quantidade']

    # Calculate VlrTotalpPac
    print ('Calcula totais')
    #print(df['ReceitaEnvioTotPac'].head())
    df['VlrTotalpPac'] = df.groupby('N.º de venda_hyperlink')['VlrTotalpSKU'].transform('sum')
    df['ReceitaEnvioTotPac'] = df.groupby('N.º de venda_hyperlink')['Receita por envio (BRL)'].transform('sum')
    df['TarifaVendaTotPac'] = df.groupby('N.º de venda_hyperlink')['Tarifa de venda e impostos'].transform('sum')
    df['TarifaEnvioTotPac'] = df.groupby('N.º de venda_hyperlink')['Tarifas de envio'].transform('sum')
    df['CancelamentosTotPac'] = df.groupby('N.º de venda_hyperlink')['Cancelamentos e reembolsos (BRL)'].transform('sum')
    df['RepasseTotPac'] = df.groupby('N.º de venda_hyperlink')['Total (BRL)'].transform('sum')

    # Calculate proportional values
    print ('Calcula Valores Proporcionais')
    #print(df['ReceitaEnvioTotPac'].head())

    df['ReceitaEnvio'] = df['ReceitaEnvioTotPac'] * (df['VlrTotalpSKU'] / df['VlrTotalpPac'])
    df['TarifaVenda'] = df['TarifaVendaTotPac'] * (df['VlrTotalpSKU'] / df['VlrTotalpPac'])
    df['TarifaEnvio'] = df['TarifaEnvioTotPac'] * (df['VlrTotalpSKU'] / df['VlrTotalpPac'])
    df['Cancelamentos'] = df['CancelamentosTotPac'] * (df['VlrTotalpSKU'] / df['VlrTotalpPac'])
    df['Repasse'] = df['RepasseTotPac'] * (df['VlrTotalpSKU'] / df['VlrTotalpPac'])
    
    # Propagate package information to SKU rows and Keep only the SKU rows
    df['SKU'] = df['SKU'].str.strip()
    df['SKU'] = df['SKU'].replace('', pd.NA)
    df = propagate_package_info(df)
    df = df.dropna(subset=['SKU'])
    
    # Drop the calculation columns
    cols_to_drop = ['VlrTotalpPac', 'ReceitaEnvioTotPac', 'TarifaVendaTotPac', 'TarifaEnvioTotPac', 'CancelamentosTotPac', 'RepasseTotPac']
    cols_to_drop.extend(['Receita por produtos (BRL)', 'Receita por envio (BRL)', 'Tarifa de venda e impostos',	'Tarifas de envio',	'Cancelamentos e reembolsos (BRL)',	'Total (BRL)'])
    cols_to_drop.extend(['Unidades01', 'Unidades02', 'URL de acompanhamento', 'URL de acompanhamento01', 'Número de rastreamento', 'País', 'Tipo de contribuinte',	'Inscrição estadual'])
    cols_to_drop.extend(['Forma de entrega01', 'Data a caminho', 'Data a caminho01', 'Motorista', 'Motorista01'])
    df = df.drop([x for x in cols_to_drop if x in df.columns], axis=1)

    #df.drop(columns=['VlrTotalpPac', 'ReceitaEnvioTotPac', 'TarifaVendaTotPac', 'TarifaEnvioTotPac', 'CancelamentosTotPac', 'RepasseTotPac'])
    #df.drop(columns=['Unidades', 'URL de acompanhamento', 'Número de rastreamento', 'xx'], errors = 'ignore')
    return df

def simplify_status(df):
    # Define the patterns and replacements
    replacements = {
        r"Pacote de \d+ produtos": "Pacote de produtos",
        r"Devolvido no dia [\w\s]+": "Devolvido",
        r"Entregue dia [\w\s]+": "Entregue",
        r"Para enviar no dia [\w\s]+": "Para Enviar"
    }
    
    # Apply the replacements
    for pattern, replacement in replacements.items():
        df['Status'] = df['Status'].str.replace(pattern, replacement, regex=True)
    
    return df

def excel_column_range(start, end):
    """Generate Excel column labels between start and end inclusive."""
    columns = []
    start_index = int(start, 36) - 9  # Convert letter to number (base 36 to decimal, adjusted for Excel)
    end_index = int(end, 36) - 9
    for i in range(start_index, end_index + 1):
        number = i
        col = ''
        while number > 0:
            number, remainder = divmod(number - 1, 26)
            col = chr(65 + remainder) + col
        columns.append(col)
    return columns

def load_and_clean_data(filepath, processor, header_name, extract_hyperlinks=False):
    """Load data from an Excel file, handle merged headers, optionally extract hyperlinks."""
    if extract_hyperlinks:  
        # Call a separate function dedicated to extracting hyperlinks
        data = extract_hyperlinks_data(filepath, header_name)
    else:
        # Continue with the original data loading method
        header_row_index = find_header_row(filepath, header_name)
        data = pd.read_excel(filepath, header=header_row_index)    
         # 🔧 NOVO: normalizar nomes e códigos imediatamente após o carregamento
        data = _normalize_basic(data)
    # Extract month and year from the filename and add as a new column if necessary
    if processor in [process_B_Estoq, process_O_CtasAPagar, process_O_Estoq, process_KON_RelGeral, process_O_NFCI]:
        month_year = int(extract_month_year_from_filename(filepath))
        data['AnoMes'] = month_year
    # Process the data using the specified processor function
    return processor(data)

def extract_month_year_from_filename(filename):
    """Extract month and year from the filename in the format YYMM."""
    base_name = os.path.basename(filename)
    match = re.search(r'(\d{4})_(\d{2})', base_name)
    if match:
        year = match.group(1)[-2:]  # Get the last two digits of the year
        month = match.group(2)  # Get the month
        return f"{year}{month}"
    else:
        return "Unknown"

def convert_currency_to_float(currency_str):
    """Convert currency string 'R$ 149,90' to float 149.90, handle mixed data types."""
    if pd.isna(currency_str):
        return None  # Handle missing values
    # Check if the value is already a numeric type (float or int)
    if isinstance(currency_str, (int, float)):
        return float(currency_str)  # Return as float if already a number
    # Assuming the value is a string that needs to be cleaned and converted
    try:
        # Remove 'R$', replace ',' with '.', and remove any spaces or periods used as thousands separators
        normalized_str = currency_str.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.').strip()
        return float(normalized_str)
    except ValueError:
        print(f"Conversion error with input '{currency_str}'")
        return None
    
def check_and_process_files():
    raw_dir = os.path.join(base_dir, 'raw')
    clean_dir = os.path.join(base_dir, 'clean')

    processing_map = {
        'O_NFCI': (process_O_NFCI, "Operação", False),
        'O_CC': (process_O_CC, "Situação", False),
        'O_CtasAPagar': (process_O_CtasAPagar, "Minha Empresa (Nome Fantasia)", False),
        'O_CtasARec': (process_O_CtasAPagar, "Minha Empresa (Nome Fantasia)", False),
        'B_Estoq': (process_B_Estoq, "Código ", False),
        'B_EFull': (process_B_Estoq, "Código ", False),
        'L_LPI': (process_L_LPI, "Data", False),
        'O_Estoq': (process_O_Estoq, "Código do Produto", False),
        'MLK_Vendas': (process_MLK_Vendas, "N.º de venda", True),  # Enable hyperlink extraction for MLK_Vendas
        'MLA_Vendas': (process_MLK_Vendas, "N.º de venda", True),  # New entry, same process as MLK_Vendas
        'T_EstTrans': (process_T_EstTrans, "CodProd", False),
        "MGK_Extrato": (process_MGK_Extrato, "Relatório solicitado em: Data/Hora", False),
        "SHK_Extrato": (process_SHK_Extrato, "Data", False),
        'Kon_RelGeral': (process_KON_RelGeral, "CONCILIACAO", False)
    }
    for subdir, dirs, files in os.walk(raw_dir):
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~$'):
                # Loop through each file type in the processing map
                for key, (processor, header_name, use_hyperlinks) in processing_map.items():
                    if key in file:  # Check if the file type matches the key in the map
                        raw_filepath = os.path.join(subdir, file)
                        clean_subdir = os.path.join(clean_dir, os.path.basename(subdir))
                        clean_filepath = os.path.join(clean_subdir, file.replace('.xlsx', '_clean.xlsx'))
                        
                        raw_mtime = os.path.getmtime(raw_filepath)  # Get modification time of raw file
                        clean_mtime = os.path.getmtime(clean_filepath) if os.path.exists(clean_filepath) else 0

                        if not os.path.exists(clean_filepath) or raw_mtime > clean_mtime:
                            print(f"Processing {file}...")
                            try:
                                data = load_and_clean_data(raw_filepath, processor, header_name, use_hyperlinks)
                                save_cleaned_data(data, clean_filepath)
                            except Exception as e:
                                print(f"Error processing {file}: {e}")
                        else:
                            pass
                            # print(f"Skipped {file}, already processed.")

def check_and_process_files_csv():
    """Process all CSV files from RAW folder and save as cleaned XLSX files."""
    raw_dir = os.path.join(base_dir, 'raw')
    clean_dir = os.path.join(base_dir, 'clean')

    processing_map_csv = {
        'MGK_Pacotes': process_MGK_Pacotes_CSV
    }

    for subdir, _, files in os.walk(raw_dir):
        for file in files:
            if file.endswith('.csv') and not file.startswith('~$'):  # Ignore temp files
                for key, processor in processing_map_csv.items():
                    if key in file:
                        raw_filepath = os.path.join(subdir, file)
                        clean_subdir = os.path.join(clean_dir, os.path.basename(subdir))
                        clean_filepath = os.path.join(clean_subdir, file.replace('.csv', '_clean.xlsx'))

                        raw_mtime = os.path.getmtime(raw_filepath)
                        clean_mtime = os.path.getmtime(clean_filepath) if os.path.exists(clean_filepath) else 0

                        if not os.path.exists(clean_filepath) or raw_mtime > clean_mtime:
                            print(f"📂 Processing CSV: {file}...")
                            try:
                                data = processor(raw_filepath)  # Process CSV file
                                save_cleaned_data(data, clean_filepath)  # Save as XLSX
                            except Exception as e:
                                print(f"❌ Error processing {file}: {e}")
                        else:
                            pass
                            # print(f"Skipped {file}, already processed.")

def check_and_process_files_multiformat():
    """
    Process multi-format files (that can come as .csv or .xlsx)
    and save standardized clean files.
    """
    raw_dir = os.path.join(base_dir, 'raw')
    clean_dir = os.path.join(base_dir, 'clean')

    # Map of files that may come in multiple formats
    processing_map_multi = {
        "MLK_ExtLib": process_MLK_ExtLib
    }

    for subdir, _, files in os.walk(raw_dir):
        for file in files:
            for key, processor in processing_map_multi.items():
                if key in file and (file.endswith(".csv") or file.endswith(".xlsx")) and not file.startswith("~$"):   
                    raw_filepath = os.path.join(subdir, file)
                    clean_subdir = os.path.join(clean_dir, os.path.basename(subdir))
                    clean_filepath = os.path.join(clean_subdir, 
                                                  file.replace(".csv", "_clean.xlsx").replace(".xlsx", "_clean.xlsx"))

                    raw_mtime = os.path.getmtime(raw_filepath)
                    clean_mtime = os.path.getmtime(clean_filepath) if os.path.exists(clean_filepath) else 0

                    if not os.path.exists(clean_filepath) or raw_mtime > clean_mtime:
                        print(f"📂 Processing MULTIFORMAT: {file}...")
                        try:
                            # Load depending on extension
                            if file.endswith(".csv"):
                                encoding, delimiter = detect_encoding_and_delimiter(raw_filepath)
                                # 👇 safeguard: ASCII is too limited, fallback to UTF-8
                                if encoding is None or encoding.lower() == "ascii":
                                    encoding = "utf-8"
                                df = pd.read_csv(raw_filepath, encoding=encoding, delimiter=delimiter)
                            else:  # .xlsx
                                df = pd.read_excel(raw_filepath)

                            # Standardize with processor
                            data = processor(df)
                            save_cleaned_data(data, clean_filepath)

                        except Exception as e:
                            print(f"❌ Error processing {file}: {e}")

def extract_hyperlinks_data(filepath, header_name):
    """Extract data and create a new column for hyperlinks for a specific header."""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb.active
    data_rows = []
    header_row_index = None
    headers = []

    # Iterate over rows to find the header and extract data
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, values_only=False):
        if header_row_index is None:
            if any(header_name == (cell.value or '') for cell in row):
                header_row_index = row[0].row
                headers = [cell.value for cell in row]
                headers.append(f"{header_name}_hyperlink")
                continue
        if header_row_index and row[0].row > header_row_index:
            row_data = []
            hyperlink_value = None
            for cell in row:
                if cell.column == headers.index(header_name) + 1 and cell.hyperlink:
                    # Replace specific parts of the hyperlink
                    hyperlink_replaced = cell.hyperlink.target.replace("https://www.mercadolivre.com.br/vendas/", "").replace("/detalhe#source=excel", "")
                    hyperlink_value = hyperlink_replaced
                row_data.append(cell.value)
            row_data.append(hyperlink_value)
            data_rows.append(row_data)
    # 🔧 Novo: cria o DF, aplica normalização e retorna
    df = pd.DataFrame(data_rows, columns=headers)
    df = _normalize_basic(df)
    return df

def save_cleaned_data(data, output_filepath):
    """Save the cleaned data to a new Excel file."""
    os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
    data.to_excel(output_filepath, index=False)

if __name__ == "__main__":
    check_and_process_files()
    check_and_process_files_csv()
    check_and_process_files_multiformat()
