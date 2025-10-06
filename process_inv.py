"""
This script, `process_inv.py`, is designed to handle inventory data files, primarily for the KBB MF project. 
Its primary goals are as follows:
1. Define and validate the base directory for accessing inventory data files.
2. Set up date range variables to specify the start and end periods for processing.
3. Implement functions to:
   - Process and stack inventory files for specific months and years.
   - Format data for consistent and streamlined reporting.
4. Organize files within a directory structure (e.g., `/clean/YYYY_MM/`) for efficient data retrieval and processing.

Key Features:
- Dynamic handling of inventory files based on specified year and month.
- Robust error handling to ensure smooth execution even if files or directories are missing.
- Integration with other scripts and workflows in the KBB MF project.

Prerequisites:
- Ensure that the base directory paths specified in `path_options` exist and contain the necessary inventory files.
- Verify that the `/clean/YYYY_MM/` directory structure is consistent with the expected format.

This script is integral to maintaining the accuracy and efficiency of inventory management workflows.
"""


import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook, load_workbook
import sys

# Define the base directory as before, now adding the /clean part
path_options = [
    '/Users/mauricioalouan/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data/',
    '/Users/simon/Library/CloudStorage/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data'
]
for path in path_options:
    if os.path.exists(path):
        base_dir = path
        break
else:
    print("None of the specified directories exist.")
    base_dir = None

# Define the date range variables
# ---- minimal prompt block (like Conc_Estoque.py) ----
def _prompt_year_month():
    import argparse
    from datetime import datetime
    ap = argparse.ArgumentParser(add_help=False)
    ap.add_argument("--year", "-y", type=int)
    ap.add_argument("--month", "-m", type=int)
    args, _ = ap.parse_known_args()
    if args.year is not None and args.month is not None:
        return args.year, args.month
    # If missing, ask interactively; defaults mimic Conc_Estoque.py (month defaults to previous)
    now = datetime.now()
    default_year = now.year if now.month > 1 else now.year - 1
    default_month = now.month - 1 if now.month > 1 else 12
    print("Year and/or month not provided.")
    year = int(input(f"Enter year (default {default_year}): ") or default_year)
    month = int(input(f"Enter month [1-12] (default {default_month}): ") or default_month)
    return year, month

start_year, start_month = _prompt_year_month()
end_year, end_month = start_year, start_month
# ---- end minimal prompt block ----

# Function to process inventory files for a given month and year
def process_inventory_files(year, month):
    """Process and stack inventory files for a given year and month."""         
    try:
        # Format the month to always be two digits (e.g., 01, 02, ..., 12)
        month_str = f'{month:02d}'
        
        # The files for each month are inside the /clean/YYYY_MM/ folder
        clean_folder = os.path.join(base_dir, f'clean/{year}_{month_str}')

        # Define all file types with their corresponding 'Local' values
        file_configs = {
            f'B_Estoq_{year}_{month_str}_clean.xlsx': 'Bling',
            f'T_EstTrans_{year}_{month_str}_clean.xlsx': 'Transito',
            f'O_Estoq_{year}_{month_str}_clean.xlsx': None,  # Special case with its own column
            f'B_EFullAj_{year}_{month_str}_clean.xlsx': 'Ajuste',
            f'B_EFullAm_{year}_{month_str}_clean.xlsx': 'Amazon Full',
            f'B_EFullMg_{year}_{month_str}_clean.xlsx': 'Magalu Full',
            f'B_EFullML_{year}_{month_str}_clean.xlsx': 'ML Full'
        }

        combined_dfs = []

        # Process each file
        for file_name, local_value in file_configs.items():
            file_path = os.path.join(clean_folder, file_name)
            try:
                if os.path.exists(file_path):
                    if 'O_Estoq' in file_name:
                        # Special handling for O_Estoq
                        df = pd.read_excel(file_path, usecols=['Código do Produto', 'Quantidade', 'Local de Estoque (Código)'])
                        df.rename(columns={
                              'Código do Produto': 'Codigo',
                            'Quantidade': 'Quantidade',
                            'Local de Estoque (Código)': 'Local'
                        }, inplace=True)
                    elif 'T_EstTrans' in file_name:
                        # Special handling for T_EstTrans
                        df = pd.read_excel(file_path, usecols=['CodProd', 'Qt'])
                        df.rename(columns={'CodProd': 'Codigo', 'Qt': 'Quantidade'}, inplace=True)
                        df['Local'] = 'Transito'
                    else:
                        # General handling
                        df = pd.read_excel(file_path, usecols=['Código', 'Quantidade'])
                        df.rename(columns={'Código': 'Codigo', 'Quantidade': 'Quantidade'}, inplace=True)
                        if local_value:
                            df['Local'] = local_value
                    combined_dfs.append(df)
                else:
                    print(f"File not found: {file_name}. Skipping this file.")
            except Exception as e:
                print(f"Error processing inventory files for {year}-{month_str}, file prefix: {file_name}: {e}")
                continue  # Skip this file and proceed with the next

        # Combine all dataframes
        if combined_dfs:
            combined_df = pd.concat(combined_dfs, ignore_index=True)
        else:
            print(f"No files found for {year}-{month_str}. Returning an empty DataFrame.")
            combined_df = pd.DataFrame(columns=['Codigo', 'Quantidade', 'Local'])

        return combined_df

    except Exception as e:
        print(f"Error processing inventory files for {year}-{month_str}: {e}")
        return None


def _write_cuem_in_place(xlsx_path, sheet_name, pai_to_cuem, cutoff_code):
    """
    Update only CUEm cells in-place on the given sheet, for rows with AnoMes == cutoff_code.
    Keeps all formatting, formulas, and other sheets intact.
    Requires headers: 'Pai', 'AnoMes', 'CUEm' in row 1 of the sheet.
    Returns number of rows updated.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)  # keep formulas
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {xlsx_path}")
    ws = wb[sheet_name]

    header_cells = list(ws[1])
    header_map = {str(c.value).strip(): c.col_idx for c in header_cells if c.value is not None}

    required = ['Pai', 'AnoMes', 'CUEm']
    missing = [c for c in required if c not in header_map]
    if missing:
        raise ValueError(
            f"Sheet '{sheet_name}' must contain header(s): {', '.join(missing)}."
        )

    col_pai    = header_map['Pai']
    col_anomes = header_map['AnoMes']
    col_cuem   = header_map['CUEm']

    updated = 0
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        anomes_val = ws.cell(row=r, column=col_anomes).value
        pai_val    = ws.cell(row=r, column=col_pai).value
        try:
            anomes_code = int(str(anomes_val).strip())
        except Exception:
            continue
        pai_key = (str(pai_val).strip() if pai_val is not None else '')
        if anomes_code == cutoff_code and pai_key in pai_to_cuem:
            ws.cell(row=r, column=col_cuem).value = float(pai_to_cuem[pai_key])
            updated += 1

    wb.save(xlsx_path)
    return updated

def lookup_cu_values(inventory_df, cutoff_date):
    """
    Father-only cost lookup with strict validation and robust matching.

    - Cutoff by AnoMes (YYMM).
    - Per-entry effective cost: CUF (>0) -> CUEm (>0) -> (Ult CU R$ + AddR).
    - Weighted average by (Pai, AnoMes) using Qt.
    - For the cutoff month, writes weighted cost to T_Entradas.CUEm (Sheet1) IN-PLACE.
    - Output KEEP 'Pai' and drop only helper mapping columns (CodPF_Prod, CodPP_Prod, UCP).
    - FAILS FAST if any required column is missing.
    - Prints explicit match counts so we can see what’s happening.
    """
    import numpy as np

    print("[lookup_cu_values] START")

    entradas_path = os.path.join(base_dir, 'Tables', 'T_Entradas.xlsx')
    entradas_df = pd.read_excel(
        entradas_path,
        dtype={'Pai': str, 'Filho': str}
    )
    prodf_df = pd.read_excel(
        os.path.join(base_dir, 'Tables', 'T_ProdF.xlsx'),
        dtype={'CodPF': str, 'CodPP': str}
    )

    # ---- cutoff YYMM (guard for 2000–2099)
    year = cutoff_date.year
    month = cutoff_date.month
    if not (2000 <= year <= 2099):
        raise ValueError(f"Unsupported year {year}: YYMM logic expects 2000–2099.")
    cutoff_code = (year - 2000) * 100 + month
    print(f"[lookup_cu_values] cutoff YYMM = {cutoff_code}")

    # ---- required columns (HARD FAIL)
    req_entradas = ['AnoMes', 'Qt', 'Ult CU R$', 'AddR', 'Pai', 'CUF', 'CUEm']
    missing_e = [c for c in req_entradas if c not in entradas_df.columns]
    if missing_e:
        raise KeyError(f"T_Entradas.xlsx missing required column(s): {', '.join(missing_e)}")

    req_prodf = ['CodPF', 'CodPP']
    missing_p = [c for c in req_prodf if c not in prodf_df.columns]
    if missing_p:
        raise KeyError(f"T_ProdF.xlsx missing required column(s): {', '.join(missing_p)}")

    # ---- normalize helpers
    def norm_str(x):
        return '' if pd.isna(x) else str(x).strip()

    def to_num_key(s):
        s = norm_str(s)
        return int(s) if s.isdigit() else np.nan

    # ---- sanitization
    entradas_df['AnoMes']     = pd.to_numeric(entradas_df['AnoMes'], errors='coerce').astype('Int64')
    entradas_df               = entradas_df.dropna(subset=['AnoMes']).copy()
    entradas_df['AnoMes']     = entradas_df['AnoMes'].astype(int)
    entradas_df['Qt']         = pd.to_numeric(entradas_df['Qt'], errors='coerce').fillna(0.0)
    entradas_df['Ult CU R$']  = pd.to_numeric(entradas_df['Ult CU R$'], errors='coerce').fillna(0.0)
    entradas_df['AddR']       = pd.to_numeric(entradas_df['AddR'], errors='coerce').fillna(0.0)
    entradas_df['CUF']        = pd.to_numeric(entradas_df['CUF'], errors='coerce').fillna(0.0)
    entradas_df['CUEm']       = pd.to_numeric(entradas_df['CUEm'], errors='coerce').fillna(0.0)
    entradas_df['Pai']        = entradas_df['Pai'].map(norm_str)

    pre_rows = len(entradas_df)
    entradas_df = entradas_df[entradas_df['AnoMes'] <= cutoff_code]
    print(f"[lookup_cu_values] entradas rows <= cutoff: {len(entradas_df)} (from {pre_rows})")

    # ---- father-only rows
    pai_df = entradas_df[entradas_df['Pai'] != ''].copy()
    print(f"[lookup_cu_values] pai rows <= cutoff: {len(pai_df)}")
    if pai_df.empty:
        print("[lookup_cu_values][WARN] No father entries up to cutoff; returning zeros.")
        out = inventory_df.copy()
        out['Codigo'] = out['Codigo'].astype(str)
        out = out.rename(columns={'Quantidade': 'Quantidade_Inv', 'Codigo': 'Codigo_Inv'})
        out['Pai'] = out['Codigo_Inv']
        print(f"[lookup_cu_values] PF→PP mapped for {(out['CodPP_Prod']!='').sum()}/{len(out)} SKUs")

        out['UCU'] = 0.0
        out['UCT'] = 0.0
        return out.drop(columns=['CodPF_Prod', 'CodPP_Prod', 'UCP'], errors='ignore')

    pai_df['Pai_key_str'] = pai_df['Pai']
    pai_df['Pai_key_num'] = pai_df['Pai'].apply(to_num_key)

    # ---- per-entry effective cost priority
    pai_df['EffCost'] = np.where(
        pai_df['CUF'] > 0, pai_df['CUF'],
        np.where(pai_df['CUEm'] > 0, pai_df['CUEm'], pai_df['Ult CU R$'] + pai_df['AddR'])
    )
    pai_df['CUxQ'] = pai_df['EffCost'] * pai_df['Qt']

    # ---- monthly weighted averages
    monthly_str = (
        pai_df.groupby(['Pai_key_str', 'AnoMes'], as_index=False)
              .agg({'CUxQ': 'sum', 'Qt': 'sum'})
    )
    monthly_num = (
        pai_df.dropna(subset=['Pai_key_num'])
             .groupby(['Pai_key_num', 'AnoMes'], as_index=False)
             .agg({'CUxQ': 'sum', 'Qt': 'sum'})
    )
    print(f"[lookup_cu_values] monthly keys: str={len(monthly_str)} rows, num={len(monthly_num)} rows")

    # ---- write CUEm for cutoff month (string key; sheet stores 'Pai' as text)
    monthly_cutoff = monthly_str.loc[monthly_str['AnoMes'] == cutoff_code, ['Pai_key_str', 'CUxQ', 'Qt']].copy()
    if not monthly_cutoff.empty:
        monthly_cutoff['UCU_month'] = np.where(monthly_cutoff['Qt'] > 0, monthly_cutoff['CUxQ'] / monthly_cutoff['Qt'], 0.0)
        pai_to_cuem = dict(zip(monthly_cutoff['Pai_key_str'], monthly_cutoff['UCU_month']))
        updated = _write_cuem_in_place(entradas_path, 'Sheet1', pai_to_cuem, cutoff_code)
        print(f"[CUEm] YYMM={cutoff_code}: {len(pai_to_cuem)} Pai(s) aggregated; wrote {updated} row(s) to Sheet1.")
    else:
        print(f"[CUEm] No entries in cutoff month YYMM={cutoff_code}; nothing to write.")

    # ---- latest month per Pai (<= cutoff) → UCU
    if monthly_str.empty and monthly_num.empty:
        raise RuntimeError("[lookup_cu_values] No monthly data to compute last costs.")

    if not monthly_str.empty:
        idx_last_s = monthly_str.groupby('Pai_key_str')['AnoMes'].idxmax()
        last_str = monthly_str.loc[idx_last_s, ['Pai_key_str', 'CUxQ', 'Qt']].copy()
        last_str['UCU'] = np.where(last_str['Qt'] > 0, last_str['CUxQ'] / last_str['Qt'], 0.0)
        last_str = last_str.rename(columns={'Pai_key_str': 'key_str'})[['key_str', 'UCU']]
    else:
        last_str = pd.DataFrame(columns=['key_str', 'UCU'])

    if not monthly_num.empty:
        idx_last_n = monthly_num.groupby('Pai_key_num')['AnoMes'].idxmax()
        last_num = monthly_num.loc[idx_last_n, ['Pai_key_num', 'CUxQ', 'Qt']].copy()
        last_num['UCU_num'] = np.where(last_num['Qt'] > 0, last_num['CUxQ'] / last_num['Qt'], 0.0)
        last_num = last_num.rename(columns={'Pai_key_num': 'key_num'})[['key_num', 'UCU_num']]
    else:
        last_num = pd.DataFrame(columns=['key_num', 'UCU_num'])

    # ---- prepare inventory & map PF→PP (father). Fallback: father = own code.
    out = inventory_df.copy()
    out['Codigo'] = out['Codigo'].astype(str)
    out = out.rename(columns={'Quantidade': 'Quantidade_Inv', 'Codigo': 'Codigo_Inv'})

    prodf_df = prodf_df.rename(columns={'CodPF': 'CodPF_Prod', 'CodPP': 'CodPP_Prod'})
    prodf_df['CodPF_Prod'] = prodf_df['CodPF_Prod'].map(norm_str)
    prodf_df['CodPP_Prod'] = prodf_df['CodPP_Prod'].map(norm_str)
    out['Codigo_Inv']      = out['Codigo_Inv'].map(norm_str)

    # --- pass 1: string key join
    out = pd.merge(
        out, prodf_df[['CodPF_Prod', 'CodPP_Prod']],
        left_on='Codigo_Inv', right_on='CodPF_Prod', how='left'
    )
    # rows still needing a father code
    need_map = out['CodPP_Prod'].isna() | (out['CodPP_Prod'] == '')

    # --- pass 2: numeric key fallback (handles '001234' vs '1234')
    def _to_num_key(s):
        s = '' if pd.isna(s) else str(s).strip()
        return int(s) if s.isdigit() else pd.NA

    prodf_df['CodPF_num'] = prodf_df['CodPF_Prod'].apply(_to_num_key)
    prodf_df['CodPP_num'] = prodf_df['CodPP_Prod'].apply(_to_num_key)
    out['Codigo_num']     = out['Codigo_Inv'].apply(_to_num_key)

    if need_map.any():
        out.loc[need_map, ['Codigo_num']] = out.loc[need_map, 'Codigo_Inv'].apply(_to_num_key)
        out = out.merge(
            prodf_df[['CodPF_num', 'CodPP_Prod']].rename(columns={'CodPF_num': 'join_num'}),
            left_on='Codigo_num', right_on='join_num', how='left'
        )
        # fill from numeric join where string join failed
        take_num = need_map & out['CodPP_Prod'].isna() & out['CodPP_Prod_y'].notna()
        out.loc[take_num, 'CodPP_Prod'] = out.loc[take_num, 'CodPP_Prod_y']
        out = out.drop(columns=['join_num', 'CodPP_Prod_y'], errors='ignore')

    # final fallback: father = own code
    out['CodPP_Prod'] = out['CodPP_Prod'].where(
        out['CodPP_Prod'].notna() & (out['CodPP_Prod'] != ''), out['Codigo_Inv']
    ).astype(str).str.strip()

    # report must keep Pai → set Pai to mapped father code
    out['Pai'] = out['CodPP_Prod']


    # two join keys for robustness (string + numeric)
    out['key_str'] = out['CodPP_Prod']
    out['key_num'] = out['CodPP_Prod'].apply(to_num_key)

    # pass 1: string key
    out = pd.merge(out, last_str, on='key_str', how='left')
    matched1 = int(out['UCU'].notna().sum()); total = len(out)
    print(f"[lookup_cu_values] matches after string-key merge: {matched1}/{total}")

    # pass 2: numeric-key fallback
    if not last_num.empty:
        need = out['UCU'].isna()
        if need.any():
            tmp = out.loc[need, ['key_num']].merge(last_num, on='key_num', how='left')
            out.loc[need, 'UCU'] = tmp['UCU_num'].values

    matched2 = int(out['UCU'].notna().sum())
    print(f"[lookup_cu_values] total matches after numeric-key fallback: {matched2}/{total}")

    # numerics & totals
    out['Quantidade_Inv'] = pd.to_numeric(out['Quantidade_Inv'], errors='coerce').fillna(0.0)
    out['UCU'] = pd.to_numeric(out['UCU'], errors='coerce').fillna(0.0)
    out['UCT'] = out['UCU'] * out['Quantidade_Inv']

    # cleanup (keep Pai)
    out = out.drop(columns=['CodPF_Prod', 'CodPP_Prod', 'UCP', 'key_str', 'key_num'], errors='ignore')

    # hard guard: if everything zero, say it loudly
    if float(out['UCU'].sum()) == 0.0:
        print("[lookup_cu_values][WARN] All UCU are 0. Likely a key mismatch. Check PF/PP vs Pai formats and AnoMes filter.")
    print("[lookup_cu_values] END")

    return out


# Main function to handle the process for all months within the date range
def process_all_months():
    # Loop through each year and month in the specified range
    for year in range(start_year, end_year + 1):
        for month in range(1, 13):
            if year == start_year and month < start_month:
                continue
            if year == end_year and month > end_month:
                break

            print(f"Processing data for year {year}, month {month:02d}")

            # Step 1: Process and stack inventory data for the given year and month
            inventory_df = process_inventory_files(year, month)
            if inventory_df is None:
                continue

            # Step 2: Lookup CU values and calculate UCU and UCT
            cutoff_date = pd.Timestamp(year=year, month=month, day=1)
            final_df = lookup_cu_values(inventory_df, cutoff_date)
            print(f"[pipeline] after lookup: UCU>0={(final_df['UCU']>0).sum()}/{len(final_df)} ; unique Pai={final_df['Pai'].nunique() if 'Pai' in final_df.columns else 'n/a'}")


            if final_df is None:
                continue

            # Add AnoMes
            final_df['AnoMes'] = (year % 100) * 100 + month
            # Step 3: Save the resulting dataframe to a new Excel file
            output_filepath = os.path.join(base_dir, 'clean',f'{year}_{month:02d}', f'R_Estoq_fdm_{year}_{month:02d}.xlsx')
            final_df.to_excel(output_filepath, index=False, sheet_name='Data')
            print(f"Saved combined inventory data for {year}-{month:02d} to {output_filepath}")
            format_and_add_pivot(output_filepath, final_df, year,month)
            print(f"Added Formating and Pivots for {year}-{month:02d} to {output_filepath}")

# Format and add pivot tables using openpyxl
def format_and_add_pivot(output_filepath, df, year, month):
    # Load the workbook
    wb = load_workbook(output_filepath)
    ws = wb['Data']
    
    # Apply number format to specified columns
    number_format = '#,##0.00'
    columns_to_format = ['UCP', 'UCF', 'UCU', 'UCT']
    # Add autofilter to the pivot table
    ws.auto_filter.ref = ws.dimensions

    for col in columns_to_format:
        if col in df.columns:
            col_idx = df.columns.get_loc(col) + 1  # Adjust for Excel's 1-based indexing
            for row in range(2, len(df) + 2):  # Start from the second row (excluding header)
                cell = ws.cell(row=row, column=col_idx)
                cell.number_format = number_format
    
    # Create a pivot table on a new sheet
    pivot_table = df.pivot_table(
       index='Codigo_Inv',
        columns='Local',
        values='Quantidade_Inv',
        aggfunc='sum',
        fill_value=0
    )
############################################
    # Ensure total column is correct
    pivot_table['Total'] = pivot_table.sum(axis=1)

    # Add a total cost column
    total_cost = df.groupby('Codigo_Inv')['UCT'].sum()
    pivot_table['Total Cost'] = total_cost

    # Add a unit cost column
    pivot_table['Unit Cost'] = pivot_table['Total Cost'] / pivot_table['Total']
    # Add AnoMes
    pivot_table['AnoMes'] = (year % 100) * 100 + month

    # Validate totals
    original_total_cost = df['UCT'].sum()
    pivot_total_cost = pivot_table['Total Cost'].sum()

    print(f"Original Total Cost: {original_total_cost}")
    print(f"Pivot Table Total Cost: {pivot_total_cost}")

    # Check for mismatched rows
    unmatched_rows = df[~df['Codigo_Inv'].isin(pivot_table.index)]
    if not unmatched_rows.empty:
        print("Unmatched rows found:")
        print(unmatched_rows)
    else:
        print('### No unmatched rows ###')
        
############################################

    # Reset the index for better readability
    pivot_table = pivot_table.reset_index()

    # Add the pivot table to a new sheet
    pivot_sheet_name = 'PT01'
    if pivot_sheet_name not in wb.sheetnames:
        wb.create_sheet(title=pivot_sheet_name)
    pivot_ws = wb[pivot_sheet_name]
    
    # Write headers
    for col_idx, header in enumerate(pivot_table.columns, start=1):
        pivot_ws.cell(row=1, column=col_idx, value=header)

    # Write the pivot table data
    for r_idx, row in enumerate(pivot_table.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            pivot_ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Add totals row at the bottom
    totals_row_idx = len(pivot_table) + 2
    pivot_ws.cell(row=totals_row_idx, column=1, value="Grand Total").font = Font(bold=True)
    for col_idx, col_name in enumerate(pivot_table.columns[1:], start=2):  # Skip 'CodPF_Prod'
        total_value = pivot_table[col_name].sum()
        cell = pivot_ws.cell(row=totals_row_idx, column=col_idx, value=total_value)
        cell.font = Font(bold=True)
        if col_name in ['Total Cost', 'Unit Cost']:
            cell.number_format = number_format

    # Apply formatting to the last 3 columns (bold + light gray fill)
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for col_name in ['Total', 'Total Cost', 'Unit Cost']:
        col_idx = pivot_table.columns.get_loc(col_name) + 1
        for row in range(2, totals_row_idx + 1):  # Include totals row
            cell = pivot_ws.cell(row=row, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = gray_fill
            cell.number_format = number_format

    # Add autofilter to the pivot table
    pivot_ws.auto_filter.ref = pivot_ws.dimensions

    # Save the workbook
    wb.save(output_filepath)


if __name__ == "__main__":
    process_all_months()
