import os
import pandas as pd
import shutil
from openpyxl import load_workbook

#Global
ano_x = 2025
mes_x = 1

# Format month as two digits (01, 02, ..., 12)
mes_str = f"{mes_x:02d}"
ano_mes = f"{ano_x}_{mes_str}"  # e.g., "2025_01"

# Define the potential base directories
path_options = [
    '/Users/mauricioalouan/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data/',
    '/Users/simon/Library/CloudStorage/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data'
]
# Iterate over the list and set base_dir to the first existing path
for path in path_options:
    if os.path.exists(path):
        base_dir = path
        break
else:
    # If no valid path is found, raise an error or handle it appropriately
    print("None of the specified directories exist.")
    base_dir = None  # Or set a default path if appropriate
print("Base directory set to:", base_dir)

# Define paths dynamically using global variables
source_file = os.path.join(base_dir, "clean", ano_mes, f"R_Resumo_{ano_mes}.xlsx")
template_file = os.path.join(base_dir, "Template", "PivotTemplate.xlsm")
output_file = os.path.join(base_dir, "clean", ano_mes, f"Pivot_Report_{ano_mes}.xlsm")

# Step 1: Copy the template (preserves macros)
shutil.copy(template_file, output_file)
print(f"✅ Copied template to {output_file}")

# Step 2: Open the template workbook with macros
print("✅ Opening template with macros...")
wb_template = load_workbook(output_file, keep_vba=True)

# Step 3: Open the source workbook
print("✅ Loading source file into memory...")
wb_source = load_workbook(source_file)

# Step 4: Remove all existing sheets from the template
print(f"✅ Removing {len(wb_template.sheetnames)} sheets from template...")
for sheet in wb_template.sheetnames:
    del wb_template[sheet]

# Step 5: Copy all sheets from source to template
print(f"✅ Copying {len(wb_source.sheetnames)} sheets...")
for sheet_name in wb_source.sheetnames:
    print(f"📄 Copying sheet: {sheet_name}")
    source_sheet = wb_source[sheet_name]

    # Create a new sheet in the template with the same name
    new_sheet = wb_template.create_sheet(title=sheet_name)

    # Copy data from the source sheet to the new sheet
    for row in source_sheet.iter_rows():
        for cell in row:
            new_sheet[cell.coordinate].value = cell.value

# Step 6: Save and close the updated workbook
wb_template.save(output_file)
wb_template.close()
wb_source.close()

print(f"✅ Final file saved at {output_file}")
print("📌 Open the file in Excel and run the macro: 'CreatePivotTable'.")
