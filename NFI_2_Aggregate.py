import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import xlwings as xw


# Define base folder and available series
path_options = [
    '/Users/mauricioalouan/Dropbox/nfs',
    '/Users/simon/Library/CloudStorage/Dropbox/nfs'
]
BASE_FOLDER = None
for path in path_options:
    if os.path.exists(path):
        BASE_FOLDER = path
        break

# The same series used in NFI_1_Create.py
SERIES_LIST = [
    "Serie 1 - Omie",
    "Serie 2 - filial",
    "Serie 3 - Bling",
    "Serie 4 - Lexos",
    "Serie 5 - Olist",
    "Serie 6 - Meli",
    "Serie 7 - Amazon",
    "Serie 8 - Magalu Full",
    "Serie 9 - Shopee Full"
]

import xlwings as xw

def run_macro(path, macro_name):
    app = xw.App(visible=False)
    wb = app.books.open(path)
    app.macro(macro_name)()
    wb.app.api.calculate()
    wb.save()
    wb.close()
    app.quit()

# === MAIN FUNCTION ===
def combine_monthly_items_excels(year, month):
    # Define output directory
    month_num = str(month).split('-')[0].zfill(2)
    output_dir = os.path.join(BASE_FOLDER, "Mauricio", "Contabilidade", f"{year}_{month_num}")
    OUTPUT_DIR = output_dir
    
    if not os.path.exists(output_dir):
        print(f"⚠️ Output directory not found: {output_dir}")
        return

    # Define path for lookup tables and Resumo
    # Assuming BASE_FOLDER ends with /nfs, we strip it to get the root Dropbox folder
    dropbox_root = os.path.dirname(BASE_FOLDER)
    tables_dir = os.path.join(dropbox_root, "KBB MF", "AAA", "Balancetes", "Fechamentos", "data", "Tables")
    
    lookup_file = os.path.join(tables_dir, "T_NFTipo.xlsx")
    resumo_file = os.path.join(tables_dir, "R_ResumoFin25.xlsx")
    prodf_file = os.path.join(tables_dir, "T_Prodf.xlsx")

    # Load Lookup Table
    try:
        lookup_df = pd.read_excel(lookup_file)
        # Create a dictionary for faster lookup: Natureza_NF -> Natureza_Grp
        # Ensure columns exist
        if "Natureza_NF" in lookup_df.columns and "Natureza_Grp" in lookup_df.columns:
            lookup_map = dict(zip(lookup_df["Natureza_NF"], lookup_df["Natureza_Grp"]))
        else:
            print(f"⚠️ Warning: Columns 'Natureza_NF' or 'Natureza_Grp' not found in {lookup_file}")
            lookup_map = {}
    except Exception as e:
        print(f"⚠️ Error reading lookup table: {e}")
        lookup_map = {}

    # Load Product Table
    prodf_df = pd.DataFrame()
    prodf_raw = pd.read_excel(prodf_file)
    prodf_df = prodf_raw[["CodPF", "CodPP"]].drop_duplicates(subset=["CodPF"])
    combined_df = pd.DataFrame()

    # Iterate through series
    for series in SERIES_LIST:
        # Input file from NFI_1_Create
        input_file = os.path.join(output_dir, f"NFI_{year}_{month_num}_{series}.xlsx")
        
        if not os.path.exists(input_file):
            print(f"Skipping {series}: File not found -> {input_file}")
            continue
            
        try:
            df = pd.read_excel(input_file)
            df.insert(0, "Series", series)  # add series name
            # Add Natureza_GRP
            df["Natureza_GRP"] = df["Natureza"].map(lookup_map).fillna(999)
            combined_df = pd.concat([combined_df, df], ignore_index=True)
            print(f"✔ Loaded {series} ({len(df)} rows)")
        except Exception as e:
            print(f"❌ Error reading {input_file}: {e}")

    if combined_df.empty:
        print("No data combined.")
        return

    # Merge with T_Prodf
    
    print("Merging with T_Prodf...")
    combined_df = combined_df.merge(prodf_df, left_on="CProd", right_on="CodPF", how="left")
    combined_df["CodPP"] = combined_df["CodPP"].fillna("xxx")

    # Save combined file
    # combined_output = os.path.join(output_dir, f"NFI_{year}_{month_num}_todos.xlsx")
    # combined_df.to_excel(combined_output, index=False)
    
    # Use Template
    template_file = os.path.join(dropbox_root, "KBB MF", "AAA", "Balancetes", "Fechamentos", "data", "Template", "NFI_XML.xlsm")
    
    if not os.path.exists(template_file):
        print(f"❌ Template file not found: {template_file}")
        return

    print(f"Loading template: {template_file}")
    wb = load_workbook(template_file, keep_vba=True)
    
    # Target sheet
    if "NFI" in wb.sheetnames:
        ws = wb["NFI"]
        # Clear existing data if any (optional, but good practice if reusing)
        # For now, assuming template is clean or we append/overwrite. 
        # Better to clear:
        ws.delete_rows(2, ws.max_row)
    else:
        ws = wb.create_sheet("NFI")
        
    # Write dataframe to sheet
    for r in dataframe_to_rows(combined_df, index=False, header=True):
        ws.append(r)
    
    # Add Autofilter
    if ws.max_row > 0:
        ws.auto_filter.ref = ws.dimensions
        
    # Delete Sheet1 if exists
    if "Sheet1" in wb.sheetnames:
        del wb["Sheet1"]
        
    # Save as XLSM
    combined_output = os.path.join(output_dir, f"NFI_{year}_{month_num}_todos.xlsm")
    wb.save(combined_output)
    
    print(f"✅ Combined file saved: {combined_output}")
    print(f"📊 Total rows combined: {len(combined_df)}")

    # === UPDATE RESUMO ===
    print("🔄 Updating ResumoFin25...")
    
    # Calculate totals by Natureza_GRP
    # We need to aggregate by Natureza_GRP
    # Metrics needed:
    # nfs: nunique(NF)
    # linhas: count()
    # itens: sum(qCom)
    # vProd: sum(vProd)
    # vICMS: sum(vICMS_Item)
    # vIPI: sum(vIPI_Item)
    # vFrete: sum(vFrete_Item)
    # Vdesc: sum(vDesc_Item)
    # vNF: sum(vProd + vIPI + vFrete - vDesc) (Calculated per item)
    
    # Calculate item-level vNF for aggregation
    combined_df["vNF_Item"] = combined_df["vProd"] + combined_df["vIPI_Item"] + combined_df["vFrete_Item"] - combined_df["vDesc_Item"]
    
    # Group by Natureza_GRP
    grouped = combined_df.groupby("Natureza_GRP")
    
    totals = {}
    for name, group in grouped:
        totals[f"MA_NFI_{name}_nfs"] = group["NF"].nunique()
        totals[f"MA_NFI_{name}_linhas"] = len(group)
        totals[f"MA_NFI_{name}_itens"] = group["qCom"].sum()
        totals[f"MA_NFI_{name}_vProd"] = group["vProd"].sum()
        totals[f"MA_NFI_{name}_vICMS"] = group["vICMS_Item"].sum()
        totals[f"MA_NFI_{name}_vIPI"] = group["vIPI_Item"].sum()
        totals[f"MA_NFI_{name}_vFrete"] = group["vFrete_Item"].sum()
        totals[f"MA_NFI_{name}_Vdesc"] = group["vDesc_Item"].sum()
        totals[f"MA_NFI_{name}_vNF"] = group["vNF_Item"].sum()

    try:
        if not os.path.exists(resumo_file):
            print(f"⚠️ Resumo file not found: {resumo_file}")
            return

        wb = load_workbook(resumo_file)
        if "Numbers" not in wb.sheetnames:
             print(f"❌ Sheet 'Numbers' not found in {resumo_file}")
             return
        ws = wb["Numbers"]
        
        # Find column for current month (YYMM)
        target_col_header = f"{year[-2:]}{month_num}" # e.g. 2511
        target_col_idx = None
        
        # Iterate through header row (assuming row 1)
        for cell in ws[1]:
            if str(cell.value) == target_col_header:
                target_col_idx = cell.column
                break
        
        if not target_col_idx:
            print(f"❌ Column '{target_col_header}' not found in Resumo file.")
            return

        # Update values
        # Iterate through rows in column A to find keys
        updates_count = 0
        for row in range(2, ws.max_row + 1):
            key = ws.cell(row=row, column=1).value
            if key in totals:
                ws.cell(row=row, column=target_col_idx).value = totals[key]
                updates_count += 1
        
        wb.save(resumo_file)
        print(f"✅ Resumo updated: {updates_count} values written to column {target_col_header}")

    except Exception as e:
        print(f"❌ Error updating Resumo: {e}")



def run_macro(path, macro_name):
    app = xw.App(visible=False)
    try:
        wb = app.books.open(path)
        app.macro(macro_name)()
        wb.save()
        wb.close()
    except Exception as e:
        print(f"❌ Error running macro {macro_name}: {e}")
    finally:
        app.quit()

# === RUN ===
def main(year, month):
    year_str = str(year)
    month_str = f"{month:02d}"

    combine_monthly_items_excels(year_str, month_str)
    
    # Path to the XLSM file
    file_path = os.path.join(BASE_FOLDER, "Mauricio", "Contabilidade", f"{year_str}_{month_str}", f"NFI_{year_str}_{month_str}_todos.xlsm")
    
    if os.path.exists(file_path):
        print(f"Running macro on: {file_path}")
        run_macro(file_path, "Pivot_NFI_XML")
    else:
        print(f"❌ File not found for macro execution: {file_path}")


if __name__ == "__main__":
    import argparse
    from datetime import datetime
    
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", "-y", type=int)
    parser.add_argument("--month", "-m", type=int)
    args = parser.parse_args()
    
    if args.year and args.month:
        main(args.year, args.month)
    else:
        # Default or interactive
        now = datetime.now()
        def_year = now.year
        def_month = now.month - 1 if now.month > 1 else 12
        if def_month == 12: def_year -= 1
        
        print(f"Using default/interactive mode. Default: {def_year}-{def_month}")
        try:
            y = int(input(f"Year [{def_year}]: ") or def_year)
            m = int(input(f"Month [{def_month}]: ") or def_month)
            main(y, m)
        except:
            main(def_year, def_month)
