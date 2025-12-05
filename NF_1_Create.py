import os
import xml.etree.ElementTree as ET
import pandas as pd
import re
from openpyxl import load_workbook

# Run processing for all series for a specific month
YEAR = "2025"
MONTH = "11-Novembro"

# Define base folder and available series
path_options = [
    '/Users/mauricioalouan/Dropbox/nfs',
    '/Users/simon/Library/CloudStorage/Dropbox/nfs'
]
BASE_FOLDER = None
for path in path_options:
    if os.path.exists(path):
        BASE_FOLDER = path
        break

if not BASE_FOLDER:
    print("⚠️ Warning: No valid BASE_FOLDER found.")
    BASE_FOLDER = "/Users/mauricioalouan/Dropbox/nfs"
SERIES_LIST = [
    "Serie 1 - Omie",
    "Serie 2 - filial",
    "Serie 3 - Bling",
    "Serie 4 - Lexos",
    "Serie 5 - Olist",
    "Serie 6 - Meli",
    "Serie 7 - Amazon",
    "Serie 8 - Magalu Full",
    "Serie 9 - Shopee Full"
]

GLOBAL_LOG_LINES = []  # log compartilhado entre todas as séries

def log_global(msg):
    print(msg)
    GLOBAL_LOG_LINES.append(msg)


def process_series(month, year, series):
    """Process XML invoices and cancellation events for a given month, year, and series."""

    import os
    import xml.etree.ElementTree as ET
    import pandas as pd
    import re
    from openpyxl import load_workbook

    # Define output directory
    month_num_for_dir = str(month).split('-')[0].zfill(2)
    output_dir = os.path.join(BASE_FOLDER, "Mauricio", "Contabilidade", f"{year}_{month_num_for_dir}")
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    month_num = month.split('-')[0]
    folder_path = os.path.join(BASE_FOLDER, year, series, month)
    output_file = os.path.join(output_dir, f"NF_{year}_{month_num}_{series}.xlsx")

    if not os.path.exists(folder_path):
        log_global(f"Skipping {series}: Folder not found -> {folder_path}")
        return

    data_list = []
    data_rows_by_key = {}

    # Classify files by XML content
    xml_files = sorted(os.listdir(folder_path))
    invoice_files = []
    event_files = []

    for f in xml_files:
        if not f.endswith(".xml"):
            continue
        file_path = os.path.join(folder_path, f)
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            ns = {"ns": root.tag.split("}")[0].strip("{")} if "}" in root.tag else {}

            if root.find(".//ns:tpEvento", ns) is not None:
                event_files.append(f)
            elif root.find(".//ns:infNFe", ns) is not None:
                invoice_files.append(f)
            else:
                log_global(f"⏭ Ignored unrecognized XML structure: {f}")
        except Exception as e:
            log_global(f"⚠️ Failed to classify {f}: {e}")

    # Process invoices
    for filename in invoice_files:
        file_path = os.path.join(folder_path, filename)
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            namespace = {"ns": root.tag.split("}")[0].strip("{")}

            infNFe = root.find(".//ns:infNFe", namespace)
            chave_nfe = infNFe.attrib.get("Id", "")[3:] if infNFe is not None else ""

            date = root.find(".//ns:ide/ns:dhEmi", namespace)
            date = date.text[:10] if date is not None else "N/A"

            nf = root.find(".//ns:ide/ns:nNF", namespace)
            nf = nf.text if nf is not None else "N/A"

            natureza = root.find(".//ns:ide/ns:natOp", namespace)
            natureza = natureza.text if natureza is not None else "N/A"

            serie = root.find(".//ns:ide/ns:serie", namespace)
            serie = serie.text if serie is not None else "N/A"

            client = root.find(".//ns:dest/ns:xNome", namespace)
            client = client.text if client is not None else "N/A"

            cpf = root.find(".//ns:dest/ns:CPF", namespace)
            cpf = cpf.text if cpf is not None else "N/A"

            pedido = root.find(".//ns:compra/ns:xPed", namespace)
            pedido = pedido.text if pedido is not None else "N/A"

            valor_produto = root.find(".//ns:ICMSTot/ns:vProd", namespace)
            valor_produto = round(float(valor_produto.text), 2) if valor_produto is not None else 0.00

            icms = root.find(".//ns:ICMSTot/ns:vICMS", namespace)
            icms = round(float(icms.text), 2) if icms is not None else 0.00

            st = root.find(".//ns:ICMSTot/ns:vST", namespace)
            st = round(float(st.text), 2) if st is not None else 0.00

            desconto = root.find(".//ns:ICMSTot/ns:vDesc", namespace)
            desconto = round(float(desconto.text), 2) if desconto is not None else 0.00

            frete = root.find(".//ns:ICMSTot/ns:vFrete", namespace)
            frete = round(float(frete.text), 2) if frete is not None else 0.00

            ipi = root.find(".//ns:ICMSTot/ns:vIPI", namespace)
            ipi = round(float(ipi.text), 2) if ipi is not None else 0.00

            desp_ass = root.find(".//ns:ICMSTot/ns:vOutro", namespace)
            desp_ass = round(float(desp_ass.text), 2) if desp_ass is not None else 0.00

            total_nf = root.find(".//ns:ICMSTot/ns:vNF", namespace)
            total_nf = round(float(total_nf.text), 2) if total_nf is not None else 0.00

            cprod = root.find(".//ns:det/ns:prod/ns:cProd", namespace)
            cprod = cprod.text if cprod is not None else "N/A"

            inf_cpl_elem = root.find(".//ns:infAdic/ns:infCpl", namespace)
            if inf_cpl_elem is not None:
                inf_cpl_text = inf_cpl_elem.text
                amazon_order_match = re.search(r"Numero do pedido da compra:\s*([\d-]+)", inf_cpl_text)
                num_ped_am = amazon_order_match.group(1) if amazon_order_match else "N/A"
            else:
                num_ped_am = "N/A"

            check_value = round(valor_produto + st - desconto + ipi + frete + desp_ass - total_nf, 2)

            row_data = [
                filename, date, nf, natureza, serie, client, cpf, pedido, num_ped_am, cprod,
                valor_produto, icms, st, desconto, ipi, frete, desp_ass, total_nf, check_value,
                "Ativa"
            ]
            data_rows_by_key[chave_nfe] = len(data_list)
            data_list.append(row_data)

        except Exception as e:
            log_global(f"❌ Error processing invoice {filename}: {e}")

    # Process event files
    for filename in event_files:
        file_path = os.path.join(folder_path, filename)
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            namespace = {"ns": root.tag.split("}")[0].strip("{")}

            tp_evento_elem = root.find(".//ns:tpEvento", namespace)
            if tp_evento_elem is None:
                continue

            tipo_evento = tp_evento_elem.text.strip()
            if tipo_evento == "110110":
                log_global(f"⏭ Ignored Carta de Correção: {filename}")
                continue
            if tipo_evento != "110111":
                log_global(f"⏭ Ignored unknown event type {tipo_evento}: {filename}")
                continue

            ch_nfe_elem = root.find(".//ns:chNFe", namespace)
            if ch_nfe_elem is not None:
                chave_nfe = ch_nfe_elem.text.strip()
                idx = data_rows_by_key.get(chave_nfe)
                if idx is not None:
                    data_list[idx][-1] = "Cancelado"
                    log_global(f"✔ NF {chave_nfe} marked as Cancelado via {filename}")
                else:
                    log_global(f"⚠ Cancelamento found for unknown NF: {chave_nfe}")
        except Exception as e:
            log_global(f"❌ Error processing event {filename}: {e}")

    df = pd.DataFrame(data_list, columns=[
        "XML_File", "Date", "NF", "Natureza", "Serie", "Client", "CPF", "Pedido", "NumPedAm", "CProd",
        "ValorProduto", "ICMS", "ST", "Desconto", "IPI", "Frete", "DespAss", "TotalNF", "Check", "Status"
    ])

    if df.empty:
        log_global(f"No valid data found for {series} - {month}/{year}. Skipping file creation.")
        return

    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active
    num_format = "#,##0.00"
    num_columns = ["K", "L", "M", "N", "O", "P", "Q", "R", "S"]

    for col in num_columns:
        for row in range(2, ws.max_row + 1):
            ws[f"{col}{row}"].number_format = num_format

    ws.auto_filter.ref = ws.dimensions
    wb.save(output_file)

    log_global(f"✅ Processed {series}: Extracted and formatted data saved to {output_file}")

def get_month_folder_name(month_int):
    months = {
        1: "01-Janeiro", 2: "02-Fevereiro", 3: "03-Março", 4: "04-Abril",
        5: "05-Maio", 6: "06-Junho", 7: "07-Julho", 8: "08-Agosto",
        9: "09-Setembro", 10: "10-Outubro", 11: "11-Novembro", 12: "12-Dezembro"
    }
    return months.get(month_int, f"{month_int:02d}")

# Call the function for all series for a given month
def main(year, month):
    """Iterate through all series and process XML files for a given month and year."""
    # Convert to string format expected by folders
    year_str = str(year)
    month_str = get_month_folder_name(month)
    
    print(f"Processing for {year_str} / {month_str}")

    for series in SERIES_LIST:
        process_series(month_str, year_str, series)

    # Save global log
    month_num = month_str.split('-')[0]
    output_dir = os.path.join(BASE_FOLDER, "Mauricio", "Contabilidade", f"{year}_{month_num}")
    os.makedirs(output_dir, exist_ok=True)
    
    log_path = os.path.join(output_dir, f"NF_{year}_{month_num}_log.txt")
    with open(log_path, "w", encoding="utf-8") as f:
        for line in GLOBAL_LOG_LINES:
            f.write(line + "\n")

    print(f"📝 Global log saved to {log_path}")

if __name__ == "__main__":
    import argparse
    from datetime import datetime
    
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", "-y", type=int)
    parser.add_argument("--month", "-m", type=int)
    args = parser.parse_args()
    
    if args.year and args.month:
        main(args.year, args.month)
    else:
        # Default or interactive
        now = datetime.now()
        def_year = now.year
        def_month = now.month - 1 if now.month > 1 else 12
        if def_month == 12: def_year -= 1
        
        print(f"Using default/interactive mode. Default: {def_year}-{def_month}")
        try:
            y = int(input(f"Year [{def_year}]: ") or def_year)
            m = int(input(f"Month [{def_month}]: ") or def_month)
            main(y, m)
        except:
            main(def_year, def_month)
