"""
This script, `remake_dataset.py`, is part of the KBB MF data pipeline for processing financial and inventory datasets. 
Its primary objectives include:
1. Setting up directories for dynamic and static data sources.
2. Importing essential libraries such as pandas and openpyxl for data manipulation and file handling.
3. Processing and merging datasets, including:
   - Combining dynamic data files with static lookup tables stored in the "Tables" directory.
   - Cleaning and transforming data for reporting and dashboard creation.
4. Leveraging utility functions for tasks like locating directories, managing dates, and formatting Excel outputs.

Prerequisites:
- Ensure that the directories specified in `path_options` exist and contain the necessary files.
- Verify that the "Tables" directory holds the required static tables for lookup and merging operations.

This script is integral to maintaining data accuracy and efficiency in the reporting workflow.
"""


import pandas as pd
from pandas.tseries.offsets import MonthEnd
import os
import shutil
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from dateutil import parser
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
import argparse
import re
import numpy as np


def _select_year_month(year=None, month=None):
    if year and month:
        return year, month

    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("-y", "--year", type=int)
    parser.add_argument("-m", "--month", type=int)
    args, _ = parser.parse_known_args()

    if args.year and args.month:
        return args.year, args.month

    # Default = previous month
    now = datetime.now()
    prev = now.replace(day=1) - relativedelta(days=1)
    def_year, def_month = prev.year, prev.month

    print("Year and/or month not provided.")
    try:
        year  = int(input(f"Enter year (default {def_year}): ") or def_year)
        month = int(input(f"Enter month [1-12] (default {def_month}): ") or def_month)
    except (EOFError, KeyboardInterrupt):
        # non-interactive fallback = previous month
        year, month = def_year, def_month

    return year, month

# Global variables will be set in main()
base_dir = None
static_dir = None
ano_mes = None


def yymm(year, month): return f"{year}_{month:02d}"

# Define the potential base directories
path_options = [
    '/Users/mauricioalouan/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data/',
    '/Users/simon/Library/CloudStorage/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data'
]
# Iterate over the list and set base_dir to the first existing path
for path in path_options:
    if os.path.exists(path):
        base_dir = path
        break
else:
    # If no valid path is found, raise an error or handle it appropriately
    print("None of the specified directories exist.")
    base_dir = None  # Or set a default path if appropriate
print("Base directory set to:", base_dir)

static_dir = os.path.join(base_dir, 'Tables')
#inventory_file_path = os.path.join(static_dir, 'R_EstoqComp.xlsx')  # Update to the correct path if needed

column_rename_dict = {
    'O_NFCI': {
        # Note: Many columns already renamed in process_data.py - don't rename them again
        # Columns already renamed: Data, Qt, PMerc_T, A_ICMSST_T, A_IPI_T, PNF_T, ICMS_T, CODPF, PMerc_U
        # Only rename columns that weren't renamed in process_data.py:
        'Operação': 'OP',
        'Nota Fiscal': 'NF',
        'Cliente (Razão Social)': 'NOMERS',
        'Cliente (Nome Fantasia)': 'NOMEF',
        'Estado': 'UF'
        # Add other columns that need renaming for O_NFCI (not already renamed in process_data.py)
    },
     'L_LPI': {
        # Note: Columns already renamed in process_data.py - don't rename them again
        # Columns already renamed: Integracao, PMerc_T, CodPed, Status, CODPF, Qt
        # Only rename columns that weren't renamed in process_data.py:
        # (no additional renames needed - all handled in process_data.py)
    },
    'MLK_Vendas' : {
        'PREÇO UNITÁRIO DE VENDA DO ANÚNCIO (BRL)': 'PRECOUNIT',
        'Quantidade' : 'QTD',
        'DATA DA VENDA' : 'DATA'
    }
    # Add dictionaries for other dataframes...
}
column_format_dict = {
    'O_NFCI': {
        'Data': 'DD-MMM-YY',  # was EMISS - renamed in process_data.py
        'Qt': '0',  # was QTD - renamed in process_data.py
        'PMerc_U': '#,##0.00',  # was PRECO CALC - renamed in process_data.py
        'PMerc_T': '#,##0.00',  # was MERCVLR - renamed in process_data.py
        'A_ICMSST_T': '#,##0.00',  # was ICMSST - renamed in process_data.py
        'A_IPI_T': '#,##0.00',  # was IPI - renamed in process_data.py
        'PNF_T': '#,##0.00',  # was TOTALNF - renamed in process_data.py
        'ICMS_T': '#,##0.00',  # was ICMS - renamed in process_data.py
        'ECU': '#,##0.00',
        'COMISSPCT': '0.00%',
        'FRETEPCT': '0.00%',
        'VERBAPCT': '0.00%',
        'ECT': '#,##0.00',
        'COMISSVLR': '#,##0.00',
        'FRETEVLR': '#,##0.00',
        'MARGVLR': '#,##0.00',
        'MARGPCT': '0.00%',
        # Add other formats for O_NFCI
    },
    'L_LPI':{
        'PMerc_T': '#,##0.00',  # was VLRVENDA - renamed in process_data.py        
        'DESCONTO PEDIDO SELLER': '#,##0.00',        
        'FRETE SELLER': '#,##0.00',        
        'ECUK': '#,##0.00',
        'ECTK': '#,##0.00',
        'ComissPctMp': '0.0%',
        'ComissPctVlr': '#,##0.00',
        'FreteFixoVlr': '#,##0.00',
        'FreteProdVlr': '#,##0.00',
        'REPASSE': '#,##0.00',
        'ImpLP': '#,##0.00',
        'ImpICMS': '#,##0.00',
        'ImpTot': '#,##0.00',
        'MargVlr': '#,##0.00',
        'MargPct': '0.0%',
    },
    'MLK_Vendas':{
        'VLRTOTALPSKU': '#,##0.00',
        'RECEITAENVIO': '#,##0.00',
        'TARIFAVENDA': '#,##0.00',
        'TARIFAENVIO': '#,##0.00',
        'CANCELAMENTOS': '#,##0.00',
        'REPASSE': '#,##0.00',
        'ECU': '#,##0.00',
        'ECTK': '#,##0.00',
        'Imposto1': '#,##0.00',
        'Imposto2': '#,##0.00',
        'ImpostoT': '#,##0.00',
        'MARGVLR': '#,##0.00',
        'MARGPCT': '0.00%',
    },
    'O_CC': {
        'DATA': 'DD-MMM-YY',  
        'VALOR (R$)': '#,##0.00',
    },
    'O_CtasAPagar': {
        'PREVISÃO': 'DD-MMM-YY',  
        'EMISSÃO': 'DD-MMM-YY',  
        'VENCIMENTO': 'DD-MMM-YY',  
        'REGISTRO': 'DD-MMM-YY',  
        'A PAGAR': '#,##0.00',
    },
    'O_CtasARec': {
        'PREVISÃO': 'DD-MMM-YY',  
        'EMISSÃO': 'DD-MMM-YY',  
        'VENCIMENTO': 'DD-MMM-YY',  
        'REGISTRO': 'DD-MMM-YY',  
        'A RECEBER': '#,##0.00',
        'DATA BASE': 'DD-MMM-YY',  
    },
    'Kon_Detail_SKUAdj': {
        'Venda': '#,##0.00',
        'QTD': '#,##0',
        'Taxa_dir': '#,##0.00',
        'Frete_dir': '#,##0.00',
        'Outros_dir': '#,##0.00',
        'TotDesc_dir': '#,##0.00',
        'Taxa_ind': '#,##0.00',
        'Frete_ind': '#,##0.00',
        'Outros_ind': '#,##0.00',
        'TotDesc_ind': '#,##0.00',
        'Taxa_tot': '#,##0.00',
        'Frete_tot': '#,##0.00',
        'Outros_tot': '#,##0.00',
        'TotDesc_tot': '#,##0.00',
        'ECU': '#,##0.00',
        'Marg_vlr': '#,##0.00',
        'Marg_pct': '0.0%'
    }
}
rows_todrop = {
    'O_NFCI': {
        'C': 0,
    }
}
cols_todrop = {
    'O_NFCI': {
        'PROJETO': 'd',
        'C': 'd',
    },
    'MLK_Vendas':{
        'RECEITA POR ACRÉSCIMO NO PREÇO (PAGO PELO COMPRADOR)': 'd',
        'TAXA DE PARCELAMENTO EQUIVALENTE AO ACRÉSCIMO': 'd',
        'ENDEREÇO': 'd',
        'ENDEREÇO01': 'd',
        'CIDADE': 'd',
        'ESTADO': 'd',
        'CEP': 'd',
    },
    'O_CC': {
        'SALDO (R$)': 'd',
    },
    'O_CtasAPagar': {
        'MINHA EMPRESA (NOME FANTASIA)': 'd',
        'MINHA EMPRESA (RAZÃO SOCIAL)': 'd',
        'MINHA EMPRESA (CNPJ)': 'd',
        'ORÇAMENTO': 'd',
        'VALOR DA CONTA': 'd',
        'VALOR PIS': 'd',
        'VALOR COFINS': 'd',
        'VALOR CSLL': 'd',
        'VALOR IR': 'd',
        'VALOR ISS': 'd',
        'VALOR INSS': 'd',
        'VALOR LÍQUIDO': 'd',
        'VALOR PAGO': 'd',
    },
    'O_CtasARec': {
        'MINHA EMPRESA (NOME FANTASIA)': 'd',
        'MINHA EMPRESA (RAZÃO SOCIAL)': 'd',
        'MINHA EMPRESA (CNPJ)': 'd',
        'ORÇAMENTO': 'd',
        'VALOR DA CONTA': 'd',
        'VALOR PIS': 'd',
        'VALOR COFINS': 'd',
        'VALOR CSLL': 'd',
        'VALOR IR': 'd',
        'VALOR ISS': 'd',
        'VALOR INSS': 'd',
        'VALOR LÍQUIDO': 'd',
        'RECEBIDO': 'd',
    },

}

audit_client_names = ['ALWE', 'COMPROU CHEGOU', 'NEXT COMPRA']  # Add other clients as needed
invaudit_client_names = ['ALWE', 'COMPROU CHEGOU', 'NEXT COMPRA']  # Add other clients as needed

def debug_df(all_data, table_name, label):
    if table_name not in all_data or all_data[table_name].empty:
        print(f"⚠️ [{label}] Tabela '{table_name}' não encontrada ou vazia.")
        return all_data

    df = all_data[table_name]
    print(f"\n🔎 DEBUG {label}: {table_name} shape = {df.shape}")
    print(f"Columns: {list(df.columns)}")
    print(df.head(5))
    return all_data

def rename_columns(all_data, column_rename_dict):
    for df_name, rename_dict in column_rename_dict.items():
        if df_name in all_data:
            df = all_data[df_name]
            # Debug print to verify columns before renaming
            #print(f"Before rename:\nTable: {df_name}\nColumns: {df.columns.tolist()}")
            df.rename(columns=rename_dict, inplace=True)
            # Debug print to verify columns after renaming
            #print(f"Renamed columns in {df_name}: {rename_dict}")
            #print(f"Columns in {df_name}: {df.columns.tolist()}")
            all_data[df_name] = df
    return all_data

def load_recent_data(base_dir, file_pattern, ds_year, ds_month):
    frames = []
    year_month = f"{ds_year}_{ds_month:02d}"
    file_path = os.path.join(base_dir, 'clean', year_month, file_pattern.format(year_month=year_month))

    # Specify which columns should always be treated as strings
    string_columns = ["ORDER_ID", "TRANSACTION_ID", "SHIPPING_ID", "SOURCE_ID", "EXTERNAL_REFERENCE"]

    if os.path.exists(file_path):
        df = pd.read_excel(file_path, dtype={col: str for col in string_columns})
        frames.append(df)
        #print(f"Loaded {file_path} with shape: {df.shape}")  # Debug print
    else:
        print(f"File not found: {file_path}")  # Debug print

    return pd.concat(frames) if frames else pd.DataFrame()

def load_static_data(static_dir, filename):
    return pd.read_excel(os.path.join(static_dir, filename))

def standardize_text_case(df):
    """Convert all text to uppercase for standardization."""
    if isinstance(df, pd.DataFrame):
        df = df.copy()  # avoid chained assignment issues
        df.columns = [str(col).strip().upper() for col in df.columns]
        for col in df.select_dtypes(include=[object]).columns:
            df[col] = df[col].str.upper()
    return df

def clean_dataframes(all_data):
    """
    Drops specified rows and columns from each table in all_data based on global variables rows_todrop and cols_todrop.
    
    Parameters:
    - all_data (dict): Dictionary containing all datasets.

    Returns:
    - all_data (dict): Updated dictionary with rows and columns removed as per defined rules.
    """
    global rows_todrop, cols_todrop

    for key, df in all_data.items():
        # Drop rows based on conditions in rows_todrop
        if key in rows_todrop:
            for col, value in rows_todrop[key].items():
                if col in df.columns:
                    df = df[df[col] != value]  # Keep rows where column is NOT equal to the specified value
                    print(f"Dropped rows in {key} where {col} = {value}")

        # Drop columns based on conditions in cols_todrop
        if key in cols_todrop:
            cols_to_remove = [col for col in cols_todrop[key].keys() if col in df.columns]
            df = df.drop(columns=cols_to_remove, errors='ignore')
            print(f"Dropped columns {cols_to_remove} from {key}")

        # Update the dataframe in all_data
        all_data[key] = df

    return all_data

def split_SKU_lines(df: pd.DataFrame) -> pd.DataFrame:
    """
    Divide linhas com múltiplos SKUs contidos na coluna 'SKU' (separados por vírgula).
    Para cada SKU individual, cria uma nova linha com os mesmos dados.
    Marca como 'MULTIVENDAS' = 1 e adiciona 'SPLIT_ID' para identificar o grupo original.
    Também preenche CODPF = SKU para todas as linhas.
    """
    print("🔍 Dividindo linhas com múltiplos SKUs (split_SKU_lines)...")
    new_rows = []
    group_counter = 0
    split_count = 0

    for _, row in df.iterrows():
        sku_raw = row.get("SKU", "")
        if pd.isna(sku_raw):
            row["CODPF"] = ""
            new_rows.append(row)
            continue

        sku_raw = str(sku_raw).strip()
        sku_list = [s.strip().upper() for s in sku_raw.split(",") if s.strip()]

        if len(sku_list) <= 1:
            row["CODPF"] = sku_list[0] if sku_list else ""
            new_rows.append(row)
            continue

        split_id = f"SPLIT_{group_counter:04d}"
        for sku in sku_list:
            new_row = row.copy()
            new_row["SKU"] = sku
            new_row["CODPF"] = sku
            new_row["MULTIVENDAS"] = 1
            new_row["SPLIT_ID"] = split_id
            new_rows.append(new_row)

        group_counter += 1
        split_count += 1

    print(f"✅ Split concluído: {split_count} linhas com múltiplos SKUs foram divididas.")
    return pd.DataFrame(new_rows)


def split_SKU_lines_by_cost_ratio(all_data: dict) -> dict:
    """
    Para linhas marcadas com 'multivenda' e 'split_id':
    - Redistribui VALOR_REPASSE, VALOR_PREVISTO e DIFERENCA proporcionalmente ao ECU
    - As demais colunas são mantidas iguais
    """
    df = all_data["Kon_RelGeral"].copy()

    print("🔧 Aplicando split_SKU_lines_by_cost_ratio (via split_id)...")
    value_cols = ["VALOR_REPASSE", "VALOR_PREVISTO", "DIFERENCA"]
    untouched = df[df["MULTIVENDAS"].fillna(0) != 1].copy()
    to_split = df[df["MULTIVENDAS"].fillna(0) == 1].copy()

    if "SPLIT_ID" not in to_split.columns:
        print("⚠️ Nenhuma coluna 'SPLIT_ID' encontrada nas linhas multivenda.")
        all_data["Kon_RelGeral"] = df
        return all_data

    result = []
    grouped = to_split.groupby("SPLIT_ID")
    print(f"🔢 Total de grupos para redistribuir: {len(grouped)}")

    for split_id, group in grouped:
        group = group.copy()
        ecus = group["ECU"].replace(0, 1).fillna(1)
        total_ecu = ecus.sum()

        if total_ecu == 0:
            shares = [1 / len(group)] * len(group)
        else:
            shares = ecus / total_ecu

        for col in value_cols:
            if col in group.columns:
                base_val = group[col].sum()
                group[col] = shares * base_val

        result.append(group)

    df_final = pd.concat([untouched] + result, ignore_index=True)
    print(f"✅ Redistribuição aplicada a {len(grouped)} grupos de multivenda.")
    all_data["Kon_RelGeral"] = df_final
    return all_data

def build_Kon_Final_Report(all_data):
    """
    Cria relatório final combinando os dados MAPPED + UNMAPPED:
    - Base: Kon_Ratios_T_MAPPED
    - Acrescenta: Unmapped_total, Unmapped_pct
    - Recalcula: Desc_total, Venda_liq
    - Salva em: all_data["Kon_Ratios_FINAL"]
    """
    df_mapped = all_data["Kon_Ratios_T_MAPPED"].copy()
    df_unmapped = all_data["Kon_Ratios_T_UNMAPPED"].copy()

    # Extrai linhas chave
    venda_total = df_mapped[df_mapped["METRICA"] == "Venda_total"].set_index("METRICA")
    venda_liq_mapped = df_mapped[df_mapped["METRICA"] == "Venda_liq"].set_index("METRICA")
    unmapped_total = df_unmapped[df_unmapped["METRICA"] == "Venda_liq"].set_index("METRICA")

    # Preenche faltantes com 0
    unmapped_total = unmapped_total.reindex(columns=venda_total.columns, fill_value=0)

    # Unmapped_pct = Unmapped_total / Venda_total
    unmapped_pct = unmapped_total.values / venda_total.replace(0, np.nan).values
    unmapped_pct = pd.DataFrame(unmapped_pct, columns=venda_total.columns, index=["Unmapped_pct"]).fillna(0)

    # Desc_final = Venda_total - (Venda_liq_mapped + Unmapped_total)
    venda_liq_final = venda_liq_mapped.values + unmapped_total.values
    desc_final = venda_total.values - venda_liq_final

    venda_liq_final_df = pd.DataFrame(venda_liq_final, columns=venda_total.columns, index=["Venda_liq"])
    desc_final_df = pd.DataFrame(desc_final, columns=venda_total.columns, index=["Desc_total"])

    # Monta resultado final
    df_final = pd.concat([
        venda_total,
        df_mapped[df_mapped["METRICA"].isin([
            "Taxa_total", "Frete_total", "Outros_total", "Taxa_pct", "Frete_pct", "Outros_pct", "Total_pct"
        ])].set_index("METRICA"),
        unmapped_total.rename(index={"Venda_liq": "Unmapped_total"}),
        unmapped_pct,
        desc_final_df,
        venda_liq_final_df
    ])

    df_final.reset_index(inplace=True)
    all_data["Kon_Ratios_FINAL"] = df_final
    print(f"✅ Kon_Ratios_FINAL created with shape: {df_final.shape}")
    return all_data


def build_all_ratio_versions(all_data):
    """
    Gera 3 versões do Kon_Ratios e Kon_Ratios_T:
    A. Total (todas as linhas)
    B. Mapped (CODPP != "")
    C. Unmapped (CODPP == "")

    Salva nas chaves:
        - Kon_Ratios
        - Kon_Ratios_T
        - Kon_Ratios_MAPPED
        - Kon_Ratios_T_MAPPED
        - Kon_Ratios_UNMAPPED
        - Kon_Ratios_T_UNMAPPED
    """
    from copy import deepcopy

    df_all = all_data["Kon_RelGeral"]
    mask_mapped = df_all["CODPP"].astype(str).str.strip() != ""
    mask_unmapped = ~mask_mapped

    # A. Total
    all_data = compute_channel_ratios(all_data)

    # B. Mapped
    temp = deepcopy(all_data)
    temp["Kon_RelGeral"] = df_all[mask_mapped].copy()
    temp = compute_channel_ratios(temp)
    all_data["Kon_Ratios_MAPPED"] = temp["Kon_Ratios"]
    all_data["Kon_Ratios_T_MAPPED"] = temp["Kon_Ratios_T"]

    # C. Unmapped
    temp = deepcopy(all_data)
    temp["Kon_RelGeral"] = df_all[mask_unmapped].copy()
    temp = compute_channel_ratios(temp)
    all_data["Kon_Ratios_UNMAPPED"] = temp["Kon_Ratios"]
    all_data["Kon_Ratios_T_UNMAPPED"] = temp["Kon_Ratios_T"]

    print("✅ Versões Total, MAPPED e UNMAPPED geradas com sucesso.")
    return all_data

def compute_channel_ratios(all_data):
    """
    Build Kon_Ratios (by CANAL) + Kon_Ratios_T (transposed).
    - Kon_Ratios now has Desc_total, Venda_liq, and a TOTAL row.
    - Both tables compute TOTAL percentages from total sums (not sum of per-channel %).
    """
    if "Kon_RelGeral" not in all_data:
        print("⚠️ Kon_RelGeral not found. Skipping ratio computation.")
        return all_data

    df = all_data["Kon_RelGeral"].copy()
    if df.empty or "KON_GR" not in df.columns:
        print("⚠️ Kon_RelGeral empty or missing KON_GR.")
        return all_data

    # Normalize
    df["VALOR_REPASSE"] = pd.to_numeric(df["VALOR_REPASSE"], errors="coerce").fillna(0)
    df["CANAL"] = df["CANAL"].astype(str).str.strip().str.upper()

    channel_order = ["AMAZON", "MAGAZINE LUIZA", "MERCADO LIVRE", "SHOPEE"]
    ratio_rows = []

    for canal in channel_order:
        df_c = df[df["CANAL"] == canal]
        if df_c.empty:
            ratio_rows.append({
                "CANAL": canal,
                "Venda_total": 0.0,
                "Taxa_total": 0.0,
                "Frete_total": 0.0,
                "Outros_total": 0.0,
                "Taxa_pct": 0.0,
                "Frete_pct": 0.0,
                "Outros_pct": 0.0,
                "Total_pct": 0.0
            })
            continue

        total_venda = df_c.loc[df_c["KON_GR"].str.upper() == "VENDA", "VALOR_REPASSE"].sum()
        total_taxa  = df_c.loc[df_c["KON_GR"].str.upper() == "TAXA-COMISSAO", "VALOR_REPASSE"].sum()
        total_frete = df_c.loc[df_c["KON_GR"].str.upper() == "FRETE", "VALOR_REPASSE"].sum()
        total_outros= df_c.loc[df_c["KON_GR"].str.upper() == "OUTROS", "VALOR_REPASSE"].sum()

        if total_venda == 0:
            ratio_rows.append({
                "CANAL": canal,
                "Venda_total": 0.0,
                "Taxa_total": 0.0,
                "Frete_total": 0.0,
                "Outros_total": 0.0,
                "Taxa_pct": 0.0,
                "Frete_pct": 0.0,
                "Outros_pct": 0.0,
                "Total_pct": 0.0
            })
            continue

        taxa_pct   = total_taxa  / total_venda
        frete_pct  = total_frete / total_venda
        outros_pct = total_outros/ total_venda
        total_pct  = taxa_pct + frete_pct + outros_pct

        ratio_rows.append({
            "CANAL": canal,
            "Venda_total": total_venda,
            "Taxa_total": total_taxa,
            "Frete_total": total_frete,
            "Outros_total": total_outros,
            "Taxa_pct": taxa_pct,
            "Frete_pct": frete_pct,
            "Outros_pct": outros_pct,
            "Total_pct": total_pct
        })

    # === Kon_Ratios base ===
    df_ratios = pd.DataFrame(ratio_rows)

    # Add Desc_total & Venda_liq
    df_ratios["Desc_total"] = df_ratios["Taxa_total"] + df_ratios["Frete_total"] + df_ratios["Outros_total"]
    df_ratios["Venda_liq"]  = df_ratios["Venda_total"] + df_ratios["Desc_total"]  # desc are negative

    # Build TOTAL row (values are sums; % recomputed from those sums)
    tot_Venda  = df_ratios["Venda_total"].sum()
    tot_Taxa   = df_ratios["Taxa_total"].sum()
    tot_Frete  = df_ratios["Frete_total"].sum()
    tot_Outros = df_ratios["Outros_total"].sum()
    tot_Desc   = df_ratios["Desc_total"].sum()
    tot_Liq    = df_ratios["Venda_liq"].sum()

    if tot_Venda != 0:
        tot_Taxa_pct   = tot_Taxa  / tot_Venda
        tot_Frete_pct  = tot_Frete / tot_Venda
        tot_Outros_pct = tot_Outros/ tot_Venda
        tot_Total_pct  = tot_Desc  / tot_Venda
    else:
        tot_Taxa_pct = tot_Frete_pct = tot_Outros_pct = tot_Total_pct = 0.0

    total_row = {
        "CANAL": "TOTAL",
        "Venda_total": tot_Venda,
        "Taxa_total":  tot_Taxa,
        "Frete_total": tot_Frete,
        "Outros_total":tot_Outros,
        "Taxa_pct":    tot_Taxa_pct,
        "Frete_pct":   tot_Frete_pct,
        "Outros_pct":  tot_Outros_pct,
        "Total_pct":   tot_Total_pct,
        "Desc_total":  tot_Desc,
        "Venda_liq":   tot_Liq
    }

    df_ratios = pd.concat([df_ratios, pd.DataFrame([total_row])], ignore_index=True)
    all_data["Kon_Ratios"] = df_ratios
    print(f"✅ Kon_Ratios created with {len(df_ratios)} rows (incl. TOTAL).")

    # === Kon_Ratios_T (transposed) ===
    df_t = (
        df_ratios.set_index("CANAL")
                 .T
                 .rename_axis("METRICA")
                 .reset_index()
    )

    # Recreate TOTAL column as sum across channels for value metrics…
    canal_cols = [c for c in df_t.columns if c not in ("METRICA")]
    if "TOTAL" in canal_cols:
        canal_cols_no_total = [c for c in canal_cols if c != "TOTAL"]
    else:
        canal_cols_no_total = canal_cols

    # First, sum across channels for every row
    df_t["TOTAL"] = df_t[canal_cols_no_total].select_dtypes(include=["number"]).sum(axis=1)

    # …but FIX percent rows to recompute from totals (not sum of %)
    pct_map = {
        "Taxa_pct":   (tot_Taxa,  tot_Venda),
        "Frete_pct":  (tot_Frete, tot_Venda),
        "Outros_pct": (tot_Outros,tot_Venda),
        "Total_pct":  (tot_Desc,  tot_Venda),
    }
    for r, (num, den) in pct_map.items():
        if den == 0:
            val = 0.0
        else:
            val = num / den
        df_t.loc[df_t["METRICA"] == r, "TOTAL"] = val

    # Optional rounding
    for c in [c for c in df_t.columns if c != "METRICA"]:
        df_t[c] = pd.to_numeric(df_t[c], errors="ignore")
        if df_t[c].dtype.kind in "fc":
            df_t[c] = df_t[c].round(2)

    all_data["Kon_Ratios_T"] = df_t
    print(f"✅ Kon_Ratios_T created with shape: {df_t.shape}")

    return all_data

def build_Kon_Report_from_df(df, name="Kon_Report_Custom"):
    """
    Generic version of build_Kon_Report1 that accepts a dataframe directly.
    Produces a pivoted report grouped by KON_GR, KON_SGR, and CANAL.
    """
    required_cols = ["CANAL", "VALOR_REPASSE", "KON_GR", "KON_SGR"]
    for c in required_cols:
        if c not in df.columns:
            print(f"⚠️ Missing column {c} in input for {name}. Skipping.")
            return pd.DataFrame()

    df = df[~df["KON_GR"].isin(["Saldo", "Saque"])]

    pivot = (
        df.groupby(["KON_GR", "KON_SGR", "CANAL"], as_index=False)["VALOR_REPASSE"]
          .sum()
    )

    pivot = pivot.pivot_table(
        index=["KON_GR", "KON_SGR"],
        columns="CANAL",
        values="VALOR_REPASSE",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    channel_order = ["Amazon", "Magazine Luiza", "Mercado Livre", "Shopee"]
    for ch in channel_order:
        if ch not in pivot.columns:
            pivot[ch] = 0

    pivot = pivot[["KON_GR", "KON_SGR"] + channel_order]
    pivot["TOTAL"] = pivot[channel_order].sum(axis=1)
    pivot["Report_Name"] = name

    print(f"✅ {name} created with shape: {pivot.shape}")
    return pivot

def allocate_nosku_deductions(all_data):
    """
    Build non-SKU report, compute deduction ratios per marketplace,
    and redistribute deductions across SKU-valid lines.
    Creates:
      - Kon_Report_NoSKU   : summary for unmatched transactions
      - Kon_Report_SKUAdj  : SKU-level report with allocated deductions
      - Kon_Ratios         : summary of per-marketplace ratios
    """
    if "Kon_RelGeral" not in all_data:
        print("⚠️ Kon_RelGeral not found. Skipping allocation.")
        return all_data

    df = all_data["Kon_RelGeral"].copy()
    if "CODPP" not in df.columns:
        print("⚠️ CODPP column missing. Cannot flag SKU validity.")
        return all_data

    # --- flag valid vs invalid SKU lines ---
    df["VALID_SKU"] = df["CODPP"].astype(str).str.strip().ne("")

    df_valid = df[df["VALID_SKU"] == True].copy()
    df_non   = df[df["VALID_SKU"] == False].copy()

    # --- build non-SKU report ---
    rep_non = build_Kon_Report_from_df(df_non, name="Kon_Report_NoSKU")
    all_data["Kon_Report_NoSKU"] = rep_non

    # --- calculate deduction ratios per CANAL ---
    channel_order = ["Amazon", "Magazine Luiza", "Mercado Livre", "Shopee"]
    ratios = {}

    for canal in channel_order:
        df_canal = df_non[df_non["CANAL"] == canal]
        if df_canal.empty:
            ratios[canal] = {"Taxa_pct": 0, "Frete_pct": 0, "Outros_pct": 0}
            continue

        rep_canal = build_Kon_Report_from_df(df_canal, name=f"Kon_NoSKU_{canal}")
        agg = rep_canal.groupby("KON_GR")[["TOTAL"]].sum()

        venda_tot  = agg.loc["Venda"]["TOTAL"] if "Venda" in agg.index else 1
        taxa_tot   = agg.loc["Taxa-Comissao"]["TOTAL"] if "Taxa-Comissao" in agg.index else 0
        frete_tot  = agg.loc["Frete"]["TOTAL"] if "Frete" in agg.index else 0
        outros_tot = agg.loc["Outros"]["TOTAL"] if "Outros" in agg.index else 0

        ratios[canal] = {
            "Taxa_pct":  taxa_tot / venda_tot,
            "Frete_pct": frete_tot / venda_tot,
            "Outros_pct": outros_tot / venda_tot,
        }

    # --- print summary of ratios ---
    print("💡 Non-SKU ratios by marketplace:")
    for canal, vals in ratios.items():
        print(f"   {canal}: {vals}")

    # --- apply per-marketplace ratios to valid SKUs ---
    df_valid["Taxa_val"] = 0.0
    df_valid["Frete_val"] = 0.0
    df_valid["Outros_val"] = 0.0

    for canal, pct in ratios.items():
        mask = df_valid["CANAL"] == canal
        df_valid.loc[mask, "Taxa_val"]  = -df_valid.loc[mask, "VALOR_REPASSE"] * pct["Taxa_pct"]
        df_valid.loc[mask, "Frete_val"] = -df_valid.loc[mask, "VALOR_REPASSE"] * pct["Frete_pct"]
        df_valid.loc[mask, "Outros_val"]= -df_valid.loc[mask, "VALOR_REPASSE"] * pct["Outros_pct"]

    # --- compute net adjusted repasse ---
    df_valid["VALOR_REPASSE_LIQ"] = (
        df_valid["VALOR_REPASSE"]
        + df_valid["Taxa_val"]
        + df_valid["Frete_val"]
        + df_valid["Outros_val"]
    )

    # --- store adjusted detail table ---
    all_data["Kon_RelGeral_SKUAdj"] = df_valid

    # --- build SKU-adjusted report summary ---
    rep_sku = build_Kon_Report_from_df(df_valid, name="Kon_Report_SKUAdj")
    all_data["Kon_Report_SKUAdj"] = rep_sku

    # --- build ratio summary for audit ---
    ratio_rows = []
    for canal, vals in ratios.items():
        ratio_rows.append({
            "CANAL": canal,
            "Taxa_pct": vals["Taxa_pct"],
            "Frete_pct": vals["Frete_pct"],
            "Outros_pct": vals["Outros_pct"],
            "Total_pct": vals["Taxa_pct"] + vals["Frete_pct"] + vals["Outros_pct"],
        })
    df_ratios = pd.DataFrame(ratio_rows)
    all_data["Kon_Ratios"] = df_ratios

    print(f"✅ Allocation completed. Added sheets: Kon_Report_NoSKU, Kon_Report_SKUAdj, Kon_Ratios")
    return all_data

def add_unmapped_skus(all_data):
    """
    Aggregate unmapped sales per channel and create synthetic 'UNMAPPED' SKUs.
    Uses mapped SKUs to estimate avg ticket and cost ratio for each channel.
    Appends one synthetic SKU row per channel to Kon_RelGeral for full reconciliation.
    """
    if "Kon_RelGeral" not in all_data or "T_ProdF" not in all_data:
        print("⚠️ Kon_RelGeral or T_ProdF not found. Skipping unmapped synthesis.")
        return all_data

    df = all_data["Kon_RelGeral"].copy()

    # --- Flag valid vs invalid SKU lines ---
    df["VALID_SKU"] = df["CODPP"].astype(str).str.strip().ne("")

    df_valid = df[df["VALID_SKU"] == True].copy()
    df_unmapped = df[df["VALID_SKU"] == False].copy()

    if df_unmapped.empty:
        print("✅ No unmapped sales found. Nothing to synthesize.")
        all_data["Kon_RelGeral"] = df
        return all_data

    channel_order = ["Amazon", "Magazine Luiza", "Mercado Livre", "Shopee"]
    unmapped_rows = []

    for canal in channel_order:
        df_valid_c = df_valid[df_valid["CANAL"] == canal]
        df_unm_c   = df_unmapped[df_unmapped["CANAL"] == canal]

        if df_unm_c.empty or df_valid_c.empty:
            continue

        # --- Core metrics from mapped SKUs ---
        total_sales_mapped = df_valid_c["VALOR_REPASSE"].sum()
        total_cost_mapped  = (
            df_valid_c.merge(all_data["T_ProdF"][["CODPF", "ECU"]],
                             left_on="CODPF", right_on="CODPF", how="left")
                      .assign(ECU=lambda d: pd.to_numeric(d["ECU"], errors="coerce").fillna(0))
                      .eval("ECU * 1")
        ).sum()
        total_qty_mapped = df_valid_c.shape[0]  # since no QTD field per sale

        avg_ticket = total_sales_mapped / total_qty_mapped if total_qty_mapped else 0
        cost_ratio = total_cost_mapped / total_sales_mapped if total_sales_mapped else 0

        # --- Unmapped totals ---
        venda_unmapped = df_unm_c["VALOR_REPASSE"].sum()
        qtd_unmapped   = round(venda_unmapped / avg_ticket) if avg_ticket else 1
        ecu_unmapped   = cost_ratio * avg_ticket

        unmapped_rows.append({
            "SKU": "UNMAPPED",
            "CODPF": "UNMAPPED",
            "CODPP": "UNMAPPED",
            "KON_GR": "Venda",
            "KON_SGR": "UNMAPPED",
            "CANAL": canal,
            "VALOR_REPASSE": venda_unmapped,
            "QTD": qtd_unmapped,
            "ECU": ecu_unmapped,
            "VALID_SKU": True
        })

        print(f"💡 Added UNMAPPED SKU for {canal}: Venda={venda_unmapped:.2f}, "
              f"QTD≈{qtd_unmapped}, ECU≈{ecu_unmapped:.2f}")

    # --- Append synthetic rows ---
    if unmapped_rows:
        df_unmapped_new = pd.DataFrame(unmapped_rows)
        df = pd.concat([df, df_unmapped_new], ignore_index=True)
        print(f"✅ Appended {len(df_unmapped_new)} synthetic UNMAPPED rows.")

    all_data["Kon_RelGeral"] = df
    return all_data

def build_Kon_Detail_SKUAdj(all_data):
    """
    Create detailed SKU-level cost table combining direct and indirect deductions 
    (Taxa/Frete/Outros) and calculate gross margin.
    Output sheet: Kon_Detail_SKUAdj
    """
    if "Kon_RelGeral" not in all_data:
        print("⚠️ Kon_RelGeral not found. Skipping detail table.")
        return all_data

    df = all_data["Kon_RelGeral"].copy()
    if df.empty:
        print("⚠️ Kon_RelGeral is empty.")
        return all_data

    # --- Check required columns ---
    required_cols = ["SKU", "CANAL", "VALOR_REPASSE", "CODPP", "KON_GR"]
    for c in required_cols:
        if c not in df.columns:
            print(f"⚠️ Missing column {c} in Kon_RelGeral.")
            return all_data

    # Normalize
    df["VALOR_REPASSE"] = pd.to_numeric(df["VALOR_REPASSE"], errors="coerce").fillna(0)
    df["KON_GR"] = df["KON_GR"].astype(str).str.upper()

    # --- Calcular valores diretos agrupando por SKU + CANAL + REF_PEDIDO ---
    group_keys = ["SKU", "CODPP", "CANAL", "REF_PEDIDO"]

    df_dir = (
        df[df["KON_GR"].isin(["TAXA-COMISSAO", "FRETE", "OUTROS"])]
        .groupby(group_keys + ["KON_GR"])["VALOR_REPASSE"]
        .sum()
        .unstack(fill_value=0)
        .rename(columns={
            "TAXA-COMISSAO": "Taxa_dir",
            "FRETE": "Frete_dir",
            "OUTROS": "Outros_dir"
        })
        .reset_index()
    )

    # --- Selecionar apenas linhas de Venda ---
    df_base = df[df["KON_GR"] == "VENDA"].copy()

    # --- Add default quantity (proxy 1 por REF_PEDIDO) ---
    df_base["QTD"] = 1

    # --- Merge valores diretos ---
    df_base = df_base.merge(df_dir, on=group_keys, how="left")
    for col in ["Taxa_dir", "Frete_dir", "Outros_dir"]:
        df_base[col] = df_base[col].fillna(0)
    df_base["TotDesc_dir"] = df_base["Taxa_dir"] + df_base["Frete_dir"] + df_base["Outros_dir"]

    # --- Trazer ratios por canal (indiretos) ---
    if "Kon_Ratios" in all_data and not all_data["Kon_Ratios"].empty:
        ratios = all_data["Kon_Ratios"].copy()
        ratios["CANAL"] = ratios["CANAL"].astype(str).str.upper()
        df_base["CANAL"] = df_base["CANAL"].astype(str).str.upper()
        df_base = df_base.merge(ratios, on="CANAL", how="left")
    else:
        df_base["Taxa_pct"] = df_base["Frete_pct"] = df_base["Outros_pct"] = df_base["Total_pct"] = 0.0

    # --- Cálculo deduções indiretas ---
    df_base["Taxa_ind"]   = -df_base["VALOR_REPASSE"] * df_base["Taxa_pct"].fillna(0)
    df_base["Frete_ind"]  = -df_base["VALOR_REPASSE"] * df_base["Frete_pct"].fillna(0)
    df_base["Outros_ind"] = -df_base["VALOR_REPASSE"] * df_base["Outros_pct"].fillna(0)
    df_base["TotDesc_ind"] = df_base["Taxa_ind"] + df_base["Frete_ind"] + df_base["Outros_ind"]

    # --- Totais (direto + indireto) ---
    df_base["Taxa_tot"]   = df_base["Taxa_dir"] + df_base["Taxa_ind"]
    df_base["Frete_tot"]  = df_base["Frete_dir"] + df_base["Frete_ind"]
    df_base["Outros_tot"] = df_base["Outros_dir"] + df_base["Outros_ind"]
    df_base["TotDesc_tot"] = df_base["TotDesc_dir"] + df_base["TotDesc_ind"]

    # --- Custo unitário e margem ---
    df_base["ECU"] = pd.to_numeric(df_base["ECU"], errors="coerce").fillna(0)
    df_base["Venda"] = df_base["VALOR_REPASSE"]

    df_base["Marg_vlr"] = df_base["Venda"] * (1 - 0.275) - df_base["TotDesc_tot"] - (df_base["ECU"] * df_base["QTD"])
    df_base["Marg_pct"] = df_base.apply(
        lambda r: r["Marg_vlr"] / abs(r["Venda"]) if r["Venda"] != 0 else 0, axis=1
    )

    # --- Selecionar colunas finais ---
    col_order = [
        "SKU", "CODPP", "CANAL", "REF_PEDIDO", "Venda", "QTD",
        "Taxa_dir", "Frete_dir", "Outros_dir", "TotDesc_dir",
        "Taxa_ind", "Frete_ind", "Outros_ind", "TotDesc_ind",
        "Taxa_tot", "Frete_tot", "Outros_tot", "TotDesc_tot",
        "ECU", "Marg_vlr", "Marg_pct"
    ]
    df_final = df_base[col_order]

    all_data["Kon_Detail_SKUAdj"] = df_final
    print(f"✅ Kon_Detail_SKUAdj created with shape: {df_final.shape}")
    return all_data

def merge_all_data(all_data):
    print(f"Creating Merged and Calculated Columns")

    # Ensure all relevant columns are in uppercase for case-insensitive comparison
    all_data = {key: standardize_text_case(df) for key, df in all_data.items()}

    # compute column ANOMES
    compute_NFCI_ANOMES(all_data)
    compute_LPI_ANOMES(all_data)
    compute_ML_ANOMES(all_data)
    compute_CC_ANOMES(all_data)

    # Merge O_NFCI with T_Remessas - REM
    all_data = merge_data(all_data, "O_NFCI", "NOMEF", "T_Remessas", "NOMEF", "REM_NF", default_value=0)

    # Merge O_NFCI with T_Prodf - CODPP
    all_data = merge_data(all_data, "O_NFCI", "CodPF", "T_ProdF", "CodPF", "CODPP", default_value="xxx")
    all_data = merge_data(all_data, "L_LPI", "CodPF", "T_ProdF", "CodPF", "CODPP", default_value="xxx")
    all_data = merge_data(all_data, "MLA_Vendas", "SKU", "T_ProdF", "CodPF", "CODPP", default_value="xxx")
    all_data = merge_data(all_data, "MLK_Vendas", "SKU", "T_ProdF", "CodPF", "CODPP", default_value="xxx")

    # Merge O_NFCI with T_GruposCli - G1
    all_data = merge_data(all_data, "O_NFCI", "NomeF", "T_GruposCli", "NomeF", "G1", default_value="V")

    # Merge O_NFCI with ECU on columns 'Data' and 'CodPF'
    # Note: Column renamed from 'EMISS' to 'Data' in process_data.py
    all_data = merge_data_lastcost(all_data, df1_name="O_NFCI",
        df1_product_col="CODPP",
        df1_date_col="Data",
        df2_name="T_Entradas",
        df2_product_col="PAI",
        df2_date_col="ULTIMA ENTRADA",
        df2_cost_col="ULT CU R$",
        new_col_name="ECU",
        default_value=999
    )
    # L_LPI
    all_data = merge_data_lastcost(all_data, df1_name="L_LPI",
        df1_product_col="CODPP",
        df1_date_col="DATA",
        df2_name="T_Entradas",
        df2_product_col="PAI",
        df2_date_col="ULTIMA ENTRADA",
        df2_cost_col="ULT CU R$",
        new_col_name="ECUK",
        default_value=999
    )
    # MLK_Vendas
    all_data = merge_data_lastcost(all_data, df1_name="MLK_Vendas",
        df1_product_col="CODPP",
        df1_date_col="DATA DA VENDA",
        df2_name="T_Entradas",
        df2_product_col="PAI",
        df2_date_col="ULTIMA ENTRADA",
        df2_cost_col="ULT CU R$",
        new_col_name="ECU",
        default_value=999
    )

    all_data = merge_data2v(all_data, "MLA_Vendas", "ANOMES", "SKU", "ECU", "ANOMES", "CODPF", "VALUE", "ECU", default_value=999)

    # Merge VENDEDOR with T_REPS for COMPCT
    all_data = merge_data(all_data, "O_NFCI", "Vendedor", "T_Reps", "Vendedor", "COMISSPCT", default_value=0)

    # Merge UF with T_Fretes for FretePCT
    all_data = merge_data(all_data, "O_NFCI", "UF", "T_Fretes", "UF", "FRETEPCT", default_value=0)
    # Set FRETEPCT = 0 where G1 = "DROP" or "ALWE"
    if 'O_NFCI' in all_data:
        all_data['O_NFCI'].loc[all_data['O_NFCI']['G1'].isin(['DROP', 'ALWE']), 'FRETEPCT'] = 0

    # Merge NomeF with T_Verbas for VerbaPct
    all_data = merge_data(all_data, "O_NFCI", "NOMEF", "T_Verbas", "NomeF", "VERBAPCT", default_value=0)

    # MP merges
    # Note: Column renamed from 'Integração' to 'Integracao' in process_data.py
    all_data = merge_data(all_data, "L_LPI", "Integracao", "T_MP", "Integração", "Empresa",  default_value='erro')
    all_data = merge_data(all_data, "L_LPI", "Integracao", "T_MP", "Integração", "MP",       default_value='erro')
    all_data = merge_data(all_data, "L_LPI", "Integracao", "T_MP", "Integração", "EmpresaF", default_value='erro')

    # OrderStatus
    all_data = merge_data(all_data, "MLA_Vendas", "STATUS", "T_MLStatus", "MLStatus", "OrderStatus", default_value='erro')
    all_data = merge_data(all_data, "MLK_Vendas", "STATUS", "T_MLStatus", "MLStatus", "OrderStatus", default_value='erro')

    # Ctas a Pagar e Receber
    all_data = merge_data(all_data, "O_CtasAPagar", "CATEGORIA", "T_CtasAPagarClass", "Categoria", "GrupoCtasAPagar", default_value='erro')
    # all_data = merge_data(all_data, "O_CtasARec", "CATEGORIA", "T_CtasARecClass", "Categoria", "GrupoCtasAPagar", default_value='erro')

    # CC
    all_data = merge_data(all_data, "O_CC", "CATEGORIA", "T_CCCats", "CC_Categoria Omie", "CC_Cat SG",  default_value='erro')
    all_data = merge_data(all_data, "O_CC", "CATEGORIA", "T_CCCats", "CC_Categoria Omie", "CC_Cat Grp", default_value='erro')
    all_data = merge_data(all_data, "O_CC", "CATEGORIA", "T_CCCats", "CC_Categoria Omie", "CC_B2X",     default_value='erro')
    all_data = merge_data(all_data, "O_CC", "CATEGORIA", "T_CCCats", "CC_Categoria Omie", "CC_Tipo",    default_value='erro')

    # Kon_RelGeral joins with T_KonCats
    all_data = merge_data(all_data,
        df1_name="Kon_RelGeral",
        df1_col="TP_Lancamento",
        df2_name="T_KonCats",
        df2_col="TP_Lancamento",
        new_col="KON_GR",
        default_value="Outros"
    )
    all_data = merge_data(all_data,
        df1_name="Kon_RelGeral",
        df1_col="TP_Lancamento",
        df2_name="T_KonCats",
        df2_col="TP_Lancamento",
        new_col="KON_SGR",
        default_value="Outros"
    )

    for key, df in all_data.items():
        if key == 'O_NFCI':
            # Note: Using new column names from process_data.py
            df['C'] = 1 - df['REM_NF']
            df['B'] = df.apply(lambda row: 1 if row['OP'] == 'REMESSA DE PRODUTO' and row['C'] == 1 else 0, axis=1)
            df['ECT'] = df['ECU'] * df['QT'] * df['C']
            df['COMISSVLR'] = df['PMERC_T'] * df['COMISSPCT'] * df['C']
            df['FRETEVLR'] = df.apply(lambda row: max(row['FRETEPCT'] * row['PNF_T'] * row['C'],
                                                      row['FRETEPCT'] * row['ECT'] * row['C'] * 2), axis=1)
            df['VERBAVLR'] = df['VERBAPCT'] * df['PNF_T'] * df['C']
            df['MARGVLR'] = df['C'] * ( df['PMERC_T'] * (1 - 0.0925) - df['ICMS_T'] ) - df['VERBAVLR'] - df['FRETEVLR'] - df['COMISSVLR'] - df['ECT']
            df['MARGPCT'] = df['MARGVLR'] / df['PMERC_T']

        elif key == 'L_LPI':
            # Note: All column names are uppercase after standardize_text_case
            # make sure join key is string
            if 'CODPED' in df.columns:
                df['CODPED'] = df['CODPED'].astype(str).str.strip()

            cols_to_drop = ['PREÇO', 'PREÇO TOTAL', 'DESCONTO ITEM', 'DESCONTO TOTAL']
            df = df.drop([x for x in cols_to_drop if x in df.columns], axis=1)

            df["MP2"] = df["MP"].str[:2]
            df['VALIDO'] = df['STATUS'].apply(lambda x: 0 if x in ['CANCELADO', 'PENDENTE', 'AGUARDANDO PAGAMENTO'] else 1)
            df['KAB'] = df.apply(lambda row: 1 if row['VALIDO'] == 1 and row['EMPRESA'] in ['K', 'A', 'B'] else 0, axis=1)
            df['ECTK'] = df['ECUK'] * df['QT'] * df['KAB']

            # ----- TipoAnuncio from MLK_Vendas -----
            if ('MLK_Vendas' in all_data and
                not all_data['MLK_Vendas'].empty and
                {'N.º DE VENDA', 'TIPO DE ANÚNCIO'}.issubset(set(all_data['MLK_Vendas'].columns))):
                all_data['MLK_Vendas']['N.º DE VENDA'] = all_data['MLK_Vendas']['N.º DE VENDA'].astype(str).str.strip()
                print_table_head(all_data, "MLK_Vendas")
                df = df.merge(
                    all_data['MLK_Vendas'][['N.º DE VENDA', 'TIPO DE ANÚNCIO']],
                    left_on='CODPED',
                    right_on='N.º DE VENDA',
                    how='left'
                )
                df['TipoAnuncioK'] = df.apply(
                    lambda row: 'ML' + row['TIPO DE ANÚNCIO'][:2] if pd.notna(row['TIPO DE ANÚNCIO']) and row['EMPRESA'] == 'K' and row['MP'] == 'ML' else None,
                    axis=1
                )
                df.drop(columns=['N.º DE VENDA', 'TIPO DE ANÚNCIO'], inplace=True)
            else:
                # ensure column exists so later drop doesn't crash
                df['TipoAnuncioK'] = None

            # ----- TipoAnuncio from MLA_Vendas -----
            if ('MLA_Vendas' in all_data and
                not all_data['MLA_Vendas'].empty and
                {'N.º DE VENDA', 'TIPO DE ANÚNCIO'}.issubset(set(all_data['MLA_Vendas'].columns))):
                all_data['MLA_Vendas']['N.º DE VENDA'] = all_data['MLA_Vendas']['N.º DE VENDA'].astype(str).str.strip()
                df = df.merge(
                    all_data['MLA_Vendas'][['N.º DE VENDA', 'TIPO DE ANÚNCIO']],
                    left_on='CODPED',
                    right_on='N.º DE VENDA',
                    how='left'
                )
                df['TipoAnuncioA'] = df.apply(
                    lambda row: 'ML' + row['TIPO DE ANÚNCIO'][:2] if pd.notna(row['TIPO DE ANÚNCIO']) and row['EMPRESA'] == 'A' and row['MP'] == 'ML' else None,
                    axis=1
                )
                df.drop(columns=['N.º DE VENDA', 'TIPO DE ANÚNCIO'], inplace=True)
            else:
                df['TipoAnuncioA'] = None  # ensure column exists

            # ----- TipoAnuncio from MLB_Vendas -----
            if ('MLB_Vendas' in all_data and
                not all_data['MLB_Vendas'].empty and
                {'N.º DE VENDA', 'TIPO DE ANÚNCIO'}.issubset(set(all_data['MLB_Vendas'].columns))):
                all_data['MLB_Vendas']['N.º DE VENDA'] = all_data['MLB_Vendas']['N.º DE VENDA'].astype(str).str.strip()
                df = df.merge(
                    all_data['MLB_Vendas'][['N.º DE VENDA', 'TIPO DE ANÚNCIO']],
                    left_on='CODPED',
                    right_on='N.º DE VENDA',
                    how='left'
                )
                df['TipoAnuncioB'] = df.apply(
                    lambda row: 'ML' + row['TIPO DE ANÚNCIO'][:2] if pd.notna(row['TIPO DE ANÚNCIO']) and row['EMPRESA'] == 'B' and row['MP'] == 'ML' else None,
                    axis=1
                )
                df.drop(columns=['N.º DE VENDA', 'TIPO DE ANÚNCIO'], inplace=True)
            else:
                # previous behavior kept a default for B
                df['TipoAnuncioB'] = df.apply(lambda row: 'MLG' if row['EMPRESA'] == 'B' and row['MP'] == 'ML' else None, axis=1)

            # Merge TipoAnuncio K/A/B into MP (only if MP == 'ML')
            df['MP'] = df.apply(
                lambda row: row['TipoAnuncioK'] if row['EMPRESA'] == 'K' and row['MP'] == 'ML' and pd.notna(row['TipoAnuncioK']) else
                            row['TipoAnuncioA'] if row['EMPRESA'] == 'A' and row['MP'] == 'ML' and pd.notna(row['TipoAnuncioA']) else
                            row['TipoAnuncioB'] if row['EMPRESA'] == 'B' and row['MP'] == 'ML' and pd.notna(row['TipoAnuncioB']) else row['MP'],
                axis=1
            )
            # Drop helper cols
            df.drop(columns=['TipoAnuncioK', 'TipoAnuncioA', 'TipoAnuncioB'], inplace=True)

            # Comissão por MP
            if 'T_RegrasMP' in all_data:
                df = df.merge(
                    all_data['T_RegrasMP'][['MPX', 'TARMP', 'FFABAIXODE', 'FRETEFIX']],
                    left_on='MP',
                    right_on='MPX',
                    how='left'
                )
                df['ComissPctMp'] = df['TARMP']
                df['ComissPctVlr'] = df['PMERC_T'] * df['ComissPctMp'] * -1
                df['FreteFixoVlr'] = df.apply(
                    lambda row: -row['FRETEFIX'] if row['PMERC_T'] < row['FFABAIXODE'] else 0,
                    axis=1
                )
                df.drop(columns=['MPX', 'TARMP', 'FFABAIXODE', 'FRETEFIX'], inplace=True)

            # Frete por produto/MP
            if 'T_FretesMP' in all_data:
                df_fretesmp = all_data['T_FretesMP']
                df['MP_2L'] = df['MP'].str[:2].str.upper()

                def get_frete_pai(row):
                    codpp = row['CODPP']
                    mp_col = row['MP_2L']
                    if mp_col not in df_fretesmp.columns:
                        return 99
                    match = df_fretesmp[df_fretesmp['CODPP'] == codpp]
                    if not match.empty:
                        return match[mp_col].values[0]
                    generic_match = df_fretesmp[df_fretesmp['CODPP'] == 'XXX']
                    if not generic_match.empty:
                        return generic_match[mp_col].values[0]
                    return 99

                df['FreteProdVlr'] = df.apply(lambda row: -get_frete_pai(row), axis=1)
                df.drop(columns=['MP_2L'], inplace=True)

            df['Rebate'] = 0.0
            df['REPASSE'] = df['PMERC_T'] + df['ComissPctVlr'] + df['FreteFixoVlr'] + df['FreteProdVlr'] + df['Rebate']
            df['ImpLP'] = df.apply(
                lambda row: -0.0925 * row['PMERC_T'] if row['EMPRESA'] == 'K' else
                            -0.14   * row['PMERC_T'] if row['EMPRESA'] == 'A' else
                            -0.10   * row['PMERC_T'] if row['EMPRESA'] == 'B' else 0,
                axis=1)
            df['ImpICMS'] = df.apply(lambda row: -0.18 * row['PMERC_T'] if row['EMPRESA'] == 'K' else 0, axis=1)
            df['ImpTot'] = df['ImpLP'] + df['ImpICMS']
            df['MargVlr'] = df.apply(
                lambda row: 0 if row['EMPRESA'] == 'NC' else
                            row['REPASSE'] + row['ImpTot'] - row['ECTK'] - 1 - (0.01)*row['PMERC_T'] if row['EMPRESA'] == 'K' else
                            row['REPASSE'] + row['ImpTot'] - 1.6 * row['ECTK'],
                axis=1)
            df['MargPct'] = df['MargVlr'] / df['PMERC_T']

        elif key == 'MLA_Vendas':
            if 'N.º DE VENDA' in df.columns:
                df['N.º DE VENDA'] = df['N.º DE VENDA'].astype(str).str.strip()
                print("Changed N.º de venda to str. Sample values:", df['N.º DE VENDA'].head().tolist())
            if 'N.º DE VENDA_HYPERLINK' in df.columns:
                df['N.º DE VENDA_HYPERLINK'] = df['N.º DE VENDA_HYPERLINK'].astype(str).str.strip()
                print("Changed N.º de venda_hyperlink to str. Sample values:", df['N.º DE VENDA_HYPERLINK'].head().tolist())

            tot = None
            if 'VLRTOTALPSKU' in df.columns:
                tot = pd.to_numeric(df['VLRTOTALPSKU'], errors='coerce').fillna(0)
            elif 'VlrTotalpSKU' in df.columns:
                tot = pd.to_numeric(df['VlrTotalpSKU'], errors='coerce').fillna(0)
            else:
                tot = pd.Series(0, index=df.index)
            if 'VLRTOTALPSKU' not in df.columns and 'VlrTotalpSKU' in df.columns:
                df.rename(columns={'VlrTotalpSKU': 'VLRTOTALPSKU'}, inplace=True)

            df['Imposto1'] = tot * 0.11
            df['Imposto2'] = 0.0
            df['ImpostoT'] = df['Imposto1'] + df['Imposto2']

            cols_to_drop = ['CODPF_x', 'CODPF_y', 'MLSTATUS']
            df = df.drop([x for x in cols_to_drop if x in df.columns], axis=1)

        elif key == 'MLK_Vendas':
            df['ECTK'] = df['ECU'] * df['QTD']
            df['Imposto1'] = df['VLRTOTALPSKU']*(0.0925)
            df['Imposto2'] = df['VLRTOTALPSKU']*(0.18)
            df['ImpostoT'] =  df['Imposto1'] + df['Imposto2']
            df['MARGVLR'] = df['REPASSE'] - df['ImpostoT'] - df['ECTK'] - (1) -(.01)*df['VLRTOTALPSKU']
            df['MARGPCT'] = df['MARGVLR'] / df['VLRTOTALPSKU']
            cols_to_drop = ['CODPF_x', 'CODPF_y', 'MLSTATUS']
            df = df.drop([x for x in cols_to_drop if x in df.columns], axis=1)

        elif key == 'O_CtasARec':
            df['DATA BASE'] = pd.to_datetime(df['ANOMES'], format='%y%m') + MonthEnd(0)
            df['DIAS ATRASO'] = (df['DATA BASE'] - df['VENCIMENTO']).dt.days
            df['DIAS ATRASO'] = df['DIAS ATRASO'].apply(lambda x: max(0, x))
            df_ctas_a_rec_class = all_data['T_CtasARecClass']
            def classify_dias_atraso(row):
                match = df_ctas_a_rec_class[
                    (df_ctas_a_rec_class['DEXDIAS'] <= row['DIAS ATRASO']) &
                    (row['DIAS ATRASO'] <= df_ctas_a_rec_class['ATEXDIAS'])
                ]
                return match['STATUS ATRASO'].iloc[0] if not match.empty else None
            df['CLASSIFICACAO'] = df.apply(classify_dias_atraso, axis=1)

        df = clean_dataframes({key: df})[key]
        all_data[key] = df

    return all_data

def preprocess_inventory_data(file_path):
    sheets = pd.read_excel(file_path, sheet_name=None, header=1)  # Load data with headers from the second row
    processed_sheets = {}

    for sheet_name, df in sheets.items():
        df = df.melt(id_vars=['CodPF'], var_name='ANOMES', value_name='Value')
        df['ANOMES'] = pd.to_datetime(df['ANOMES'], errors='coerce').dt.strftime('%y%m')
        processed_sheets[sheet_name] = df
    
    return processed_sheets

def merge_data(all_data, df1_name, df1_col, df2_name, df2_col, new_col=None, indicator_name=None, default_value=None):
    df1_col = df1_col.upper()
    df2_col = df2_col.upper()
    if new_col:
        new_col = new_col.upper()

    # Verifica se os dataframes existem
    if df1_name not in all_data or df2_name not in all_data:
        print(f"❌ Dataframe '{df1_name}' or '{df2_name}' not found in all_data.")
        return all_data

    df1 = all_data[df1_name]
    df2 = all_data[df2_name]

    # Verifica se os dataframes não estão vazios
    if df1.empty or df2.empty:
        print(f"❌ Dataframe '{df1_name}' or '{df2_name}' is empty. Merge skipped.")
        return all_data

    # Verifica se as colunas existem
    if df1_col not in df1.columns or df2_col not in df2.columns:
        print(f"❌ Column '{df1_col}' or '{df2_col}' not found in dataframes '{df1_name}' or '{df2_name}'. Merge skipped.")
        return all_data

    print(f"Columns in {df1_name} BEFORE merge: {df1.columns.tolist()}")

    # Drop potential conflicting columns
    cols_to_drop = []
    if new_col and new_col in df1.columns:
        cols_to_drop.append(new_col)
    if df2_col != df1_col and df2_col in df1.columns:
        cols_to_drop.append(df2_col)
    if cols_to_drop:
        print(f"⚠️  Dropping existing columns {cols_to_drop} from '{df1_name}' before merge.")
        all_data[df1_name] = all_data[df1_name].drop(columns=cols_to_drop)
        df1 = all_data[df1_name]

        # --- Ensure dtype alignment for common numeric-like ids (rollback behavior) ---
        # Note: CodPed is the new name from process_data.py (was CÓDIGO PEDIDO)
        for colname, frame in ((df1_col, df1), (df2_col, df2)):
            if colname in frame.columns and colname in ['CodPed', 'CÓDIGO PEDIDO', 'N.º DE VENDA']:  # Support both old and new names
                frame[colname] = frame[colname].astype(str).str.strip()

    # Faz o merge
    merged = df1.merge(
        df2[[df2_col] + ([new_col] if new_col else [])],
        left_on=df1_col,
        right_on=df2_col,
        how='left',
        indicator=indicator_name is not None
    )

    # Adiciona coluna com valor default se merge falhar
    if new_col and default_value is not None:
        merged[new_col] = merged[new_col].fillna(default_value)

    if indicator_name:
        merged[indicator_name] = merged[indicator_name].fillna('no_match')

    all_data[df1_name] = merged
    print(f"✅ Merge applied: {df1_name} ← {df2_name} by '{df1_col}'")
    return all_data

def merge_data2v(all_data, df1_name, df1_col1, df1_col2, df2_name, df2_col1, df2_col2, df2_val_col, new_col_name, default_value=None, negative=False):
    df1_col1 = df1_col1.upper()
    df1_col2 = df1_col2.upper()
    df2_col1 = df2_col1.upper()
    df2_col2 = df2_col2.upper()
    df2_val_col = df2_val_col.upper()
    new_col_name = new_col_name.upper()

    if df1_name in all_data and df2_name in all_data:
        df1 = all_data[df1_name]
        df2 = all_data[df2_name]

        # Standardize column names
        df1.columns = [col.upper() for col in df1.columns]
        df2.columns = [col.upper() for col in df2.columns]

        if df1_col1 not in df1.columns or df1_col2 not in df1.columns or df2_col1 not in df2.columns or df2_col2 not in df2.columns:
            raise KeyError(f"One of the columns '{df1_col1}', '{df1_col2}', '{df2_col1}', '{df2_col2}' not found in dataframes.")

        if negative:
            df2[df2_val_col] = df2[df2_val_col] * -1  # Make the VALUE column negative

        df2_cols = [df2_col1, df2_col2, df2_val_col]

        # Remove colunas que já existam com o mesmo nome da futura 'new_col_name'
        cols_to_drop = [col for col in df1.columns if col.upper().startswith(new_col_name)]
        if cols_to_drop:
            print(f"⚠️  Dropping existing columns {cols_to_drop} from '{df1_name}' before merge_data2v.")
            df1 = df1.drop(columns=cols_to_drop)

        merged_df = df1.merge(
            df2[df2_cols].drop_duplicates(),
            left_on=[df1_col1, df1_col2],
            right_on=[df2_col1, df2_col2],
            how='left'
        )

        if df2_val_col and default_value is not None:
            merged_df[df2_val_col] = merged_df[df2_val_col].fillna(default_value)

        # Rename the value column to the new column name
        merged_df.rename(columns={df2_val_col: new_col_name}, inplace=True)

        all_data[df1_name] = merged_df
    return all_data

def merge_data_lastcost(all_data, df1_name, df1_product_col, df1_date_col, df2_name, df2_product_col, df2_date_col, df2_cost_col, new_col_name, default_value=None):
    """
    Merge df1 (sales table) with df2 (cost table) to get the last recorded cost before the sale date.

    Parameters:
    - all_data (dict): Dictionary containing all datasets.
    - df1_name (str): Key for the main table (sales) in all_data.
    - df1_product_col (str): Column name for the product code in the main table.
    - df1_date_col (str): Column name for the sale date in the main table.
    - df2_name (str): Key for the cost table in all_data.
    - df2_product_col (str): Column name for the product code in the cost table.
    - df2_date_col (str): Column name for the purchase date in the cost table.
    - df2_cost_col (str): Column name for the cost value in the cost table.
    - new_col_name (str): Name of the new column to store the retrieved cost.
    - default_value (optional): Default value to use if no match is found.

    Returns:
    - all_data (dict): Updated dictionary with the main table modified to include last cost.
    """

    # Standardize column names to uppercase
    df1_product_col = df1_product_col.upper()
    df1_date_col = df1_date_col.upper()
    df2_product_col = df2_product_col.upper()
    df2_date_col = df2_date_col.upper()
    df2_cost_col = df2_cost_col.upper()
    new_col_name = new_col_name.upper()
    
    if df1_name in all_data and df2_name in all_data:
        df1 = all_data[df1_name]
        df2 = all_data[df2_name]

        # Standardize column names
        df1.columns = [col.upper() for col in df1.columns]
        df2.columns = [col.upper() for col in df2.columns]

        # Check if required columns exist
        missing_cols = [col for col in [df1_product_col, df1_date_col] if col not in df1.columns] + \
                       [col for col in [df2_product_col, df2_date_col, df2_cost_col] if col not in df2.columns]
        if missing_cols:
            raise KeyError(f"Missing columns in dataframes: {missing_cols}")

        # Convert dates to datetime format
        df1[df1_date_col] = pd.to_datetime(df1[df1_date_col])
        df2[df2_date_col] = pd.to_datetime(df2[df2_date_col])

        # Sort df2 (cost table) by product and date descending
        df2 = df2.sort_values(by=[df2_product_col, df2_date_col], ascending=[True, False])

        # Merge based on product and latest entry before sale date
        def get_last_cost(row):
            product = row[df1_product_col]
            sale_date = row[df1_date_col]

            # Filter cost table for matching product and valid entry dates
            valid_costs = df2[(df2[df2_product_col] == product) & (df2[df2_date_col] <= sale_date)]

            # Return the most recent cost before the sale date
            return valid_costs[df2_cost_col].iloc[0] if not valid_costs.empty else default_value

        df1[new_col_name] = df1.apply(get_last_cost, axis=1)

        # Update the dataset in all_data dictionary
        all_data[df1_name] = df1

    return all_data

def merge_data_sum(all_data, df1_name, df1_col, df2_name, df2_col, new_col, indicator_name=None, default_value=0):
    """
    Merges two datasets, summing up multiple occurrences of the same key in df2 before merging.
    
    Parameters:
    - all_data: Dictionary containing dataframes
    - df1_name: Name of the first dataframe in all_data (target dataframe)
    - df1_col: Column in df1 to match on
    - df2_name: Name of the second dataframe in all_data (source dataframe)
    - df2_col: Column in df2 to match on
    - new_col: Column in df2 that should be summed before merging
    - indicator_name: Optional, column name to indicate merge status
    - default_value: Value to use for missing matches (default = 0)
    
    Returns:
    - Updated all_data dictionary with the merged dataframe
    """
    df1_col = df1_col.upper()
    df2_col = df2_col.upper()
    new_col = new_col.upper()

    # ✅ Check if both dataframes exist
    if df1_name not in all_data:
        print(f"❌ Dataframe '{df1_name}' not found. Merge skipped.")
        return all_data
    if df2_name not in all_data:
        print(f"❌ Dataframe '{df2_name}' not found. Merge skipped.")
        return all_data

    df1 = all_data[df1_name]
    df2 = all_data[df2_name]

    # ✅ Check if source dataframe is empty (df2)
    if df2.empty:
        print(f"❌ Dataframe '{df2_name}' is empty. Merge skipped.")
        return all_data

    # Standardize column names
    df1.columns = [col.upper() for col in df1.columns]
    df2.columns = [col.upper() for col in df2.columns]

    if df1_col not in df1.columns or df2_col not in df2.columns or new_col not in df2.columns:
        raise KeyError(f"Column '{df1_col}', '{df2_col}', or '{new_col}' not found in dataframes.")

    # ✅ Aggregate df2 by summing new_col for each df2_col
    df2_agg = df2.groupby(df2_col, as_index=False)[new_col].sum()

    # ✅ Merge the summed values into df1
    merged_df = df1.merge(df2_agg, left_on=df1_col, right_on=df2_col, how='left', indicator=indicator_name, suffixes=('', '_DROP'))

    # Remove the '_DROP' columns
    merged_df.drop([col for col in merged_df.columns if col.endswith('_DROP')], axis=1, inplace=True)

    # ✅ Fill missing values with default_value
    merged_df[new_col] = merged_df[new_col].fillna(default_value)

    # ✅ If using an indicator, replace unmatched values
    if indicator_name:
        merged_df[indicator_name] = merged_df[indicator_name].apply(lambda x: default_value if x == 'left_only' else merged_df[new_col])
        merged_df.drop(columns=[new_col, indicator_name], inplace=True)

    all_data[df1_name] = merged_df
    
    return all_data

def compute_NFCI_ANOMES(all_data):
    for key, df in all_data.items():
        # Add the ANOMES column to O_NFCI
        # Note: Column renamed from 'EMISS' to 'Data' in process_data.py
        if key == 'O_NFCI' and 'Data' in df.columns:
            df['Data'] = pd.to_datetime(df['Data'], errors='coerce')  # Ensure the date is parsed correctly
            df['ANOMES'] = df['Data'].dt.strftime('%y%m')  # Format date as YYMM
            print(f"Added ANOMES column to {key}")
        all_data[key] = df
    return all_data

def compute_LPI_ANOMES(all_data):
    for key, df in all_data.items():
        # Add the ANOMES column to L_LPI
        if key == 'L_LPI' and 'DATA' in df.columns:
            df['DATA'] = pd.to_datetime(df['DATA'], errors='coerce')  # Ensure the date is parsed correctly
            df['ANOMES'] = df['DATA'].dt.strftime('%y%m')  # Format date as YYMM
            print(f"Added ANOMES column to {key}")
        all_data[key] = df
    return all_data

def compute_ML_ANOMES(all_data):
    for key, df in all_data.items():
        # Add the ANOMES column to MLA_Vendas and MLK_Vendas
        if key in ['MLA_Vendas', 'MLK_Vendas'] and 'DATA DA VENDA' in df.columns:
            # Use dateutil parser to parse the date string
            df['DATA DA VENDA'] = df['DATA DA VENDA'].apply(lambda x: parser.parse(x, fuzzy=True) if pd.notnull(x) else pd.NaT)
            df['ANOMES'] = df['DATA DA VENDA'].dt.strftime('%y%m')  # Format date as YYMM
            print(f"Added ANOMES column to {key}")
        all_data[key] = df
    return all_data

def compute_LPI_ANOMES(all_data):
    for key, df in all_data.items():
        # Add the ANOMES column to L_LPI
        if key == 'L_LPI' and 'DATA' in df.columns:
            df['DATA'] = pd.to_datetime(df['DATA'], errors='coerce')  # Ensure the date is parsed correctly
            df['ANOMES'] = df['DATA'].dt.strftime('%y%m')  # Format date as YYMM
            print(f"Added ANOMES column to {key}")
        all_data[key] = df
    return all_data

def compute_CC_ANOMES(all_data):
    for key, df in all_data.items():
        # Add the ANOMES column to L_LPI
        if key == 'O_CC' and 'DATA' in df.columns:
            df['DATA'] = pd.to_datetime(df['DATA'], errors='coerce')  # Ensure the date is parsed correctly
            df['ANOMES'] = df['DATA'].dt.strftime('%y%m')  # Format date as YYMM
            print(f"Added ANOMES column to {key}")
        all_data[key] = df
    return all_data

def mlcustom_date_parser(date_str):
    # Remove the 'hs.' part if it exists
    date_str = re.sub(r'\s*hs\.\s*$', '', date_str)
    
    # Replace the Portuguese month names with English month names
    month_map = {
        'JANEIRO': 'January', 'FEVEREIRO': 'February', 'MARÇO': 'March', 'ABRIL': 'April',
        'MAIO': 'May', 'JUNHO': 'June', 'JULHO': 'July', 'AGOSTO': 'August',
        'SETEMBRO': 'September', 'OUTUBRO': 'October', 'NOVEMBRO': 'November', 'DEZEMBRO': 'December'
    }
    
    for pt_month, en_month in month_map.items():
        date_str = date_str.replace(pt_month, en_month)
    
    # Parse the date
    return pd.to_datetime(date_str, format='%d DE %B DE %Y %H:%M HS.')

def compute_ML_ANOMES(all_data):
    for key, df in all_data.items():
        # Add the ANOMES column to MLA_Vendas and MLK_Vendas
        if key in ['MLA_Vendas', 'MLK_Vendas'] and 'DATA DA VENDA' in df.columns:
            # Use custom date parser to parse the date string
            df['DATA DA VENDA'] = df['DATA DA VENDA'].apply(lambda x: mlcustom_date_parser(x) if pd.notnull(x) else pd.NaT)
            df['ANOMES'] = df['DATA DA VENDA'].dt.strftime('%y%m')  # Format date as YYMM
            print(f"Added ANOMES column to {key}")
        all_data[key] = df
    return all_data

def load_inventory_data(file_path):
    return pd.read_excel(file_path)

def print_all_tables_and_columns(all_data):
    for table_name, df in all_data.items():
        #print(f"Table: {table_name}")
        #print("Columns:", df.columns.tolist())
        print("-" * 50)

def print_table_and_columns(all_data, table_name):
    if table_name in all_data:
        #print(f"Table: {table_name}")
        #print("Columns:", all_data[table_name].columns.tolist())
        print("-" * 50)
    else:
        print(f"Table '{table_name}' not found in the dataset.")

def print_table_head(all_data, table_name):
    """
    Print the column names and the first 10 rows of a DataFrame.

    Parameters:
    df (pandas.DataFrame): The DataFrame to print.
    """
    #print("Columns:", all_data[table_name].columns.tolist())
    print("\nTop 10 Rows:")
    #print(all_data[table_name].head(10))

def excel_format(output_file, column_format_dict):
    # Load the workbook with macros
    wb = load_workbook(output_file, keep_vba=True)

    # Predefine header style if not already added
    if "header_style" not in wb.named_styles:
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True)
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        wb.add_named_style(header_style)  # Add only if it doesn't exist

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"✅ Formatting sheet: {sheet_name}")

        # Apply column formatting first
        if sheet_name in column_format_dict:
            format_dict = column_format_dict[sheet_name]

            for col in ws.iter_cols():
                col_letter = col[0].column_letter  # Get column letter
                col_name = col[0].value  # Get header name

                if col_name in format_dict:
                    format_code = format_dict[col_name]

                    # Apply format to the entire column including header
                    for cell in col:
                        cell.number_format = format_code

        # Apply header styling separately
        for cell in ws[1]:  # First row (header)
            cell.style = "header_style"

    # **CRITICAL**: Save the workbook while keeping macros
    try:
        wb.save(output_file)
        print(f"✅ Successfully formatted and saved {output_file}")
    except Exception as e:
        print(f"❌ Error saving {output_file}: {e}")

def excel_autofilters(output_path):
    print(f"✅ Adding auto-filters to {output_path}")
    
    # Open workbook while preserving macros
    workbook = load_workbook(output_path, keep_vba=True)

    for sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]
        print(f"🔹 Processing sheet: {sheetname}")

        # Check if the worksheet has any data
        if worksheet.max_row > 1 and worksheet.max_column > 1:
            data_range = worksheet.dimensions
            if data_range and ":" in data_range:  # Ensure valid range
                worksheet.auto_filter.ref = data_range
                print(f"✅ Auto-filter applied to {sheetname}: {data_range}")
            else:
                print(f"⚠️ Skipping {sheetname}: No valid data range found.")
        else:
            print(f"⚠️ Skipping {sheetname}: Sheet is empty or has only headers.")

    # Save changes while keeping macros
    try:
        workbook.save(output_path)
        print(f"✅ Auto-filters added and saved: {output_path}")
    except Exception as e:
        print(f"❌ Error saving {output_path}: {e}")

# Define the audit function
def perform_audit(df, client_name):
    # Note: Using new column names from process_data.py
    audit_columns = [
        'CODPF',
        'Qt',  # was QTD
        'PMerc_U',  # was PRECO CALC
        'PMerc_T',  # was MERCVLR
        'A_ICMSST_T',  # was ICMSST
        'A_IPI_T',  # was IPI
        'PNF_T',  # was TOTALNF
        'Data']  # was EMISS

    audit_df = df[df['NOMEF'] == client_name][audit_columns]  
    return audit_df

def AuditMP_SH(all_data, mp2, empresa):
    # Note: Using new column names from process_data.py
    lpi_columns = [
        'CodPed',  # was CÓDIGO PEDIDO
        'EMPRESA',
        'MP',
        'MP2',
        'Status',  # was STATUS PEDIDO
        'CODPP',
        'PMerc_T',  # was VLRVENDA
        'Qt',  # was QTD
        'REPASSE'
    ]

    sh_columns = [
        'ID DO PEDIDO',
        'VALOR'
    ]
    
    # Select and filter L_LPI data
    dfa = all_data['L_LPI'][lpi_columns].copy()
    dfa = dfa[(dfa["MP2"] == mp2) & (dfa["EMPRESA"] == empresa)]

    if dfa.empty:
        print(f"Warning: No matching data found for MP2={mp2} and EMPRESA={empresa}")

    # Rename columns
    # Note: PMerc_T is the new name from process_data.py, renaming to VENDATOTAL for audit
    rename_map = {
        'PMerc_T': 'VENDATOTAL',  # was VLRVENDA
        'REPASSE': 'REPASSEESPERADO_TODOSPEDIDOS'
    }
    dfa.rename(columns=rename_map, inplace=True)

    # Store in all_data dictionary
    all_data[f'Aud_{mp2}'] = dfa

    # Merge with SHK_Extrato
    # Note: CodPed is the new name from process_data.py
    all_data = merge_data(all_data, f'Aud_{mp2}', "CodPed", "SHK_Extrato", "ID DO PEDIDO", "VALOR", default_value=0)

    # Ensure columns exist before renaming
    if 'VALOR' in all_data[f'Aud_{mp2}'].columns:
        all_data[f'Aud_{mp2}'].rename(columns={'VALOR': 'REPASSEEFETIVO_PEDIDOSPAGOS'}, inplace=True)
    else:
        print(f"Warning: 'VALOR' column not found after merge in Aud_{mp2}")

    # Check for missing columns before applying the lambda function
    required_columns = ['REPASSEESPERADO_TODOSPEDIDOS', 'REPASSEEFETIVO_PEDIDOSPAGOS']
    missing_columns = [col for col in required_columns if col not in all_data[f'Aud_{mp2}'].columns]

    if missing_columns:
        print(f"Error: Missing columns {missing_columns} in Aud_{mp2}")
    else:
        all_data[f'Aud_{mp2}']['REPASSEESPERADO_PEDIDOSPAGOS'] = all_data[f'Aud_{mp2}'].apply(
            lambda row: 0 if row['REPASSEEFETIVO_PEDIDOSPAGOS'] == 0 else row['REPASSEESPERADO_TODOSPEDIDOS'], axis=1
        )

    return all_data

def AuditMP_ML(all_data, mp2, empresa):
    # Note: Using new column names from process_data.py
    lpi_columns = [
        'CodPed',  # was CÓDIGO PEDIDO
        'EMPRESA',
        'MP',
        'MP2',
        'Status',  # was STATUS PEDIDO
        'CODPP',
        'PMerc_T',  # was VLRVENDA
        'Qt',  # was QTD
        'REPASSE'
    ]

    sh_columns = [
        'ORDER_ID',
        'NETVALUE',
        'DESC'
    ]
    
    # Select and filter L_LPI data
    dfa = all_data['L_LPI'][lpi_columns].copy()
    dfa = dfa[(dfa["MP2"] == mp2) & (dfa["EMPRESA"] == empresa)]

    if dfa.empty:
        print(f"Warning: No matching data found for MP2={mp2} and EMPRESA={empresa}")

    # Rename columns
    # Note: PMerc_T is the new name from process_data.py, renaming to VENDATOTAL for audit
    rename_map = {
        'PMerc_T': 'VENDATOTAL',  # was VLRVENDA
        'REPASSE': 'REPASSEESPERADO_TODOSPEDIDOS'
    }
    dfa.rename(columns=rename_map, inplace=True)

    # Store in all_data dictionary
    all_data[f'Aud_{mp2}'] = dfa

    # Merge with MLK_ExtLib
    # Note: CodPed is the new name from process_data.py
    all_data = merge_data_sum(all_data, f'Aud_{mp2}', "CodPed", "MLK_ExtLib", "ORDER_ID", "NETVALUE", default_value=0)

    # Ensure columns exist before renaming
    if 'NETVALUE' in all_data[f'Aud_{mp2}'].columns:
        all_data[f'Aud_{mp2}'].rename(columns={'NETVALUE': 'REPASSEEFETIVO_PEDIDOSPAGOS'}, inplace=True)
    else:
        print(f"Warning: 'VALOR' column not found after merge in Aud_{mp2}")

    # Check for missing columns before applying the lambda function
    required_columns = ['REPASSEESPERADO_TODOSPEDIDOS', 'REPASSEEFETIVO_PEDIDOSPAGOS']
    missing_columns = [col for col in required_columns if col not in all_data[f'Aud_{mp2}'].columns]

    if missing_columns:
        print(f"Error: Missing columns {missing_columns} in Aud_{mp2}")
    else:
        all_data[f'Aud_{mp2}']['REPASSEESPERADO_PEDIDOSPAGOS'] = all_data[f'Aud_{mp2}'].apply(
            lambda row: 0 if row['REPASSEEFETIVO_PEDIDOSPAGOS'] == 0 else row['REPASSEESPERADO_TODOSPEDIDOS'], axis=1
        )

    return all_data

def AuditMP_MA(all_data, mp2, empresa):
    # Note: Using new column names from process_data.py
    lpi_columns = [
        'CodPed',  # was CÓDIGO PEDIDO
        'EMPRESA',
        'MP',
        'MP2',
        'Status',  # was STATUS PEDIDO
        'CODPP',
        'PMerc_T',  # was VLRVENDA
        'Qt',  # was QTD
        'REPASSE'
    ]

    sh_columns = [
        'NÚMERO DO PEDIDO',
        'VALOR LÍQUIDO ESTIMADO A RECEBER (****)',
        'DESC'
    ]
    
    # Select and filter L_LPI data
    dfa = all_data['L_LPI'][lpi_columns].copy()
    dfa = dfa[(dfa["MP2"] == mp2) & (dfa["EMPRESA"] == empresa)]

    if dfa.empty:
        print(f"Warning: No matching data found for MP2={mp2} and EMPRESA={empresa}")

    # Rename columns
    # Note: PMerc_T is the new name from process_data.py, renaming to VENDATOTAL for audit
    rename_map = {
        'PMerc_T': 'VENDATOTAL',  # was VLRVENDA
        'REPASSE': 'REPASSEESPERADO_TODOSPEDIDOS'
    }
    dfa.rename(columns=rename_map, inplace=True)

    # Store in all_data dictionary
    all_data[f'Aud_{mp2}'] = dfa

    # Merge with MGK_Extrato
    # Note: CodPed is the new name from process_data.py
    all_data = merge_data_sum(all_data, f'Aud_{mp2}', "CodPed", "MGK_Extrato", "NÚMERO DO PEDIDO", "VALOR LÍQUIDO ESTIMADO A RECEBER (****)", default_value=0)

    # Ensure columns exist before renaming
    if 'VALOR LÍQUIDO ESTIMADO A RECEBER (****)' in all_data[f'Aud_{mp2}'].columns:
        all_data[f'Aud_{mp2}'].rename(columns={'VALOR LÍQUIDO ESTIMADO A RECEBER (****)': 'REPASSEEFETIVO_PEDIDOSPAGOS'}, inplace=True)
    else:
        print(f"Warning: 'VALOR' column not found after merge in Aud_{mp2}")

    # Check for missing columns before applying the lambda function
    required_columns = ['REPASSEESPERADO_TODOSPEDIDOS', 'REPASSEEFETIVO_PEDIDOSPAGOS']
    missing_columns = [col for col in required_columns if col not in all_data[f'Aud_{mp2}'].columns]

    if missing_columns:
        print(f"Error: Missing columns {missing_columns} in Aud_{mp2}")
    else:
        all_data[f'Aud_{mp2}']['REPASSEESPERADO_PEDIDOSPAGOS'] = all_data[f'Aud_{mp2}'].apply(
            lambda row: 0 if row['REPASSEEFETIVO_PEDIDOSPAGOS'] == 0 else row['REPASSEESPERADO_TODOSPEDIDOS'], axis=1
        )

    return all_data


# Define the function to perform audits for all specified clients
def perform_all_MP_audits(all_data):
    all_data = AuditMP_SH(all_data, 'SH','K')
    all_data = AuditMP_ML(all_data, 'ML','K')
    all_data = AuditMP_MA(all_data, 'MA','K')
    return all_data


# Define the function to perform audits for all specified clients
def perform_all_audits(all_data):
    for client_name in audit_client_names:
        audit_df = perform_audit(all_data['O_NFCI'], client_name)
        all_data[f'Audit_{client_name}'] = audit_df
        print(f"Performed audit for {client_name}")  # Debug print
    return all_data

# Function to track inventory movements
# Function to calculate realized cost
def track_inventory(sales_data, purchase_data):
    inventory_movements = []

    # Process sales data
    for index, row in sales_data.iterrows():
        movement = {
            'Date': row['DATA'],
            'Invoice Number': None,
            'Product Code': row['CODPF'],
            'Quantity': -row['QTD'],
            'CV': 'V',
            'QTD E': row['QTD'],
            'CMV Unit E': None,
            'CMV Mov E': None,
            'QTD R': None,
            'CMV Unit R': None,
            'CMV Mov R': None,
            'NF Compra': None
        }
        inventory_movements.append(movement)

    # Process purchase data
    for index, row in purchase_data.iterrows():
        movement = {
            # Note: Column renamed from 'EMISS' to 'Data' in process_data.py
            'Date': row.get('Data') if 'Data' in purchase_data.columns else row.get('EMISS'),
            'Invoice Number': row['NF'],
            'Product Code': row['CODPF'],
            # Note: Column renamed from 'QTD' to 'Qt' in process_data.py
            'Quantity': row.get('Qt') if 'Qt' in purchase_data.columns else row.get('QTD'),
            'CV': 'C',
            'QTD E': None,
            # Note: Column renamed from 'PRECO CALC' to 'PMerc_U' in process_data.py
            'CMV Unit E': row.get('PMerc_U') if 'PMerc_U' in purchase_data.columns else row.get('PRECO CALC'),
            # Note: Column renamed from 'MERCVLR' to 'PMerc_T' in process_data.py
            'CMV Mov E': row.get('PMerc_T') if 'PMerc_T' in purchase_data.columns else row.get('MERCVLR'),
            'QTD R': None,
            'CMV Unit R': None,
            'CMV Mov R': None,
            'NF Compra': row['NF'],
            # Note: Columns renamed: TOTALNF→PNF_T, QTD→Qt in process_data.py
            'Custo Total Unit': (row.get('PNF_T') if 'PNF_T' in purchase_data.columns else row.get('TOTALNF')) / (row.get('Qt') if 'Qt' in purchase_data.columns else row.get('QTD')) if (row.get('Qt') if 'Qt' in purchase_data.columns else row.get('QTD')) != 0 else 0
        }
        inventory_movements.append(movement)

    inventory_df = pd.DataFrame(inventory_movements)
    inventory_df.sort_values(by='Date', inplace=True)
    return inventory_df

def calculate_realized_cost(inventory_df):
    # Get all purchases (C) in ascending date order
    purchase_data = inventory_df[inventory_df['CV'] == 'C'].sort_values(by='Date')

    # Create a list of purchases as objects with necessary details
    purchase_list = []
    for _, row in purchase_data.iterrows():
        purchase_list.append({
            'Product Code': row['Product Code'],
            'Invoice Number': row['Invoice Number'],
            'Quantity': row['Quantity'],
            'Custo Total Unit': row['Custo Total Unit']
        })
    #print("Purchase List:", purchase_list)  # Debug print

    # Iterate through the sales (V) and populate the realized cost details
    for index, row in inventory_df[inventory_df['CV'] == 'V'].iterrows():
        quantity_needed = -row['Quantity']
        for purchase in purchase_list:
            if purchase['Product Code'] == row['Product Code'] and quantity_needed > 0:
                if purchase['Quantity'] > 0:
                    quantity_to_apply = min(purchase['Quantity'], quantity_needed)

                    # Update the realized cost details
                    inventory_df.at[index, 'QTD R'] = quantity_to_apply
                    inventory_df.at[index, 'CMV Unit R'] = purchase['Custo Total Unit']
                    inventory_df.at[index, 'CMV Mov R'] = quantity_to_apply * purchase['Custo Total Unit']
                    inventory_df.at[index, 'NF Compra'] = purchase['Invoice Number']

                    # Update the purchase details
                    purchase['Quantity'] -= quantity_to_apply
                    quantity_needed -= quantity_to_apply

        # If there's still quantity needed, populate the expected cost details
        if quantity_needed > 0:
            inventory_df.at[index, 'QTD E'] = quantity_needed
            if row['CMV Unit E'] is not None:
                inventory_df.at[index, 'CMV Mov E'] = quantity_needed * row['CMV Unit E']

    # Add remaining purchase quantities back to the corresponding purchase rows
    for purchase in purchase_list:
        if purchase['Quantity'] > 0:
            purchase_indices = inventory_df[
                (inventory_df['Product Code'] == purchase['Product Code']) &
                (inventory_df['Invoice Number'] == purchase['Invoice Number'])
            ].index
            if not purchase_indices.empty:
                purchase_index = purchase_indices[0]
                inventory_df.at[purchase_index, 'QTD E'] = purchase['Quantity']
                inventory_df.at[purchase_index, 'CMV Unit E'] = purchase['Custo Total Unit']
                inventory_df.at[purchase_index, 'CMV Mov E'] = purchase['Quantity'] * purchase['Custo Total Unit']

    #print(inventory_df)  # Debug print
    return inventory_df

# Function to perform the inventory audit
def perform_invaudit(o_nfci_df, l_lpi_df, client_name):
    sales_data = l_lpi_df[l_lpi_df['EMPRESAF'] == client_name]
    purchase_data = o_nfci_df[o_nfci_df['NOMEF'] == client_name]

    inventory_df = track_inventory(sales_data, purchase_data)
    inventory_df = calculate_realized_cost(inventory_df)
    return inventory_df

# Function to perform all inventory audits
def perform_all_invaudits(all_data):
    o_nfci_df = all_data['O_NFCI']
    l_lpi_df = all_data['L_LPI']

    invaudit_results = {}
    for client in invaudit_client_names:
        invaudit_results[client] = perform_invaudit(o_nfci_df, l_lpi_df, client)

    # Add the audit results to the all_data dictionary
    for client, df in invaudit_results.items():
        all_data[f'InvAudit_{client}'] = df

    return all_data

def Kon_FixSums(all_data):

    df = all_data["Kon_RelGeral"].copy()
    saldos_raw = all_data["T_SaldosCC"].copy()

    # -------------------------------------------------------------
    # 1. REMOVE EXISTING SALDO LINES
    # -------------------------------------------------------------
    df = df[df["KON_GR"].astype(str).str.upper() != "SALDO"]

    # -------------------------------------------------------------
    # 2. PREPARE KON DATA
    # -------------------------------------------------------------
    df["VALOR_REPASSE"] = pd.to_numeric(df["VALOR_REPASSE"], errors="coerce").fillna(0)
    current_anomes = df["ANOMES"].max()

    # -------------------------------------------------------------
    # 3. RESHAPE T_SaldosCC (wide → long)
    # -------------------------------------------------------------
    saldos = saldos_raw.melt(
        id_vars=["CONTA"],
        var_name="ANOMES",
        value_name="SALDO"
    )

    saldos["ANOMES"] = pd.to_numeric(saldos["ANOMES"], errors="coerce")
    saldos["SALDO"] = pd.to_numeric(saldos["SALDO"], errors="coerce").fillna(0)

    # -------------------------------------------------------------
    # 4. WORK ONLY WITH MARKETPLACES
    # -------------------------------------------------------------
    canais_validos = ["MERCADO LIVRE", "SHOPEE", "AMAZON", "MAGAZINE LUIZA"]

    saldos = saldos[saldos["CONTA"].str.upper().isin([c.upper() for c in canais_validos])]

    # map table names → channel names in Kon_RelGeral
    canal_map = {
        "MERCADO LIVRE": "MERCADO LIVRE",
        "SHOPEE": "SHOPEE",
        "AMAZON": "AMAZON",
        "MAGAZINE LUIZA": "MAGAZINE LUIZA"
    }

    saldos["CANAL"] = saldos["CONTA"].str.upper().map(
        {k.upper(): v for k, v in canal_map.items()}
    )

    # -------------------------------------------------------------
    # 5. FIND PREVIOUS ANOMES
    # -------------------------------------------------------------
    prev_candidates = sorted(saldos["ANOMES"].unique())
    prev_candidates = [x for x in prev_candidates if x < current_anomes]
    prev_anomes = max(prev_candidates) if prev_candidates else None

    new_rows = []

    # -------------------------------------------------------------
    # 6. PROCESS EACH MARKETPLACE
    # -------------------------------------------------------------
    for canal in canal_map.values():

        # saldo atual
        atual_saldo = saldos[
            (saldos["CANAL"] == canal) &
            (saldos["ANOMES"] == current_anomes)
        ]["SALDO"].sum()

        # saldo anterior
        if prev_anomes:
            prev_saldo = saldos[
                (saldos["CANAL"] == canal) &
                (saldos["ANOMES"] == prev_anomes)
            ]["SALDO"].sum()
        else:
            prev_saldo = 0

        saldo_diff = atual_saldo - prev_saldo

        # soma de movimentos (Kon_RelGeral)
        movimento_sum = df[df["CANAL"] == canal]["VALOR_REPASSE"].sum()

        # quanto falta
        ajuste = saldo_diff - movimento_sum

        if abs(ajuste) < 0.01:
            continue

        # ---------------------------------------------------------
        # 7. CREATE THE AJUSTE ROW
        # ---------------------------------------------------------
        new_rows.append({
            "CONCILIACAO": "zzz",
            "CANAL": canal,
            "TP_LANCAMENTO": f"Ajuste de Saldo [{canal}]",
            "VALOR_REPASSE": ajuste,
            "CATEGORIA_LANCAMENTO": "AJUSTE",
            "ANOMES": current_anomes,
            "HASSKU": False,
            "KON_GR": "AJUSTE CC",
            "KON_SGR": "Ajuste de Saldo",
        })

    # -------------------------------------------------------------
    # 8. APPEND RESULTS
    # -------------------------------------------------------------
    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)

    all_data["Kon_RelGeral"] = df
    return all_data

def Kon_MatchTransfers(all_data):

    df_kon = all_data["Kon_RelGeral"].copy()
    df_cc  = all_data["O_CC"].copy()

    # -------------------------------------------------------------
    # PREP
    # -------------------------------------------------------------
    df_kon["VALOR_REPASSE"] = pd.to_numeric(df_kon["VALOR_REPASSE"], errors="coerce").fillna(0)
    df_cc["VALOR (R$)"] = pd.to_numeric(df_cc["VALOR (R$)"], errors="coerce").fillna(0)

    current_anomes = df_kon["ANOMES"].max()

    # mapping of channels → CC account name in O_CC
    cc_map = {
        "AMAZON": "AMAZON BR",
        "MAGAZINE LUIZA": "MAGALU PAY",
        "MERCADO LIVRE": "MERCADO PAGO",
        "SHOPEE": "SHOPEE",
    }

    new_rows = []

    # -------------------------------------------------------------
    # PROCESS EACH CHANNEL
    # -------------------------------------------------------------
    for canal, cc_account in cc_map.items():

        # 1. SAQUES from Kon_RelGeral
        kon_saque = df_kon[
            (df_kon["CANAL"].astype(str).str.upper() == canal) &
            (df_kon["KON_GR"].astype(str).str.upper() == "SAQUE")
        ]["VALOR_REPASSE"].sum()

        # 2. TRANSF from O_CC
        cc_transf = df_cc[
            (df_cc["CONTA CORRENTE"].astype(str).str.upper() == cc_account.upper()) &
            (df_cc["CC_CAT GRP"].astype(str).str.upper() == "TRANSF")
        ]["VALOR (R$)"].sum()

        # 3. DIFF
        diff = cc_transf - kon_saque

        if abs(diff) < 0.01:
            continue

        # 4. CREATE NEW ROW
        new_rows.append({
            "CONCILIACAO": "zzz",
            "CANAL": canal,
            "VALOR_REPASSE": diff,
            "TP_LANCAMENTO": "Diff de Saques",
            "CATEGORIA_LANCAMENTO": "Diff de Saques",
            "ANOMES": current_anomes,
            "HASSKU": False,
            "KON_GR": "AJUSTE SAQUE",
            "KON_SGR": "Ajuste de Saque por CC",
        })

    # -------------------------------------------------------------
    # APPEND NEW ROWS IF ANY
    # -------------------------------------------------------------
    if new_rows:
        df_kon = pd.concat([df_kon, pd.DataFrame(new_rows)], ignore_index=True)

    all_data["Kon_RelGeral"] = df_kon
    return all_data


def main(year: int, month: int):
    """
    Build R_Resumo for the selected (year, month).
    Expects a global `base_dir` already set.
    Fails fast if required inputs are missing (preferred over silent bad results).
    """
    import os, shutil
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    # ---- month tag & paths
    ano_mes = f"{year}_{month:02d}"
    print(f"Base directory set to: {base_dir}")

    template_file = os.path.join(base_dir, "Template", "PivotTemplate.xlsm")
    out_dir       = os.path.join(base_dir, "clean", ano_mes)
    os.makedirs(out_dir, exist_ok=True)
    output_file   = os.path.join(out_dir, f"R_Resumo_{ano_mes}.xlsm")

    # ---- copy & open template
    shutil.copy(template_file, output_file)
    print(f"✅ Copied template to {output_file}")
    print("✅ Opening template with macros...")
    wb_template = load_workbook(output_file, keep_vba=True)

    # remove all template sheets to start fresh
    for sh in wb_template.sheetnames[:]:
        del wb_template[sh]
    # for sh in wb_template.sheetnames[:]: # Removed
    #     del wb_template[sh] # Removed
    print(f"✅ Removed template sheets. Ready to write data...")

    # ----------------------------------------------------------------------
    # Load monthly CLEAN data
    # ----------------------------------------------------------------------
    file_patterns = {
        # REQUIRED clean files for a valid merge (fail-fast if missing):
        "O_NFCI":        "O_NFCI_{ym}_clean.xlsx",
        "L_LPI":         "L_LPI_{ym}_clean.xlsx",
        "MLK_Vendas":    "MLK_Vendas_{ym}_clean.xlsx",
        "O_CC":          "O_CC_{ym}_clean.xlsx",
        "O_CtasAPagar":  "O_CtasAPagar_{ym}_clean.xlsx",
        "O_CtasARec":    "O_CtasARec_{ym}_clean.xlsx",
        "Kon_RelGeral":  "Kon_RelGeral_{ym}_clean.xlsx",
        # OPTIONAL (load if present; absence will not abort):
        "MLA_Vendas":    "MLA_Vendas_{ym}_clean.xlsx",
        "MLK_ExtLib":    "MLK_ExtLib_{ym}_clean.xlsx",
        "SHK_Extrato":   "SHK_Extrato_{ym}_clean.xlsx",
        "MGK_Pacotes":   "MGK_Pacotes_{ym}_clean.xlsx",
        "MGK_Extrato":   "MGK_Extrato_{ym}_clean.xlsx",
    }
    required_clean_keys = [
        "O_NFCI", "L_LPI", "MLK_Vendas", "O_CC", "O_CtasAPagar", "O_CtasARec"
    ]

    all_data = {}
    missing_clean = []
    for key, pattern in file_patterns.items():
        fpath = os.path.join(base_dir, "clean", ano_mes, pattern.format(ym=ano_mes))
        if os.path.exists(fpath):
            try:
                df = pd.read_excel(fpath)
                print(f"{key} data shape: {df.shape}")
            except Exception as e:
                print(f"Error reading {fpath}: {e}")
                if key in required_clean_keys:
                    missing_clean.append(f"{key} -> {fpath} (read error)")
                df = pd.DataFrame()
        else:
            print(f"File not found: {fpath}")
            if key in required_clean_keys:
                missing_clean.append(f"{key} -> {fpath}")
            df = pd.DataFrame()

        all_data[key] = df

    if missing_clean:
        raise FileNotFoundError(
            "Required CLEAN inputs missing for this month:\n  - " + "\n  - ".join(missing_clean)
        )

    # ----------------------------------------------------------------------
    # Load STATIC lookup tables (required for correct merges; fail-fast)
    # ----------------------------------------------------------------------
    # static_dir = os.path.join(base_dir, "Tables") # This is now handled at the beginning of main
    static_files = {
        "T_CondPagto":         "T_CondPagto.xlsx",
        "T_Fretes":            "T_Fretes.xlsx",
        "T_GruposCli":         "T_GruposCli.xlsx",
        "T_MP":                "T_MP.xlsx",
        "T_RegrasMP":          "T_RegrasMP.xlsx",
        "T_Remessas":          "T_Remessas.xlsx",
        "T_Reps":              "T_Reps.xlsx",
        "T_Verbas":            "T_Verbas.xlsx",
        "T_Vol":               "T_Vol.xlsx",
        "T_ProdF":             "T_ProdF.xlsx",
        "T_ProdP":             "T_ProdP.xlsx",
        "T_Entradas":          "T_Entradas.xlsx",
        "T_FretesMP":          "T_FretesMP.xlsx",
        "T_MLStatus":          "T_MLStatus.xlsx",
        "T_CtasAPagarClass":   "T_CtasAPagarClass.xlsx",
        "T_CtasARecClass":     "T_CtasARecClass.xlsx",
        "T_CCCats":            "T_CCCats.xlsx",
        "T_KonCats":           "T_KonCats.xlsx",
        "T_SaldosCC":           "T_SaldosCC.xlsx",
    }
    missing_static = []
    for key, fname in static_files.items():
        spath = os.path.join(static_dir, fname)
        if os.path.exists(spath):
            try:
                sdf = pd.read_excel(spath)
                all_data[key] = sdf
                print(f"Static data {key} shape: {sdf.shape}")
            except Exception as e:
                print(f"Error reading static {spath}: {e}")
                missing_static.append(f"{key} -> {spath} (read error)")
        else:
            print(f"⚠️  Static file not found: {spath}")
            missing_static.append(f"{key} -> {spath}")

    if missing_static:
        raise FileNotFoundError(
            "Required STATIC tables missing:\n  - " + "\n  - ".join(missing_static)
        )

    # --- restore old behavior: normalize headers BEFORE merges ---
    all_data = rename_columns(all_data, column_rename_dict)

    # --- ROLLBACK: cast keys to string *after* header normalization (old behavior) ---
    if 'MLK_Vendas' in all_data and not all_data['MLK_Vendas'].empty:
        df_mlk = all_data['MLK_Vendas']
        if 'N.º DE VENDA' in df_mlk.columns:
            df_mlk['N.º DE VENDA'] = df_mlk['N.º DE VENDA'].astype(str).str.strip()
            print("Changed N.º de venda to str. Sample values:", df_mlk['N.º DE VENDA'].head().tolist())
        if 'N.º DE VENDA_HYPERLINK' in df_mlk.columns:
            df_mlk['N.º DE VENDA_HYPERLINK'] = df_mlk['N.º DE VENDA_HYPERLINK'].astype(str).str.strip()
            print("Changed N.º de venda_hyperlink to str. Sample values:", df_mlk['N.º DE VENDA_HYPERLINK'].head().tolist())
        all_data['MLK_Vendas'] = df_mlk

    # Also fix L_LPI keys if needed
    for k in ['L_LPI']:
        if k in all_data and not all_data[k].empty:
            df_any = all_data[k]
            if 'CodPed' in df_any.columns:
                all_data[k]['CodPed'] = all_data[k]['CodPed'].astype(str).str.strip()
            elif 'CÓDIGO PEDIDO' in df_any.columns:
                all_data[k]['CÓDIGO PEDIDO'] = all_data[k]['CÓDIGO PEDIDO'].astype(str).str.strip()
    # --- end rollback block ---

    for k, df_any in list(all_data.items()):
        # Note: Support both old and new column names
        if df_any is not None and not df_any.empty:
            if 'CodPed' in df_any.columns:
                all_data[k]['CodPed'] = all_data[k]['CodPed'].astype(str).str.strip()
            elif 'CÓDIGO PEDIDO' in df_any.columns:
                all_data[k]['CÓDIGO PEDIDO'] = all_data[k]['CÓDIGO PEDIDO'].astype(str).str.strip()
    # --- end rollback block ---

    # ----------------------------------------------------------------------
    # Merge + calculated columns (let it crash if something is structurally wrong)
    # ----------------------------------------------------------------------
    print("Creating Merged and Calculated Columns")
    all_data = merge_all_data(all_data)

    # --- Split lines with multiple SKUs (before unmapped and ECU calc) ---
    #all_data = debug_df(all_data, "Kon_RelGeral", "AAAAA")
    df = all_data["Kon_RelGeral"].copy()
    df = split_SKU_lines(df)
    all_data["Kon_RelGeral"] = df
    #all_data = debug_df(all_data, "Kon_RelGeral", "BBBBB")

    # Kon_RelGeral joins with T_ProdF to get CODPP from SKU = CODPF
    all_data = merge_data(all_data,
        df1_name="Kon_RelGeral",
        df1_col="CODPF",
        df2_name="T_ProdF",
        df2_col="CODPF",
        new_col="CODPP",
        default_value=""
    )
    #all_data = debug_df(all_data, "Kon_RelGeral", "CCCCC")

    # --- Add synthetic UNMAPPED SKUs per channel ---
    #all_data = add_unmapped_skus(all_data)
    #all_data = debug_df(all_data, "Kon_RelGeral", "DELTA")

    # --- Add last cost (ECU) lookup directly on Kon_RelGeral ---
    all_data = merge_data_lastcost(
        all_data,
        df1_name="Kon_RelGeral",      # or "Kon_Geral" if that’s your real key
        df1_product_col="CODPP",
        df1_date_col="DATA_PEDIDO",  # or your actual sale date col
        df2_name="T_Entradas",
        df2_product_col="PAI",
        df2_date_col="ULTIMA ENTRADA",
        df2_cost_col="ULT CU R$",
        new_col_name="ECU",
        default_value=0
    )
    #all_data = debug_df(all_data, "Kon_RelGeral", "ECO")
    all_data = split_SKU_lines_by_cost_ratio(all_data)
    #all_data = debug_df(all_data, "Kon_RelGeral", "FOX")
    all_data = Kon_MatchTransfers(all_data)
    all_data = Kon_FixSums(all_data)

    # --- Build Kon summary ---
    #all_data = build_Kon_Report1(all_data)
    #all_data = compute_channel_ratios(all_data)
    all_data = build_all_ratio_versions(all_data)
    all_data = build_Kon_Final_Report(all_data)

    all_data = build_Kon_Detail_SKUAdj(all_data)


    # ----------------------------------------------------------------------
    # Write each dataframe to the workbook
    # ----------------------------------------------------------------------
    # --- Define list of static tables to exclude ---
    excluded_sheets = {
        "T_CondPagto", "T_Fretes", "T_GruposCli", "T_MP", "T_RegrasMP",
        "T_Remessas", "T_Reps", "T_Verbas", "T_Vol", "T_ProdF", "T_ProdP",
        "T_Entradas", "T_FretesMP", "T_MLStatus", "T_CtasAPagarClass",
        "T_CtasARecClass", "T_CCCats", "Kon_Report1", "MLK_ExtLib", "SHK_Extrato", "MGK_Pacotes", 
        "MGK_Extrato", "MLA_Vendas" 
    }

    # Prepare output file
    output_file = os.path.join(base_dir, "clean", ano_mes, f"Kon_Report_{ano_mes}.xlsx")
    wb_template = Workbook()
    
    # --- Write each dataframe except excluded ones ---
    for key, df in all_data.items():
        if key in excluded_sheets:
            print(f"⏩ Skipped static sheet: {key}")
            continue
        sheet_name = (key or "Sheet")[:31]  # Excel sheet name limit
        ws = wb_template.create_sheet(title=sheet_name)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        print(f"✅ Added {key} to workbook")

    # Remove default sheet if empty
    if "Sheet" in wb_template.sheetnames and len(wb_template.sheetnames) > 1:
        del wb_template["Sheet"]

    wb_template.save(output_file)
    print(f"✅ All merged data saved to {output_file}")

    # Optional formatting (only if your helpers are defined)
    try:
        #excel_format(output_file, column_format_dict)
        #excel_autofilters(output_file)
        print(f"✅ Skipped Formatting and autofilters")
    except NameError:
        pass


if __name__ == "__main__":
    # use the month picked at the top (ano_x/mes_x) — identical to Conc_Estoque behavior
    # main(ano_x, mes_x) # Original call
    main() # New call
