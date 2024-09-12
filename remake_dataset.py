import pandas as pd
from pandas.tseries.offsets import MonthEnd
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
import re

# Define the potential base directories
path_options = [
    '/Users/mauricioalouan/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data/',
    '/Users/simon/Library/CloudStorage/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data'
]
# Iterate over the list and set base_dir to the first existing path
for path in path_options:
    if os.path.exists(path):
        base_dir = path
        break
else:
    # If no valid path is found, raise an error or handle it appropriately
    print("None of the specified directories exist.")
    base_dir = None  # Or set a default path if appropriate
print("Base directory set to:", base_dir)
static_dir = os.path.join(base_dir, 'Tables')
inventory_file_path = os.path.join(static_dir, 'R_EstoqComp.xlsx')  # Update to the correct path if needed

column_rename_dict = {
    'O_NFCI': {
        'Operação': 'OP',
        'Nota Fiscal': 'NF',
        'Data de Emissão (completa)': 'EMISS',
        'Cliente (Razão Social)': 'NOMERS',
        'Cliente (Nome Fantasia)': 'NOMEF',
        'Código do Produto': 'CODPF', 
        'Quantidade': 'QTD', 
        'Total de Mercadoria': 'MERCVLR',
        'Valor do ICMS ST': 'ICMSST',
        'Valor do IPI': 'IPI',
        'Total da Nota Fiscal': 'TOTALNF',
        'Valor do ICMS': 'ICMS',
        'Estado': 'UF'
        # Add other columns that need renaming for O_NFCI
    },
     'L_LPI': {
        'Preço Com Desconto': 'VLRVENDA',
        'SKU': 'CODPF',
        'Vendas': 'QTD'
    },
    'MLK_Vendas' : {
        'PREÇO UNITÁRIO DE VENDA DO ANÚNCIO (BRL)': 'PRECOUNIT',
        'Quantidade' : 'QTD'
    }
    # Add dictionaries for other dataframes...
}

column_format_dict = {
    'O_NFCI': {
        'EMISS': 'DD-MMM-YY',
        'QTD': '0',
        'PRECO CALC': '#,##0.00',
        'MERCVLR': '#,##0.00',
        'ICMSST': '#,##0.00',
        'IPI': '#,##0.00',
        'TOTALNF': '#,##0.00',
        'ICMS': '#,##0.00',
        'ECU': '#,##0.00',
        'COMISSPCT': '0.00%',
        'FRETEPCT': '0.00%',
        'VERBAPCT': '0.00%',
        'ECT': '#,##0.00',
        'COMISSVLR': '#,##0.00',
        'FRETEVLR': '#,##0.00',
        'MARGVLR': '#,##0.00',
        'MARGPCT': '0.00%',
        # Add other formats for O_NFCI
    },
    'L_LPI':{
        'VLRVENDA': '#,##0.00',        
        'ECU': '#,##0.00',
        'ECTK': '#,##0.00',
    },
    'MLK_Vendas':{
        'MARGVLR': '#,##0.00',        
        'MARGPCT': '0.00%',
    },

    # Add dictionaries for other dataframes...
}

audit_client_names = ['ALWE', 'COMPROU CHEGOU', 'NEXT COMPRA']  # Add other clients as needed
invaudit_client_names = ['ALWE', 'COMPROU CHEGOU', 'NEXT COMPRA']  # Add other clients as needed

def rename_columns(all_data, column_rename_dict):
    for df_name, rename_dict in column_rename_dict.items():
        if df_name in all_data:
            df = all_data[df_name]
            # Debug print to verify columns before renaming
            print(f"Before rename:\nTable: {df_name}\nColumns: {df.columns.tolist()}")
            df.rename(columns=rename_dict, inplace=True)
            # Debug print to verify columns after renaming
            print(f"Renamed columns in {df_name}: {rename_dict}")
            print(f"Columns in {df_name}: {df.columns.tolist()}")
            all_data[df_name] = df
    return all_data

def load_recent_data(base_dir, file_pattern, months=1):
    #end_date = datetime.now()
    end_date = datetime(2024, 7, 15)
    start_date = end_date - timedelta(days=months * 30)  # Approximately three months
    frames = []
    for month_count in range(months + 1):  # Current month + last three months
        year_month = (start_date + timedelta(days=30 * month_count)).strftime('%Y_%m')
        file_path = os.path.join(base_dir, 'clean', year_month, file_pattern.format(year_month=year_month))
        if os.path.exists(file_path):
            df = pd.read_excel(file_path)
            frames.append(df)
            print(f"Loaded {file_path} with shape: {df.shape}")  # Debug print
        else:
            print(f"File not found: {file_path}")  # Debug print
    return pd.concat(frames) if frames else pd.DataFrame()

def load_static_data(static_dir, filename):
    return pd.read_excel(os.path.join(static_dir, filename))

def standardize_text_case(df):
    """Convert all text to uppercase for standardization."""
    if isinstance(df, pd.DataFrame):
        df.columns = [col.upper() for col in df.columns]
        for col in df.select_dtypes(include=[object]).columns:
            df[col] = df[col].str.upper()
    return df

def merge_all_data(all_data):
    print(f"Creating Merged and Calculated Columns")

    # Ensure all relevant columns are in uppercase for case-insensitive comparison
    all_data = {key: standardize_text_case(df) for key, df in all_data.items()}

    # compute column ANOMES
    compute_NFCI_ANOMES(all_data)
    compute_LPI_ANOMES(all_data)
    compute_ML_ANOMES(all_data)

    # Merge O_NFCI with T_Remessas - REM
    all_data = merge_data(all_data, "O_NFCI", "NOMEF", "T_Remessas", "NOMEF", "REM_NF", default_value=0)

    # Merge O_NFCI with T_Prodf - CODPP
    all_data = merge_data(all_data, "O_NFCI", "CodPF", "T_ProdF", "CodPF", "CODPP", default_value="xxx")
    all_data = merge_data(all_data, "L_LPI", "CodPF", "T_ProdF", "CodPF", "CODPP", default_value="xxx")
    all_data = merge_data(all_data, "MLA_Vendas", "SKU", "T_ProdF", "CodPF", "CODPP", default_value="xxx")
    all_data = merge_data(all_data, "MLK_Vendas", "SKU", "T_ProdF", "CodPF", "CODPP", default_value="xxx")

    # Merge O_NFCI with T_GruposCli - G1
    all_data = merge_data(all_data, "O_NFCI", "NomeF", "T_GruposCli", "NomeF", "G1", default_value="V")

    # Merge O_NFCI with ECU on columns 'EMISS' and 'CodPF'
    all_data = merge_data2v(all_data, "O_NFCI", "ANOMES", "CodPF", "ECU", "ANOMES", "CODPF", "VALUE", "ECU", default_value=999)
    all_data = merge_data2v(all_data, "L_LPI", "ANOMES", "CodPF", "ECU", "ANOMES", "CODPF", "VALUE", "ECU", default_value=999)
    all_data = merge_data2v(all_data, "MLA_Vendas", "ANOMES", "SKU", "ECU", "ANOMES", "CODPF", "VALUE", "ECU", default_value=999)
    all_data = merge_data2v(all_data, "MLK_Vendas", "ANOMES", "SKU", "ECU", "ANOMES", "CODPF", "VALUE", "ECU", default_value=999)
 
    # Merge VENDEDOR with T_REPS for COMPCT
    all_data = merge_data(all_data, "O_NFCI", "Vendedor", "T_Reps", "Vendedor", "COMISSPCT", default_value=0)

    # Merge UF with T_Fretes for FretePCT
    all_data = merge_data(all_data, "O_NFCI", "UF", "T_Fretes", "UF", "FRETEPCT", default_value=0)
    # Set FRETEPCT = 0 where G1 = "DROP" or "ALWE" in O_NFCI table
    if 'O_NFCI' in all_data:
        all_data['O_NFCI'].loc[all_data['O_NFCI']['G1'].isin(['DROP', 'ALWE']), 'FRETEPCT'] = 0

    # Merge NomeF with T_Verbas for VerbaPct
    all_data = merge_data(all_data, "O_NFCI", "NOMEF", "T_Verbas", "NomeF", "VERBAPCT", default_value=0)

    # Perform the merge (example merge, adjust as necessary)
    all_data = merge_data(all_data, "L_LPI", "INTEGRAÇÃO", "T_MP", "Integração", "Empresa", default_value='erro')
    all_data = merge_data(all_data, "L_LPI", "INTEGRAÇÃO", "T_MP", "Integração", "MP", default_value='erro')
    all_data = merge_data(all_data, "L_LPI", "INTEGRAÇÃO", "T_MP", "Integração", "EmpresaF", default_value='erro')

    # OrderStatus Merge
    all_data = merge_data(all_data, "MLA_Vendas", "STATUS", "T_MLStatus", "MLStatus", "OrderStatus", default_value='erro')
    all_data = merge_data(all_data, "MLK_Vendas", "STATUS", "T_MLStatus", "MLStatus", "OrderStatus", default_value='erro')

    # Ctas a Pagar e Receber
    all_data = merge_data(all_data, "O_CtasAPagar", "CATEGORIA", "T_CtasAPagarClass", "Categoria", "GrupoCtasAPagar", default_value='erro')
    all_data = merge_data(all_data, "O_CtasARec", "CATEGORIA", "T_CtasAPagarClass", "Categoria", "GrupoCtasAPagar", default_value='erro')

    
    for key, df in all_data.items():
        if key == 'O_NFCI':
            # print_table_and_columns(all_data, "O_NFCI")

            # Create column "C"
            df['C'] = 1 - df['REM_NF']
            
            # Create column "B"
            df['B'] = df.apply(lambda row: 1 if row['OP'] == 'REMESSA DE PRODUTO' and row['C'] == 1 else 0, axis=1)
            
            # Create column ECT (ECU x QTD)
            df['ECT'] = df['ECU'] * df['QTD'] * df['C']
 
            # Create column COMVLR (VLRMERC x COMPCT)
            df['COMISSVLR'] = df['MERCVLR'] * df['COMISSPCT'] * df['C']

            # Create column FreteVLR (FretePCT x TotalNF)            
            #df['FRETEVLR'] = df['FRETEPCT'] * df['TOTALNF'] * df['C']
            df['FRETEVLR'] = df.apply(lambda row: max(row['FRETEPCT'] * row['TOTALNF'] * row['C'], row['FRETEPCT'] * row['ECT'] * row['C'] * 2), axis=1)

            # Create column VerbaVLR (VerbaPCT x TotalNF)
            df['VERBAVLR'] = df['VERBAPCT'] * df['TOTALNF'] * df['C']

            # Create column MargCVlr
            df['MARGVLR'] = df['C'] * ( df['MERCVLR'] * (1 - 0.0925) - df['ICMS'] ) - df['VERBAVLR'] - df['FRETEVLR'] - df['COMISSVLR'] - df['ECT']

            # Create column VerbaVLR (VerbaPCT x TotalNF)
            df['MARGPCT'] = df['MARGVLR'] / df['MERCVLR']

        elif key == 'L_LPI':
            cols_to_drop = ['PREÇO', 'PREÇO TOTAL', 'DESCONTO ITEM', 'DESCONTO TOTAL']
            df = df.drop([x for x in cols_to_drop if x in df.columns], axis=1)
            # Add the 'Valido' column directly
            df['VALIDO'] = df['STATUS PEDIDO'].apply(lambda x: 0 if x in ['CANCELADO', 'PENDENTE', 'AGUARDANDO PAGAMENTO'] else 1)
            df['KAB'] = df.apply(lambda row: 1 if row['VALIDO'] == 1 and row['EMPRESA'] in ['K', 'A', 'B'] else 0, axis=1)
            df['ECTK'] = df['ECU'] * df['QTD'] * df['KAB']

            # Add the 'TipoAnuncio' column directly from 'MLK_Vendas'
            if 'MLK_Vendas' in all_data:
                df = df.merge(
                    all_data['MLK_Vendas'][['N.º DE VENDA_HYPERLINK', 'TIPO DE ANÚNCIO']],
                    left_on='CÓDIGO PEDIDO',
                    right_on='N.º DE VENDA_HYPERLINK',
                    how='left'
                )
                df['TipoAnuncio'] = df.apply(lambda row: row['TIPO DE ANÚNCIO'] if row['EMPRESA'] == 'K' and row['MP'] == 'ML' else None, axis=1)
                df.drop(columns=['N.º DE VENDA_HYPERLINK', 'TIPO DE ANÚNCIO'], inplace=True)

            # Add the 'TipoAnuncio' column for 'A' and lookup in 'MLA_Vendas'
            if 'MLA_Vendas' in all_data:
                df = df.merge(
                    all_data['MLA_Vendas'][['N.º DE VENDA_HYPERLINK', 'TIPO DE ANÚNCIO']],
                    left_on='CÓDIGO PEDIDO',
                    right_on='N.º DE VENDA_HYPERLINK',
                    how='left'
                )
                df['TipoAnuncio'] = df.apply(lambda row: row['TIPO DE ANÚNCIO'] if row['EMPRESA'] == 'A' and row['MP'] == 'ML' else row['TipoAnuncio'], axis=1)
                df.drop(columns=['N.º DE VENDA_HYPERLINK', 'TIPO DE ANÚNCIO'], inplace=True)

            # Add colum Compctmp (Comissão pct por Marketplace)
            if 'T_RegrasMP' in all_data:
                df = df.merge(
                    all_data['T_RegrasMP'][['MPX', 'TARMP']],
                    left_on='MP',
                    right_on='MPX',
                    how='left'
                )
                df['Compctmp'] = df['TARMP']
                df.drop(columns=['MPX', 'TARMP'], inplace=True)

            # Add colum Compctml (Comissão pct pro ML Classico/Premium)
                df = df.merge(
                    all_data['T_RegrasMP'][['MPX', 'TARMP']],
                    left_on='TipoAnuncio',
                    right_on='MPX',
                    how='left'
                )
                df['Compctml'] = df['TARMP']
                df.drop(columns=['MPX', 'TARMP'], inplace=True)

            # Create the ComPct column based on the condition
            df['ComPct'] = df.apply(lambda row: row['Compctml'] if pd.notnull(row['Compctml']) else row['Compctmp'], axis=1)
            df['Com'] = df['VLRVENDA'] * df['ComPct'] * df['KAB']

        elif key == 'MLA_Vendas':
            # Add the 'VALIDO' column directly
            df['Imposto1'] = df['VLRTOTALPSKU']*(0.11)
            df['Imposto2'] = 0
            df['ImpostoT'] =  df['Imposto1'] + df['Imposto2']

            cols_to_drop = ['CODPF_x', 'CODPF_y', 'MLSTATUS']
            df = df.drop([x for x in cols_to_drop if x in df.columns], axis=1)

        elif key == 'MLK_Vendas':
            # Create column ECT (ECU x QTD)
            df['ECTK'] = df['ECU'] * df['QTD']

            # Add the 'Impostos' columns directly
            df['Imposto1'] = df['VLRTOTALPSKU']*(0.0925)
            df['Imposto2'] = df['VLRTOTALPSKU']*(0.18)
            df['ImpostoT'] =  df['Imposto1'] + df['Imposto2']

            # Create column MargCVlr
            df['MARGVLR'] = df['REPASSE'] - df['ImpostoT'] - df['ECTK'] - (1)
            df['MARGPCT'] = df['MARGVLR'] / df['VLRTOTALPSKU']

            cols_to_drop = ['CODPF_x', 'CODPF_y', 'MLSTATUS']
            df = df.drop([x for x in cols_to_drop if x in df.columns], axis=1)

        elif key == 'O_CtasARec':
            # Step 2: Create the 'DATA BASE' column which is the last day of the month
            df['DATA BASE'] = pd.to_datetime(df['ANOMES'], format='%y%m') + MonthEnd(0)
            # Step 3: Calculate 'DIAS ATRASO'
            df['DIAS ATRASO'] = (df['DATA BASE'] - df['VENCIMENTO']).dt.days
            # Step 4: Apply condition to set DIAS ATRASO to 0 if VENCIMENTO is greater than DATA BASE
            df['DIAS ATRASO'] = df['DIAS ATRASO'].apply(lambda x: max(0, x))
            # Step 5: Classify 'DIAS ATRASO' using the classification table from all_data['T_CtasARecClass']
            df_ctas_a_rec_class = all_data['T_CtasARecClass']

            # Merge based on the 'DIAS ATRASO' column and classification table
            df = pd.merge(df, df_ctas_a_rec_class, how='left', left_on='DIAS ATRASO', right_on='DEXDIAS')
        
            # Apply the range condition for 'DIAS ATRASO' to determine classification
            df['CLASSIFICACAO'] = df.apply(lambda row: row['STATUS ATRASO'] if row['DEXDIAS'] <= row['DIAS ATRASO'] <= row['DEXDIAS'] else None, axis=1)
        
            # Filter out rows where the classification was not within the proper range
            df = df.dropna(subset=['CLASSIFICACAO'])

        # Update the dataframe in all_data
        all_data[key] = df

    return all_data

def preprocess_inventory_data(file_path):
    sheets = pd.read_excel(file_path, sheet_name=None, header=1)  # Load data with headers from the second row
    processed_sheets = {}

    for sheet_name, df in sheets.items():
        df = df.melt(id_vars=['CodPF'], var_name='ANOMES', value_name='Value')
        df['ANOMES'] = pd.to_datetime(df['ANOMES'], errors='coerce').dt.strftime('%y%m')
        processed_sheets[sheet_name] = df
    
    return processed_sheets

def merge_data(all_data, df1_name, df1_col, df2_name, df2_col, new_col=None, indicator_name=None, default_value=None):
    df1_col = df1_col.upper()
    df2_col = df2_col.upper()
    if new_col:
        new_col = new_col.upper()

    if df1_name in all_data and df2_name in all_data:
        df1 = all_data[df1_name]
        df2 = all_data[df2_name]

        # Standardize column names
        df1.columns = [col.upper() for col in df1.columns]
        df2.columns = [col.upper() for col in df2.columns]

        if df1_col not in df1.columns or df2_col not in df2.columns:
            raise KeyError(f"Column '{df1_col}' or '{df2_col}' not found in dataframes.")

        df2_cols = [df2_col] + ([new_col] if new_col else [])
        merged_df = df1.merge(df2[df2_cols].drop_duplicates(), left_on=df1_col, right_on=df2_col, how='left', indicator=indicator_name, suffixes=('', '_DROP'))

        # Remove the '_DROP' columns
        merged_df.drop([col for col in merged_df.columns if col.endswith('_DROP')], axis=1, inplace=True)

        if indicator_name and default_value is not None:
            merged_df[indicator_name] = merged_df[indicator_name].apply(lambda x: default_value if x == 'left_only' else merged_df[new_col])
            merged_df.drop(columns=[new_col, indicator_name], inplace=True)
        elif new_col and default_value is not None:
            merged_df[new_col] = merged_df[new_col].fillna(default_value)

        all_data[df1_name] = merged_df
    return all_data

def merge_data2v(all_data, df1_name, df1_col1, df1_col2, df2_name, df2_col1, df2_col2, df2_val_col, new_col_name, default_value=None, negative=False):
    df1_col1 = df1_col1.upper()
    df1_col2 = df1_col2.upper()
    df2_col1 = df2_col1.upper()
    df2_col2 = df2_col2.upper()
    df2_val_col = df2_val_col.upper()
    new_col_name = new_col_name.upper()

    if df1_name in all_data and df2_name in all_data:
        df1 = all_data[df1_name]
        df2 = all_data[df2_name]

        # Standardize column names
        df1.columns = [col.upper() for col in df1.columns]
        df2.columns = [col.upper() for col in df2.columns]

        #print(f"Columns in {df1_name} before merge: {df1.columns}")
        #print(f"Columns in {df2_name} before merge: {df2.columns}")

        if df1_col1 not in df1.columns or df1_col2 not in df1.columns or df2_col1 not in df2.columns or df2_col2 not in df2.columns:
            raise KeyError(f"One of the columns '{df1_col1}', '{df1_col2}', '{df2_col1}', '{df2_col2}' not found in dataframes.")

        if negative:
            df2[df2_val_col] = df2[df2_val_col] * -1  # Make the VALUE column negative

        df2_cols = [df2_col1, df2_col2, df2_val_col]
        merged_df = df1.merge(df2[df2_cols].drop_duplicates(), left_on=[df1_col1, df1_col2], right_on=[df2_col1, df2_col2], how='left')

        if df2_val_col and default_value is not None:
            merged_df[df2_val_col] = merged_df[df2_val_col].fillna(default_value)

        # Rename the value column to the new column name
        merged_df.rename(columns={df2_val_col: new_col_name}, inplace=True)

        #print(f"Columns after merge: {merged_df.columns}")
        all_data[df1_name] = merged_df
    return all_data

def compute_NFCI_ANOMES(all_data):
    for key, df in all_data.items():
        # Add the ANOMES column to O_NFCI
        if key == 'O_NFCI' and 'EMISS' in df.columns:
            df['EMISS'] = pd.to_datetime(df['EMISS'], errors='coerce')  # Ensure the date is parsed correctly
            df['ANOMES'] = df['EMISS'].dt.strftime('%y%m')  # Format date as YYMM
            print(f"Added ANOMES column to {key}")
        all_data[key] = df
    return all_data

def compute_LPI_ANOMES(all_data):
    for key, df in all_data.items():
        # Add the ANOMES column to L_LPI
        if key == 'L_LPI' and 'DATA' in df.columns:
            df['DATA'] = pd.to_datetime(df['DATA'], errors='coerce')  # Ensure the date is parsed correctly
            df['ANOMES'] = df['DATA'].dt.strftime('%y%m')  # Format date as YYMM
            print(f"Added ANOMES column to {key}")
        all_data[key] = df
    return all_data

def compute_ML_ANOMES(all_data):
    for key, df in all_data.items():
        # Add the ANOMES column to MLA_Vendas and MLK_Vendas
        if key in ['MLA_Vendas', 'MLK_Vendas'] and 'DATA DA VENDA' in df.columns:
            # Use dateutil parser to parse the date string
            df['DATA DA VENDA'] = df['DATA DA VENDA'].apply(lambda x: parser.parse(x, fuzzy=True) if pd.notnull(x) else pd.NaT)
            df['ANOMES'] = df['DATA DA VENDA'].dt.strftime('%y%m')  # Format date as YYMM
            print(f"Added ANOMES column to {key}")
        all_data[key] = df
    return all_data

def compute_LPI_ANOMES(all_data):
    for key, df in all_data.items():
        # Add the ANOMES column to L_LPI
        if key == 'L_LPI' and 'DATA' in df.columns:
            df['DATA'] = pd.to_datetime(df['DATA'], errors='coerce')  # Ensure the date is parsed correctly
            df['ANOMES'] = df['DATA'].dt.strftime('%y%m')  # Format date as YYMM
            print(f"Added ANOMES column to {key}")
        all_data[key] = df
    return all_data

def mlcustom_date_parser(date_str):
    # Remove the 'hs.' part if it exists
    date_str = re.sub(r'\s*hs\.\s*$', '', date_str)
    
    # Replace the Portuguese month names with English month names
    month_map = {
        'JANEIRO': 'January', 'FEVEREIRO': 'February', 'MARÇO': 'March', 'ABRIL': 'April',
        'MAIO': 'May', 'JUNHO': 'June', 'JULHO': 'July', 'AGOSTO': 'August',
        'SETEMBRO': 'September', 'OUTUBRO': 'October', 'NOVEMBRO': 'November', 'DEZEMBRO': 'December'
    }
    
    for pt_month, en_month in month_map.items():
        date_str = date_str.replace(pt_month, en_month)
    
    # Parse the date
    return pd.to_datetime(date_str, format='%d DE %B DE %Y %H:%M HS.')

def compute_ML_ANOMES(all_data):
    for key, df in all_data.items():
        # Add the ANOMES column to MLA_Vendas and MLK_Vendas
        if key in ['MLA_Vendas', 'MLK_Vendas'] and 'DATA DA VENDA' in df.columns:
            # Use custom date parser to parse the date string
            df['DATA DA VENDA'] = df['DATA DA VENDA'].apply(lambda x: mlcustom_date_parser(x) if pd.notnull(x) else pd.NaT)
            df['ANOMES'] = df['DATA DA VENDA'].dt.strftime('%y%m')  # Format date as YYMM
            print(f"Added ANOMES column to {key}")
        all_data[key] = df
    return all_data

def load_inventory_data(file_path):
    return pd.read_excel(file_path)

def print_all_tables_and_columns(all_data):
    for table_name, df in all_data.items():
        print(f"Table: {table_name}")
        print("Columns:", df.columns.tolist())
        print("-" * 50)

def print_table_and_columns(all_data, table_name):
    if table_name in all_data:
        print(f"Table: {table_name}")
        print("Columns:", all_data[table_name].columns.tolist())
        print("-" * 50)
    else:
        print(f"Table '{table_name}' not found in the dataset.")

def excel_format(output_path, column_format_dict):
    print("Formatting all sheets")
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True)
    header_style.fill = PatternFill("solid", fgColor="6ac5fe")  # Light blue background color
    header_style.alignment = Alignment(horizontal="center", vertical="center")

    workbook = load_workbook(output_path)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Apply header style
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.style = header_style

        if sheet_name in column_format_dict:
            formats = column_format_dict[sheet_name]
            for col_name, col_format in formats.items():
                # Find the column index based on header name
                for col_idx, cell in enumerate(worksheet[1], start=1):
                    if cell.value == col_name:
                        break
                else:
                    continue  # Skip if column name is not found

                # Apply the format to the entire column
                for row_idx in range(2, worksheet.max_row + 1):  # Start from the second row to avoid header
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = col_format

    workbook.save(output_path)
    print(f"All sheets formatted")

def excel_autofilters(output_path):
    print("Adding auto-filters to all sheets")
    workbook = load_workbook(output_path)
    for sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]
        worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(output_path)
    print("Added auto-filters to all sheets")

# Define the audit function
def perform_audit(df, client_name):
    audit_columns = [
        'CODPF',
        'QTD',
        'PRECO CALC',
        'MERCVLR',
        'ICMSST',
        'IPI',
        'TOTALNF',
        'EMISS']

    audit_df = df[df['NOMEF'] == client_name][audit_columns]  
    return audit_df

# Define the function to perform audits for all specified clients
def perform_all_audits(all_data):
    for client_name in audit_client_names:
        audit_df = perform_audit(all_data['O_NFCI'], client_name)
        all_data[f'Audit_{client_name}'] = audit_df
        print(f"Performed audit for {client_name}")  # Debug print
    return all_data

# Function to track inventory movements
# Function to calculate realized cost
def track_inventory(sales_data, purchase_data):
    inventory_movements = []

    # Process sales data
    for index, row in sales_data.iterrows():
        movement = {
            'Date': row['DATA'],
            'Invoice Number': None,
            'Product Code': row['CODPF'],
            'Quantity': -row['QTD'],
            'CV': 'V',
            'QTD E': row['QTD'],
            'CMV Unit E': None,
            'CMV Mov E': None,
            'QTD R': None,
            'CMV Unit R': None,
            'CMV Mov R': None,
            'NF Compra': None
        }
        inventory_movements.append(movement)

    # Process purchase data
    for index, row in purchase_data.iterrows():
        movement = {
            'Date': row['EMISS'],
            'Invoice Number': row['NF'],
            'Product Code': row['CODPF'],
            'Quantity': row['QTD'],
            'CV': 'C',
            'QTD E': None,
            'CMV Unit E': row['PRECO CALC'],
            'CMV Mov E': row['MERCVLR'],
            'QTD R': None,
            'CMV Unit R': None,
            'CMV Mov R': None,
            'NF Compra': row['NF'],
            'Custo Total Unit': row['TOTALNF'] / row['QTD']
        }
        inventory_movements.append(movement)

    inventory_df = pd.DataFrame(inventory_movements)
    inventory_df.sort_values(by='Date', inplace=True)
    return inventory_df

def calculate_realized_cost(inventory_df):
    # Get all purchases (C) in ascending date order
    purchase_data = inventory_df[inventory_df['CV'] == 'C'].sort_values(by='Date')

    # Create a list of purchases as objects with necessary details
    purchase_list = []
    for _, row in purchase_data.iterrows():
        purchase_list.append({
            'Product Code': row['Product Code'],
            'Invoice Number': row['Invoice Number'],
            'Quantity': row['Quantity'],
            'Custo Total Unit': row['Custo Total Unit']
        })
    print("Purchase List:", purchase_list)  # Debug print

    # Iterate through the sales (V) and populate the realized cost details
    for index, row in inventory_df[inventory_df['CV'] == 'V'].iterrows():
        quantity_needed = -row['Quantity']
        for purchase in purchase_list:
            if purchase['Product Code'] == row['Product Code'] and quantity_needed > 0:
                if purchase['Quantity'] > 0:
                    quantity_to_apply = min(purchase['Quantity'], quantity_needed)

                    # Update the realized cost details
                    inventory_df.at[index, 'QTD R'] = quantity_to_apply
                    inventory_df.at[index, 'CMV Unit R'] = purchase['Custo Total Unit']
                    inventory_df.at[index, 'CMV Mov R'] = quantity_to_apply * purchase['Custo Total Unit']
                    inventory_df.at[index, 'NF Compra'] = purchase['Invoice Number']

                    # Update the purchase details
                    purchase['Quantity'] -= quantity_to_apply
                    quantity_needed -= quantity_to_apply

        # If there's still quantity needed, populate the expected cost details
        if quantity_needed > 0:
            inventory_df.at[index, 'QTD E'] = quantity_needed
            if row['CMV Unit E'] is not None:
                inventory_df.at[index, 'CMV Mov E'] = quantity_needed * row['CMV Unit E']

    # Add remaining purchase quantities back to the corresponding purchase rows
    for purchase in purchase_list:
        if purchase['Quantity'] > 0:
            purchase_indices = inventory_df[
                (inventory_df['Product Code'] == purchase['Product Code']) &
                (inventory_df['Invoice Number'] == purchase['Invoice Number'])
            ].index
            if not purchase_indices.empty:
                purchase_index = purchase_indices[0]
                inventory_df.at[purchase_index, 'QTD E'] = purchase['Quantity']
                inventory_df.at[purchase_index, 'CMV Unit E'] = purchase['Custo Total Unit']
                inventory_df.at[purchase_index, 'CMV Mov E'] = purchase['Quantity'] * purchase['Custo Total Unit']

    print(inventory_df)  # Debug print
    return inventory_df



# Function to perform the inventory audit
def perform_invaudit(o_nfci_df, l_lpi_df, client_name):
    sales_data = l_lpi_df[l_lpi_df['EMPRESAF'] == client_name]
    purchase_data = o_nfci_df[o_nfci_df['NOMEF'] == client_name]

    inventory_df = track_inventory(sales_data, purchase_data)
    inventory_df = calculate_realized_cost(inventory_df)
    return inventory_df

# Function to perform all inventory audits
def perform_all_invaudits(all_data):
    o_nfci_df = all_data['O_NFCI']
    l_lpi_df = all_data['L_LPI']

    invaudit_results = {}
    for client in invaudit_client_names:
        invaudit_results[client] = perform_invaudit(o_nfci_df, l_lpi_df, client)

    # Add the audit results to the all_data dictionary
    for client, df in invaudit_results.items():
        all_data[f'InvAudit_{client}'] = df

    return all_data

def main():
    # Define file patterns for each data type
    file_patterns = {
        'O_NFCI': 'O_NFCI_{year_month}_clean.xlsx',
        'L_LPI': 'L_LPI_{year_month}_clean.xlsx',
        'MLA_Vendas': 'MLA_Vendas_{year_month}_clean.xlsx',
        'MLK_Vendas': 'MLK_Vendas_{year_month}_clean.xlsx',
        'O_CC': 'O_CC_{year_month}_clean.xlsx',
        'O_CtasAPagar': 'O_CtasAPagar_{year_month}_clean.xlsx',
        'O_CtasARec': 'O_CtasARec_{year_month}_clean.xlsx',
    }

    all_data = {}

    for key, pattern in file_patterns.items():
        recent_data = load_recent_data(base_dir, pattern)
        print(f"{key} data shape: {recent_data.shape}")  # Debug print

        # Ensure 'N.º de venda' is treated as string if the column exists
        if 'N.º de venda' in recent_data.columns:
            recent_data['N.º de venda'] = recent_data['N.º de venda'].astype(str)
        if 'N.º de venda_hyperlink' in recent_data.columns:
            recent_data['N.º de venda_hyperlink'] = recent_data['N.º de venda_hyperlink'].astype(str)

        all_data[key] = recent_data

    # Load static data
    static_tables = ['T_CondPagto.xlsx', 'T_Fretes.xlsx', 'T_GruposCli.xlsx', 'T_MP.xlsx', 
                     'T_RegrasMP.xlsx', 'T_Remessas.xlsx', 'T_Reps.xlsx', 'T_Verbas.xlsx','T_Vol.xlsx', 'T_ProdF.xlsx', 
                     'T_ProdP.xlsx', 'T_Entradas.xlsx', 'T_FretesMP.xlsx', 'T_MLStatus.xlsx', 'T_CtasAPagarClass.xlsx', 'T_CtasARecClass.xlsx']
    static_data_dict = {table.replace('.xlsx', ''): load_static_data(static_dir, table) for table in static_tables}
    
    # Check static data shapes
    for key, df in static_data_dict.items():
        print(f"Static data {key} shape: {df.shape}")  # Debug print
    
    inventory_data = preprocess_inventory_data(inventory_file_path)
 
    # Add static data to all_data dictionary
    all_data.update(static_data_dict) 
    all_data.update(inventory_data)
    
    all_data = rename_columns(all_data, column_rename_dict)

    # Merge all data with static data
    all_data = merge_all_data(all_data) 

    # Perform audits for the specified clients
    all_data = perform_all_audits(all_data)
    print(f"Audit completed for clients: {', '.join(audit_client_names)}")

    # Perform innventory audits for the specified clients
    all_data = perform_all_invaudits(all_data)
    print(f"INVENTORY Audit completed for clients: {', '.join(audit_client_names)}")

    # Save all data to one Excel file with multiple sheets
    output_path = os.path.join(base_dir, 'clean', 'merged_data.xlsx')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for key, df in all_data.items():
            df.to_excel(writer, sheet_name=key, index=False)
            print(f"Added {key} data to {output_path} in sheet {key}")  # Debug print

    print(f"All merged data saved to {output_path}")

    excel_format(output_path, column_format_dict)
    excel_autofilters(output_path)

if __name__ == "__main__":
    main()
