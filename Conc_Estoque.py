#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Inventory reconciliation for a given Year/Month.

Inputs (auto-resolved from 2 candidate base folders):
- Previous month inventory summary: R_Estoque_fdm_YYYY_MM.xlsx (sheet PT01)
- Current  month inventory summary: R_Estoque_fdm_YYYY_MM.xlsx (sheet PT01)
- Monthly sales summary:            R_Resumo_YYYY_MM.xlsx (sheets O_NFCI, L_LPI)
- Arrivals (table):                 Tables/T_Entradas.xlsx  (all months; we filter by the target month)

Output:
CSV + Excel with columns:
CODPF, QT_I, CU_I, CT_I, VENDAS_2b, VENDAS_2c, Qt_E, CU_E, CT_E, Qt_SS, CU_F, CT_Ger
"""

from __future__ import annotations
import os
from pathlib import Path
import pandas as pd
import numpy as np
from typing import List, Optional, Tuple
from itertools import cycle

# -----------------------
# CONFIG – adjust these!
# -----------------------
# Two possible base folders, like your process_inv.py pattern
path_options = [
    '/Users/mauricioalouan/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data/',
    '/Users/simon/Library/CloudStorage/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data'
]
for candidate in path_options:
    if os.path.exists(candidate):
        base_dir = candidate
        break
else:
    print("None of the specified directories exist.")
    base_dir = None

# Subfolder with Tables
TABLES_SUBDIR = "Tables"
INPUT_SUBDIR = "clean"
OUTPUT_SUBDIR = "clean"

# <<< HARD-CODE AQUI: nomes dos arquivos (prefixos) e tabela de Entradas
INV_PREFIX      = "R_Estoq_fdm_"     # gera "R_Estoque_fdm_YYYY_MM.xlsx"
RESUMO_PREFIX   = "R_Resumo_"          # gera "R_Resumo_YYYY_MM.xlsx"

# <<< HARD-CODE AQUI: nomes das ABAS
SHEET_PT01   = "PT01"
SHEET_ONFCI  = "O_NFCI"
SHEET_LLPI   = "L_LPI"

# <<< HARD-CODE AQUI: nomes das COLUNAS por aba/tabela
# PT01 (estoque)

# O_NFCI (vendas 2b)
ONFCI_QTY_COL  = "QT"                  # ex.: "QT", "Quantidade", "QT"

# L_LPI (vendas 2c)
LLPI_CODE_COL    = "CODPP"
LLPI_QTY_COL     = "QT"
LLPI_STATUS_COL  = "STATUS"     # usado para filtrar != "CANCELADO"
LLPI_EMPRESA_COL = "EMPRESA"           # usado para filtrar == "K"

# Filtragem por mês:
ENTR_ANOMES_COL = "AnoMes"             # yymm (ex.: 2507). Se não existir, use None.
ENTR_DATE_COL   = None                 # OU nome da data (ex.: "Emissao" ou "Ultima Entrada") se preferir filtrar por data

CLEAN_ROOT = os.path.join(base_dir, INPUT_SUBDIR) 
TABLES_DIR = os.path.join(base_dir, TABLES_SUBDIR)

# -----------------------
# Helpers
# -----------------------
def find_existing_excel(base_path: Path, base_name: str) -> Path:
    """
    Tenta encontrar arquivo com base em base_name + .xlsx ou .xlsm
    """
    for ext in [".xlsx", ".xlsm"]:
        candidate = base_path / f"{base_name}{ext}"
        if candidate.exists():
            return candidate
    raise FileNotFoundError(f"Arquivo não encontrado: {base_path}/{base_name}.xlsx ou .xlsm")

def ym_to_prev(year: int, month: int) -> Tuple[int, int]:
    if month == 1:
        return (year - 1, 12)
    return (year, month - 1)

def yymm_to_str(year: int, month: int) -> str:
    return f"{year:04d}_{month:02d}"

def find_existing_file(candidates: List[Path]) -> Optional[Path]:
    for p in candidates:
        if p.exists():
            return p
    return None

def resolve_month_dir(year: int, month: int) -> Path:
    tag = f"{year:04d}_{month:02d}"
    p = Path(os.path.join(CLEAN_ROOT, tag))
    if not p.exists():
        raise FileNotFoundError(f"Pasta do mês não encontrada: {p}")
    return p

def resolve_tables_dir(year: int, month: int) -> Path:
    p = Path(os.path.join(TABLES_DIR))
    if not p.exists():
        raise FileNotFoundError(f"Pasta das Tabelas não encontrada: {p}")
    return p

def read_excel_safe(path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name=sheet_name, dtype=str) if sheet_name else pd.read_excel(path, dtype=str)
    except ValueError:
        # some files contain mixed dtypes; retry without dtype enforcement
        return pd.read_excel(path, sheet_name=sheet_name)

def norm_code(s: pd.Series) -> pd.Series:
    return s.astype(str).str.strip().str.upper()

def first_existing_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    lower_map = {c.lower(): c for c in df.columns}
    for name in candidates:
        if name in df.columns:
            return name
        if name.lower() in lower_map:
            return lower_map[name.lower()]
    return None

def coerce_numeric(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce").fillna(0)

def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

# -----------------------
# Loaders / Normalizers
# -----------------------

def load_prodf(tables_dir: Path) -> pd.DataFrame:
    path = tables_dir / "T_Prodf.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"Arquivo Prodf.xlsx não encontrado em {tables_dir}")
    df = pd.read_excel(path)
    df["CODPF"] = df["CodPF"].astype(str).str.strip().str.upper()
    df["CODPP"] = df["CodPP"].astype(str).str.strip().str.upper()
    return df[["CODPF", "CODPP"]]

def load_curr_inventory_data(file_path: Path) -> pd.DataFrame:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    ws = wb["PT_pp"]
    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[0]
    df = df.drop(index=0)
    df = df[df["CODPP"] != "GRAND TOTAL"]
    return df

def load_prev_inventory_data(file_path: Path) -> pd.DataFrame:
    """
    Load previous-month inventory (Conc sheet).
    Reads evaluated formula values using openpyxl (data_only=True).
    Cleans 'GRAND TOTAL' and ensures numeric conversions.
    """
    from openpyxl import load_workbook
    import pandas as pd

    print(f"🟡 Loading previous inventory from: {file_path}")

    # --- Load evaluated values ---
    wb = load_workbook(file_path, data_only=True)
    if "PT_pp" not in wb.sheetnames:
        raise ValueError("❌ Sheet 'PT_pp' not found in previous inventory file.")
    ws = wb["PT_pp"]

    data = list(ws.values)
    if not data:
        raise ValueError("❌ No data found in sheet 'Conc'.")

    # first row is header
    headers = [str(h).strip() if h is not None else "" for h in data[0]]
    df = pd.DataFrame(data[1:], columns=headers)

    # Remove total rows (like GRAND TOTAL)
    if "CODPP" in df.columns:
        df = df[df["CODPP"] != "GRAND TOTAL"]

    # Identify code column
    code_col = "CODPP"
    if code_col not in df.columns:
        print("❌ Column 'CODPP' not found — please check file header names.")
    else:
        print(f"✅ Using code column: {code_col}")

    # Map expected columns safely
    for src_col, dst_col in [("Qt_SS", "Qt_I"), ("CT_F", "CT_I"), ("CU_F", "CU_I")]:
        if src_col not in df.columns:
            print(f"⚠️ Missing expected column '{src_col}' in previous inventory file.")
            df[dst_col] = 0
        else:
            df[dst_col] = pd.to_numeric(df[src_col], errors="coerce").fillna(0)

    # Normalize product code
    df["CODPP"] = df[code_col].astype(str).str.strip().str.upper()

    # Keep only the relevant columns
    interest_cols = ["CODPP", "Qt_I", "CT_I", "CU_I"]
    dfkeep = df[interest_cols].drop_duplicates("CODPP", keep="first")

    print(f"📦 Loaded {len(dfkeep)} rows from previous inventory (after cleaning)")
    return dfkeep

def load_sales_onfci(resumo_path: Path) -> pd.DataFrame:
    """
    Load O_NFCI sales sheet (2b) reading evaluated formula results.
    Cleans GRAND TOTAL rows and normalizes numeric columns.
    """
    from openpyxl import load_workbook
    import pandas as pd

    print(f"🟢 Loading sales O_NFCI from: {resumo_path}")

    print(f"🟢 Loading sales O_NFCI from: {resumo_path}")
    df = pd.read_excel(resumo_path, sheet_name="O_NFCI", dtype=str)

    # Remove total rows
    if "CODPP" in df.columns:
        df = df[df["CODPP"].astype(str).str.upper() != "GRAND TOTAL"]

    # Convert and clean
    df["CODPP"] = df["CODPP"].astype(str).str.strip().str.upper()
    df["QT"] = pd.to_numeric(df.get("QT", 0), errors="coerce").fillna(0)
    df["PMERC_T"] = pd.to_numeric(df.get("PMERC_T", 0), errors="coerce").fillna(0)
    df["MARGVLR"] = pd.to_numeric(df.get("MARGVLR", 0), errors="coerce").fillna(0)

    out = pd.DataFrame({
        "CODPP": df["CODPP"],
        "VENDAS_2b": df["QT"],
        "VV_2b": df["PMERC_T"],
        "Mrg_2b": df["MARGVLR"]
    })

    onfci_out = out.groupby("CODPP", as_index=False).agg({
        "VENDAS_2b": "sum",
        "VV_2b": "sum",
        "Mrg_2b": "sum"
    })

    print(f"📦 Loaded {len(onfci_out)} O_NFCI rows (after cleaning totals)")
    return onfci_out

def load_sales_llpi(resumo_path: Path) -> pd.DataFrame:
    """
    Load L_LPI sales sheet (2c) using pandas for speed and stability.
    Cleans GRAND TOTAL rows and applies fixed filters.
    """
    import pandas as pd

    print(f"🟣 Loading sales L_LPI from: {resumo_path}")

    try:
        df = pd.read_excel(resumo_path, sheet_name="L_LPI", dtype=str)
    except Exception as e:
        print(f"❌ Error reading L_LPI sheet: {e}")
        return pd.DataFrame()

    if df.empty:
        print("⚠️ L_LPI sheet is empty.")
        return df

    # Clean GRAND TOTAL rows
    if "CODPP" in df.columns:
        df = df[df["CODPP"].astype(str).str.upper() != "GRAND TOTAL"]

    # Apply filters
    if "STATUS" in df.columns and "EMPRESA" in df.columns:
        df = df[df["STATUS"].astype(str).str.upper() != "CANCELADO"]
        df = df[df["EMPRESA"].astype(str).str.upper() == "K"]

    # Normalize and convert numerics
    df["CODPP"] = df["CODPP"].astype(str).str.strip().str.upper()
    df["QT"] = pd.to_numeric(df.get("QT", 0), errors="coerce").fillna(0)
    df["PMERC_T"] = pd.to_numeric(df.get("PMERC_T", 0), errors="coerce").fillna(0)
    df["MargVlr"] = pd.to_numeric(df.get("MargVlr", 0), errors="coerce").fillna(0)

    out = pd.DataFrame({
        "CODPP": df["CODPP"],
        "VENDAS_2c": df["QT"],
        "VV_2c": df["PMERC_T"],
        "Mrg_2c": df["MargVlr"]
    })

    lpi_out = out.groupby("CODPP", as_index=False).agg({
        "VENDAS_2c": "sum",
        "VV_2c": "sum",
        "Mrg_2c": "sum"
    })

    print(f"📦 Loaded {len(lpi_out)} L_LPI rows (after filters & cleaning)")
    return lpi_out
# -----------------------
# Main reconciliation
# -----------------------

def reconcile_inventory(year: int, month: int) -> pd.DataFrame:
    import numpy as np
    import pandas as pd
    from pathlib import Path

    def norm_code(s: pd.Series) -> pd.Series:
        return s.astype(str).str.strip().str.upper()

    # prev / this tags
    prev_y = year if month > 1 else year - 1
    prev_m = month - 1 if month > 1 else 12
    this_tag = f"{year:04d}_{month:02d}"
    prev_tag = f"{prev_y:04d}_{prev_m:02d}"

    # dirs & files
    this_dir   = resolve_month_dir(year, month)
    prev_dir   = resolve_month_dir(prev_y, prev_m)
    tables_dir = resolve_tables_dir(year, month)

    prev_inv_path = find_existing_excel(prev_dir, f"{INV_PREFIX}{prev_tag}")
    this_inv_path = find_existing_excel(this_dir,  f"{INV_PREFIX}{this_tag}")
    resumo_path   = find_existing_excel(this_dir,  f"{RESUMO_PREFIX}{this_tag}")

    # --- load base data ---
    inv_prev = load_prev_inventory_data(prev_inv_path)    # CODPP, Qt_I, CU_I, CT_I
    inv_this = load_curr_inventory_data(this_inv_path)    # CODPP, Qt_SS, CU_F, CT_F, Qt_E, CU_E, CU_S, PGE
    vendas_b = load_sales_onfci(resumo_path)              # CODPP, VENDAS_2b, VV_2b, Mrg_2b
    vendas_c = load_sales_llpi(resumo_path)               # CODPP, VENDAS_2c, VV_2c, Mrg_2c
    prodf    = load_prodf(tables_dir)                     # CODPP, CODPF (estrutura)

    # normalize keys
    for df_ in (inv_prev, inv_this, vendas_b, vendas_c, prodf):
        if "CODPP" in df_.columns:
            df_["CODPP"] = norm_code(df_["CODPP"])
        if "CODPF" in df_.columns:
            df_["CODPF"] = norm_code(df_["CODPF"])

    # --- CONJUNTO UNIÃO DE CÓDIGOS (evita perder SKUs que não estão no inv_this) ---
    codes = set()
    for df_ in (inv_this, inv_prev, vendas_b, vendas_c):
        if "CODPP" in df_.columns:
            codes.update(df_["CODPP"].dropna().tolist())
    dp = pd.DataFrame({"CODPP": sorted(codes)})

    # --- merges (sempre LEFT a partir da união) ---
    dp = dp.merge(inv_prev[["CODPP","Qt_I","CT_I","CU_I"]], on="CODPP", how="left")
    keep_cols_this = [c for c in ["CODPP","Qt_SS","CU_F","CT_F","Qt_E","CU_E","CU_S","PGE"] if c in inv_this.columns]
    dp = dp.merge(inv_this[keep_cols_this], on="CODPP", how="left")
    dp = dp.merge(vendas_b, on="CODPP", how="left")
    dp = dp.merge(vendas_c, on="CODPP", how="left")
    dp = dp[dp["CODPP"] != "GRAND TOTAL"]

    # num fills
    num_cols = ["Qt_I","CT_I","CU_I","Qt_SS","CU_F","CT_F","Qt_E","CU_E","CU_S",
                "VENDAS_2b","VENDAS_2c","VV_2b","VV_2c","Mrg_2b","Mrg_2c"]
    for c in num_cols:
        if c in dp.columns:
            dp[c] = pd.to_numeric(dp[c], errors="coerce").fillna(0)

    # vendas & margens
    dp["VENDAS_tot"] = dp.get("VENDAS_2b", 0) + dp.get("VENDAS_2c", 0)
    dp["Qt_S"]       = dp["VENDAS_tot"]
    dp["VV_tot"]     = dp.get("VV_2b", 0) + dp.get("VV_2c", 0)
    dp["Mrg_tot"]    = dp.get("Mrg_2b", 0) + dp.get("Mrg_2c", 0)

    vv2b   = dp.get("VV_2b", 0)
    vv2c   = dp.get("VV_2c", 0)
    vvtot  = dp.get("VV_tot", 0)
    mrg2b  = dp.get("Mrg_2b", 0)
    mrg2c  = dp.get("Mrg_2c", 0)
    mrgtot = dp.get("Mrg_tot", 0)

    dp["MrgPct_2b"]  = np.where(vv2b != 0, mrg2b / vv2b, 0)
    dp["MrgPct_2c"]  = np.where(vv2c != 0, mrg2c / vv2c, 0)
    dp["MrgPct_tot"] = np.where(vvtot != 0, mrgtot / vvtot, 0)
    for c in ["MrgPct_2b","MrgPct_2c","MrgPct_tot"]:
        dp[c] = dp[c].apply(lambda x: max(x, -1))

    # custos / saldos
    # fallback CU_S
    if "CU_S" not in dp.columns:
        dp["CU_S"] = dp.get("CU_F", 0)

    dp["CT_S"]  = (dp.get("CU_S", 0) * dp.get("Qt_S", 0)).round(2)
    dp["CT_E"]  = (dp.get("CU_E", 0) * dp.get("Qt_E", 0))
    dp["CT_SE"] = (dp.get("CT_I", 0) + dp["CT_E"] - dp["CT_S"]).round(2)
    dp["CT_SS"] = (dp.get("Qt_SS", 0) * dp.get("CU_F", 0)).round(2)

    dp["Qt_SE"]   = dp.get("Qt_I", 0) + dp.get("Qt_E", 0) - dp.get("Qt_S", 0)
    dp["Qt_Diff"] = dp.get("Qt_SS", 0) - dp["Qt_SE"]

    # inicializa gerado SEM filtrar por Ins
    #dp["Qt_Ger"] = np.where(dp["Qt_Diff"] == 0, dp.get("Qt_SS", 0), np.nan)
    dp["CT_Diff"] = (dp["CT_SS"] - dp["CT_SE"]).round(2)
    #dp["CT_Ger"]  = np.where(dp["Qt_Diff"] == 0, dp["CT_SS"], dp["CT_SE"])  # evita NaN

    # Ins agora calculado na base unificada (não só em inv_this)
    prodf_parts = prodf["CODPP"].drop_duplicates()
    dp["Ins"] = np.where(~dp["CODPP"].isin(prodf_parts), "I", None)

    dp["CU_Diff"] = np.where(dp.get("Qt_SS", 0) > 0, dp.get("CU_F", 0) - dp.get("CU_I", 0), 0).round(2)
    dp["AnoMes"]  = (year - 2000) * 100 + month

    dp["CODPP"] = norm_code(dp["CODPP"])
    dp = dp.sort_values("CODPP", kind="stable").reset_index(drop=True)

    final_cols_order = [
        "CODPP", "Ins",
        "Qt_I", "Qt_E", "Qt_S", "Qt_SE", "Qt_SS", "Qt_Diff", #"Qt_Ger",
        "CT_I", "CT_E", "CT_S", "CT_SE", "CT_SS", "CT_Diff", #"CT_Ger",
        "CU_I", "CU_E", "CU_S", "CU_F",
        "VENDAS_2b", "VENDAS_2c", "VENDAS_tot",
        "VV_2b", "VV_2c", "VV_tot",
        "Mrg_2b", "Mrg_2c", "Mrg_tot",
        "MrgPct_2b", "MrgPct_2c", "MrgPct_tot",
        "CU_Diff", "AnoMes"
    ]
    dp = dp[[c for c in final_cols_order if c in dp.columns]]
    return dp

def apply_excel_formatting(ws, df, wb):
    """Apply column widths and styles to an Excel worksheet."""
    col_widths = {
        "CODPP": 10,
        "Ins": 2,

        # Quantities
        "Qt_I": 6, "Qt_E": 5, "Qt_S": 5, "Qt_SE": 6, "Qt_SS": 6,
        "Qt_Diff": 5, "Qt_Ger": 6,

        # Adjustments
        "Qt_Aj": 5, "Qt_AjF": 5, "CT_Aj": 9, "CT_AjF": 9,

        # Costs
        "CT_I": 10, "CT_E": 10, "CT_S": 10, "CT_SE": 10,
        "CT_SS": 10, "CT_Diff": 10, "CT_Ger": 10,

        # Unit costs
        "CU_I": 6, "CU_E": 6, "CU_S": 6, "CU_F": 6,

        # Sales and margins
        "VV_2b": 10, "VV_2c": 10, "VV_tot": 10,
        "Mrg_2b": 10, "Mrg_2c": 10, "Mrg_tot": 10,
        "MrgPct_2b": 6, "MrgPct_2c": 6, "MrgPct_tot": 6,

        # Vendas
        "VENDAS_2b": 9, "VENDAS_2c": 9, "VENDAS_tot": 9,
    }

    # --- Base formats ---
    qt_blue   = wb.add_format({'num_format': '#,##0',   'bg_color': '#DDEBF7'})
    qt_gray   = wb.add_format({'num_format': '#,##0',   'bg_color': "#E3EEF7"})
    ct_orange = wb.add_format({'num_format': '#,##0.00','bg_color': "#D7A167"})
    ct_gray   = wb.add_format({'num_format': '#,##0.00','bg_color': "#E3EEF7"})
    cu_green  = wb.add_format({'num_format': '#,##0.00','bg_color': "#8ADBEA"})

    # --- Additional formats ---
    blue_money_fmt   = wb.add_format({'num_format': '#,##0.00', 'bg_color': '#DDEBF7'})
    green_money_fmt  = wb.add_format({'num_format': '#,##0.00', 'bg_color': '#E2EFDA'})
    blue_pct_fmt = wb.add_format({'num_format': '0%', 'bg_color': '#DDEBF7'})
    orange_money_fmt = wb.add_format({'num_format': '#,##0.00', 'bg_color': '#FCE4D6'})
    lightblue_int_fmt = wb.add_format({'num_format': '#,##0', 'bg_color': '#BDD7EE'})  # for Qt_Aj / Qt_AjF
    lightblue_money   = wb.add_format({'num_format': '#,##0.00', 'bg_color': '#BDD7EE'})  # for CT_Aj / CT_AjF

    # --- Column groups ---
    qt_blue_cols  = {"Qt_I","Qt_E","Qt_S","Qt_SE","Qt_SS"}
    qt_gray_cols  = {"Qt_Diff","Qt_Ger"}
    ct_orange_cols= {"CT_I","CT_E","CT_S","CT_SE","CT_SS"}
    ct_gray_cols  = {"CT_Diff","CT_Ger"}
    cu_green_cols = {"CU_I","CU_E","CU_S","CU_F"}
    blue_cols     = {"VV_2b","VV_2c","VV_tot"}  # VV_tot same style
    green_cols    = {"Mrg_2b","Mrg_2c","Mrg_tot"}
    pct_cols      = {"MrgPct_2b","MrgPct_2c","MrgPct_tot"}
    orange_cols   = {"VENDAS_2b","VENDAS_2c","VENDAS_tot"}
    aj_int_cols   = {"Qt_Aj","Qt_AjF"}
    aj_money_cols = {"CT_Aj","CT_AjF"}

    # --- Apply ---
    ws.autofilter(0, 0, df.shape[0], df.shape[1] - 1)
    for idx, col in enumerate(df.columns):
        width = col_widths.get(col, 12)
        fmt = None
        if   col in qt_gray_cols:   fmt = qt_gray
        elif col in qt_blue_cols:   fmt = qt_blue
        elif col in ct_orange_cols: fmt = ct_orange
        elif col in ct_gray_cols:   fmt = ct_gray
        elif col in cu_green_cols:  fmt = cu_green
        elif col in blue_cols:      fmt = blue_money_fmt
        elif col in green_cols:     fmt = green_money_fmt
        elif col in pct_cols:       fmt = blue_pct_fmt
        elif col in orange_cols:    fmt = orange_money_fmt
        elif col in aj_int_cols:    fmt = lightblue_int_fmt
        elif col in aj_money_cols:  fmt = lightblue_money

        ws.set_column(idx, idx, width, fmt)


def adjust_missing_inventory_progressive(dp):
    """
    Adjusts missing/excess inventory progressively within ±2% of CT_S total.
    Step 1: Fully offsets the smaller of the two sides (positive vs. negative Qt_Diff),
            removing neutralized items from further consideration.
    Step 2: Adds the value of the neutralized side to the available budget.
    Step 3: Uses that budget to adjust the remaining side progressively,
            starting with the cheapest items first.
    """

    import numpy as np

    # --- Step 1. Setup and budget ---
    total_ct_s = dp["CT_S"].sum(skipna=True)
    budget_limit = total_ct_s * 0.02
    print(f"💰 Total cost of goods sold (CT_S): {total_ct_s:,.2f}")
    print(f"🎯 Budget range: ±{budget_limit:,.2f}")

    # Pending rows = not reconciled and not already 'Ins'
    mask_pending = (dp["CT_Ger"].isna() | dp["Qt_Ger"].isna()) & (dp["Ins"] != "I")
    pending = dp.loc[mask_pending].copy()
    pending = pending.dropna(subset=["CU_F", "Qt_Diff"])
    pending["Qt_Aj"] = 0.0
    pending["CT_DiffVal"] = pending["Qt_Diff"] * pending["CU_F"]

    # --- Step 2. Separate sides ---
    pos = pending[pending["Qt_Diff"] > 0].copy()
    neg = pending[pending["Qt_Diff"] < 0].copy()
    total_pos = (pos["Qt_Diff"] * pos["CU_F"]).sum()
    total_neg = (neg["Qt_Diff"] * neg["CU_F"]).sum()  # will be negative
    print(f"➕ Positive side total: {total_pos:,.2f}")
    print(f"➖ Negative side total: {total_neg:,.2f}")

    # --- Step 3. Determine smaller side and neutralize it completely ---
    smaller_side = "pos" if abs(total_pos) <= abs(total_neg) else "neg"
    offset_value = min(abs(total_pos), abs(total_neg))
    print(f"⚖️  Natural offset applied (no budget impact): {offset_value:,.2f}")
    print(f"🧭 Smaller side: {smaller_side}")

    # Fully neutralize smaller side
    if smaller_side == "pos":
        pending.loc[pos.index, "Qt_Aj"] = pending.loc[pos.index, "Qt_Diff"]
    else:
        pending.loc[neg.index, "Qt_Aj"] = pending.loc[neg.index, "Qt_Diff"]

    # Remove those items from further adjustment
    if smaller_side == "pos":
        remaining = pending.loc[neg.index].copy()
    else:
        remaining = pending.loc[pos.index].copy()

    # --- Step 4. Add offset value to the usable budget ---
    remaining_budget = budget_limit + offset_value
    print(f"💵 Budget available for adjustment after neutralization: {remaining_budget:,.2f}")

    # --- Step 5. Adjust remaining side progressively (round-robin, cheapest first) ---
    from itertools import cycle

    remaining = remaining.sort_values("CU_F", ascending=True).copy()
    total_used = 0.0

    # Initialize helper columns if missing
    if "Qt_Aj" not in remaining.columns:
        remaining["Qt_Aj"] = 0.0

    # Create cycling iterator over item indices
    active = remaining.index.tolist()
    iter_cycle = cycle(active)

    while active and total_used < remaining_budget:
        i = next(iter_cycle)
        if i not in active:
            continue

        unit_cost = remaining.at[i, "CU_F"]
        diff = remaining.at[i, "Qt_Diff"]
        adj  = remaining.at[i, "Qt_Aj"]
        direction = np.sign(diff)

        # Still has room to adjust and budget left?
        if abs(adj) < abs(diff) and total_used + unit_cost <= remaining_budget:
            remaining.at[i, "Qt_Aj"] += direction
            total_used += unit_cost
        else:
            # This item is fully adjusted or over budget—remove it from cycle
            active.remove(i)
            if not active:
                break
            iter_cycle = cycle(active)
    # --- Sync back round-robin results into pending ---
    pending.loc[remaining.index, "Qt_Aj"] = remaining["Qt_Aj"].astype(float)

    print(f"✅ Budget used for remaining side: {total_used:,.2f}")
    print(f"📊 Final budget usage: {total_used / remaining_budget * 100:+.1f}%")

    # --- Step 6. Compute costs and merge back ---
    pending["CT_Aj"]  = pending["Qt_Aj"] * pending["CU_F"]
    pending["Qt_AjF"] = pending["Qt_Diff"] - pending["Qt_Aj"]
    pending["CT_AjF"] = pending["Qt_AjF"] * pending["CU_F"]


    # --- Merge back into dp ---
    for col in ["Qt_Aj", "CT_Aj", "Qt_AjF", "CT_AjF"]:
        if col not in dp.columns:
            dp.loc[:, col] = 0.0
        dp.loc[pending.index, col] = pending[col].values

    # --- Final reconciled balances (uniform, avoids notna pitfalls) ---
    dp.loc[:, "Qt_Ger"] = dp["Qt_SE"] + dp["Qt_Aj"]
    dp.loc[:, "CT_Ger"] = dp["CT_SE"] + dp["CT_Aj"]

    print(f"✅ Offset neutralized: {offset_value:,.2f}")
    print(f"✅ Budget used for remaining side: {total_used:,.2f}")
    print(f"📊 Final budget usage: {total_used / remaining_budget * 100:+.1f}%")

    # --- Update final reconciled quantities and costs after adjustment ---
    dp["Qt_Ger"] = dp["Qt_SE"] + dp["Qt_Aj"]
    dp["CT_Ger"] = dp["CT_SE"] + dp["CT_Aj"]


    return dp

def adjust_missing_inventory_budget(dp):
    """
    Adjusts missing/excess inventory by booking the difference into a single item 'Ajuste_Estoque'.
    
    Step 1: Calculates natural offset between positive and negative Qt_Diff sides.
    Step 2: Adds the smaller side (offset value) to the available adjustment budget (2% of CT_S total).
    Step 3: Applies the available budget as a value adjustment (CT_Ger) into an 'Ajuste_Estoque' item.
            - If it doesn’t exist, creates it.
            - If it exists, increments/decrements its CT_Ger accordingly.
    """

    import numpy as np
    import pandas as pd

    # --- Step 1. Setup and budget ---
    total_ct_s = dp["CT_S"].sum(skipna=True)
    budget_limit = total_ct_s * 0.02
    print(f"💰 Total cost of goods sold (CT_S): {total_ct_s:,.2f}")
    print(f"🎯 Budget range: ±{budget_limit:,.2f}")

    # --- Step 1. Compute difference for Ins = "I" items ---
    pDiff = dp.loc[dp["Ins"] != "I", "CT_Diff"].sum(skipna=True)
    print(f"📊 Total CT_Diff for Ins!='I': {pDiff:,.2f}")

    # --- Step 2. Determine how much we can adjust (smaller of budget or diff) ---
    pAdjust = np.sign(pDiff) * min(abs(pDiff), budget_limit)
    print(f"⚖️ Difference value: {pDiff:,.2f}")
    print(f"💵 Will adjust by (limited to budget): {pAdjust:,.2f}")

    # --- Step 3. Apply or create 'Ajuste_Estoque' record ---
    if "Ajuste_Estoque" in dp["CODPP"].values:
        idx_adj = dp.index[dp["CODPP"] == "Ajuste_Estoque"][0]
        dp.at[idx_adj, "CT_SS"] = dp.get("CT_SS", pd.Series(0, index=dp.index)).iloc[idx_adj] + pAdjust
        print(f"🔁 Updated existing 'Ajuste_Estoque' CT_SS → {dp.at[idx_adj, 'CT_SS']:,.2f}")
    else:
        print("🆕 Creating 'Ajuste_Estoque' item in dataset...")
        new_row = {
            "CODPP": "Ajuste_Estoque",
            "Qt_I": 0.0, "Qt_E": 0.0, "Qt_S": 0.0, "Qt_SE": 0.0,
            "CT_I": 0.0, "CT_E": 0.0, "CT_S": 0.0, "CT_SE": 0.0,
            "CU_F": 0.0,
            "CT_SS": pAdjust, "CT_Diff": -pAdjust
        }
        dp = pd.concat([dp, pd.DataFrame([new_row])], ignore_index=True)
        print(f"✅ Inserted new 'Ajuste_Estoque' with CT_SS = {pAdjust:,.2f}")

    return dp

# -----------------------
# CLI / Runner
# -----------------------

def main(year: int, month: int, save_excel: bool = True) -> Path:
    this_dir = resolve_month_dir(year, month)
    if not this_dir:
        raise FileNotFoundError(f"Could not resolve base dir for {yymm_to_str(year, month)}")

    out_dir = this_dir
    ensure_dir(out_dir)

    # Run reconciliation (child level)
    report = reconcile_inventory(year, month)

    # Load parent mapping
    tables_dir = resolve_tables_dir(year, month)
    prodf = load_prodf(tables_dir)

    # Build parent-level aggregation
    first_cols = ["Qt_E", "CU_F", "CU_Pai"]  # add any others you want

    agg_map = {}
    for col in report.columns:
        if col in ["CODPF","CU_F", "CODPP"]:
            continue
        elif col in first_cols:
            agg_map[col] = "first"   # keep as is (same for all children under same parent)
        elif pd.api.types.is_numeric_dtype(report[col]):
            agg_map[col] = "sum"
        else:
            agg_map[col] = "first"

    parent_report = report

    # Define colunas na ordem exata da aba 'Parent' desejada
    final_cols_order = [
        "CODPP", "Ins", "Qt_I", "Qt_E", "Qt_S", "Qt_SE", "Qt_SS", "Qt_Diff", "Qt_Ger",
        "CT_I", "CT_E", "CT_S", "CT_SE", "CT_SS", "CT_Diff", "CT_Ger",
        "CU_I", "CU_E", "CU_S", "CU_F",
        "VENDAS_2b", "VENDAS_2c", "VENDAS_tot",
        "VV_2b", "VV_2c", "VV_tot",
        "Mrg_2b", "Mrg_2c", "Mrg_tot",
        "MrgPct_2b", "MrgPct_2c", "MrgPct_tot"
    ]
    # Filtra colunas existentes (caso alguma falte)
    parent_report = parent_report[[col for col in final_cols_order if col in parent_report.columns]]
    parent_report = adjust_missing_inventory_budget(parent_report)

    # Save CSV and XLSX
    tag = yymm_to_str(year, month)
    xlsx_path = out_dir / f"Conc_Estoq_{tag}.xlsx"

    if save_excel:
        with pd.ExcelWriter(xlsx_path, engine="xlsxwriter") as writer:
            parent_report.to_excel(writer, index=False, sheet_name="Conc")
            wb = writer.book
            ws_child = writer.sheets["Conc"]
            apply_excel_formatting(ws_child, parent_report, wb)

    print(f"Saved: {xlsx_path}")
    return xlsx_path

if __name__ == "__main__":
    import argparse
    from datetime import datetime

    ap = argparse.ArgumentParser(description="Inventory reconciliation for a given year/month.")
    ap.add_argument("--year", "-y", type=int, help="Year, e.g. 2025")
    ap.add_argument("--month", "-m", type=int, help="Month, 1-12")
    args = ap.parse_args()

    # If missing, ask interactively
    if args.year is None or args.month is None:
        now = datetime.now()
        print("Year and/or month not provided.")
        year = int(input(f"Enter year (default {now.year}): ") or now.year)
        month = int(input(f"Enter month [1-12] (default {now.month -1}): ") or (now.month -1))
    else:
        year, month = args.year, args.month

    main(year, month)
