import os
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import gc  # Garbage collector to free memory
import psutil  # To check memory usage


# Global variables
timeframe = 6  # Default: Last 6 months
ano_x = 2025
mes_x = 10

# Define potential base directories
path_options = [
    '/Users/mauricioalouan/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data/',
    '/Users/simon/Library/CloudStorage/Dropbox/KBB MF/AAA/Balancetes/Fechamentos/data/'
]

# Find the correct base directory
base_dir = None
for path in path_options:
    if os.path.exists(path):
        base_dir = path
        break

if base_dir is None:
    raise FileNotFoundError("❌ None of the specified base directories exist.")

print("✅ Base directory set to:", base_dir)

# Function to get file paths for the last `timeframe` months
def get_last_n_files(base_dir, ano_x, mes_x, n):
    file_paths = []
    
    for i in range(n):
        # Compute the year and month dynamically
        year = ano_x
        month = mes_x - i

        if month <= 0:  # Adjust for year change
            month += 12
            year -= 1

        folder = os.path.join(base_dir, f"{year:04}_{month:02}")
        filename = f"R_Resumo_{year:04}_{month:02}.xlsm"
        file_path = os.path.join(folder, filename)

        if os.path.exists(file_path):
            file_paths.append(file_path)
        else:
            print(f"⚠️ Warning: File {file_path} not found. Skipping...")

    return file_paths

# Function to stack sheets from multiple files
def stack_sheets(file_paths):
    stacked_data = {}
    first_file = True  # Track the first file (latest month)
    latest_file = None  # Store the latest file path

    for file_path in file_paths:
        print(f"\n📂 Attempting to load file: {file_path}")

        try:
            if first_file:
                print(f"🔍 Opening {file_path} with VBA macros (KEEP FORMATTING)...")
                latest_file = file_path  # Store the latest file for copying
                wb = load_workbook(file_path, keep_vba=True, read_only=True)  # Read-only mode first
                print(f"✅ Workbook loaded with macros: {file_path}")
                first_file = False
            else:
                print(f"🔍 Opening {file_path} WITHOUT macros (FASTER)...")
                wb = load_workbook(file_path, keep_vba=False, read_only=True)
                print(f"✅ Workbook loaded without macros: {file_path}")

        except Exception as e:
            print(f"❌ Error loading {file_path}: {e}")
            continue

        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("Pivot"):  # Skip Pivot sheets
                continue

            print(f"🔹 Checking sheet: {sheet_name} in {file_path}")

            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
                print(f"✅ Loaded {sheet_name}: {df.shape}")

                if sheet_name not in stacked_data:
                    stacked_data[sheet_name] = df
                else:
                    # Alinha as colunas atuais com as do acumulado, ignorando extras
                    df_aligned = df.reindex(columns=stacked_data[sheet_name].columns)
                    stacked_data[sheet_name] = pd.concat([stacked_data[sheet_name], df_aligned], ignore_index=True)
                    print(f"🔄 Stacked {sheet_name}: {stacked_data[sheet_name].shape}")

            except Exception as e:
                print(f"❌ Error reading {sheet_name}: {e}")

        # Free memory
        del wb
        gc.collect()

    return stacked_data, latest_file

# Function to save the stacked data
def save_stacked_data(stacked_data, output_file):
    print(f"✅ Loading template {output_file} to insert stacked data...")

    # Load workbook (with macros)
    wb = load_workbook(output_file, keep_vba=True)

    # Add stacked data as new sheets
    for sheet_name, df in stacked_data.items():
        ws = wb.create_sheet(title=sheet_name)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        print(f"✅ Added stacked sheet: {sheet_name}")

    # Save final output
    wb.save(output_file)
    print(f"✅ Stacked data saved as: {output_file}")

# Main function
def main():
    # Get last `timeframe` files
    file_paths = get_last_n_files(base_dir + "clean/", ano_x, mes_x, timeframe)

    if not file_paths:
        print("❌ No valid files found. Exiting...")
        return

    # Stack sheets from all months
    stacked_data, _ = stack_sheets(file_paths)

    # Set template path
    template_file = os.path.join(base_dir, "template", "Stacktemplate.xlsm")

    # Ensure template exists
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"❌ Template file {template_file} not found.")

    # Define output file path in the last month’s folder
    last_month_folder = os.path.dirname(file_paths[0])
    output_file = os.path.join(last_month_folder, f"R_ResumoU6M_{ano_x:04}_{mes_x:02}.xlsm")

    # Copy the template file to the output file location
    print(f"✅ Copying template {template_file} to {output_file} (preserving macros)...")
    shutil.copy(template_file, output_file)

    # Save stacked data into the copied template file
    save_stacked_data(stacked_data, output_file)

# Run script
if __name__ == "__main__":
    main()
