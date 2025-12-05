import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# === CONFIG ===
YEAR = "2025"
MONTH = "11-Novembro"

# Define base folder and available series
path_options = [
    '/Users/mauricioalouan/Dropbox/nfs',
    '/Users/simon/Library/CloudStorage/Dropbox/nfs'
]
BASE_FOLDER = None
for path in path_options:
    if os.path.exists(path):
        BASE_FOLDER = path
        break

if not BASE_FOLDER:
    print("⚠️ Warning: No valid BASE_FOLDER found.")
    BASE_FOLDER = "/Users/mauricioalouan/Dropbox/nfs"

# The same series used in Tabela_NFs.py
SERIES_LIST = [
    "Serie 1 - Omie",
    "Serie 2 - filial",
    "Serie 3 - Bling",
    "Serie 4 - Lexos",
    "Serie 5 - Olist",
    "Serie 6 - Meli",
    "Serie 7 - Amazon",
    "Serie 8 - Magalu Full",
    "Serie 9 - Shopee Full"
]

# === MAIN FUNCTION ===
def combine_monthly_excels(year, month):
    # Define output directory
    month_num = str(month).split('-')[0].zfill(2)
    output_dir = os.path.join(BASE_FOLDER, "Mauricio", "Contabilidade", f"{year}_{month_num}")
    
    if not os.path.exists(output_dir):
        print(f"⚠️ Output directory not found: {output_dir}")
        # Depending on logic, maybe we should create it or just return if inputs are expected there?
        # Inputs are expected there, so if it doesn't exist, we can't aggregate.
        return

    # Define path for lookup tables and Resumo
    # Assuming BASE_FOLDER ends with /nfs, we strip it to get the root Dropbox folder
    dropbox_root = os.path.dirname(BASE_FOLDER)
    tables_dir = os.path.join(dropbox_root, "KBB MF", "AAA", "Balancetes", "Fechamentos", "data", "Tables")
    
    lookup_file = os.path.join(tables_dir, "T_NFTipo.xlsx")
    resumo_file = os.path.join(tables_dir, "R_ResumoFin25.xlsx")

    # Load Lookup Table
    try:
        lookup_df = pd.read_excel(lookup_file)
        # Create a dictionary for faster lookup: Natureza_NF -> Natureza_Grp
        # Ensure columns exist
        if "Natureza_NF" in lookup_df.columns and "Natureza_Grp" in lookup_df.columns:
            lookup_map = dict(zip(lookup_df["Natureza_NF"], lookup_df["Natureza_Grp"]))
        else:
            print(f"⚠️ Warning: Columns 'Natureza_NF' or 'Natureza_Grp' not found in {lookup_file}")
            lookup_map = {}
    except Exception as e:
        print(f"❌ Error reading lookup table {lookup_file}: {e}")
        lookup_map = {}

    all_data = []
    for series in SERIES_LIST:
        filename = f"NF_{year}_{month_num}_{series}.xlsx"
        file_path = os.path.join(output_dir, filename)
        
        if not os.path.exists(file_path):
            print(f"⚠️ Skipping missing file: {file_path}")
            continue
        
        try:
            df = pd.read_excel(file_path)
            df.insert(0, "Series", series)  # add series name
            
            # Add Natureza_GRP column
            # Use map and fillna with 999
            df["Natureza_GRP"] = df["Natureza"].map(lookup_map).fillna(999)
            
            all_data.append(df)
            print(f"📂 Added: {series} ({len(df)} rows)")
        except Exception as e:
            print(f"❌ Error reading {series}: {e}")
    
    if not all_data:
        print("No data files found — nothing to combine.")
        return
    
    # Save combined file
    # combined_file = os.path.join(output_dir, f"NF_{year}_{month_num}_todos.xlsx")
    # combined_df.to_excel(combined_file, index=False)
    
    # Use Template
    template_file = os.path.join(dropbox_root, "KBB MF", "AAA", "Balancetes", "Fechamentos", "data", "Template", "NF_XML.xlsm")
    
    if not os.path.exists(template_file):
        print(f"❌ Template file not found: {template_file}")
        return

    print(f"Loading template: {template_file}")
    wb = load_workbook(template_file, keep_vba=True)
    
    # Target sheet
    if "NF" in wb.sheetnames:
        ws = wb["NF"]
        ws.delete_rows(2, ws.max_row)
    else:
        ws = wb.create_sheet("NF")
        
    # Write dataframe to sheet
    for r in dataframe_to_rows(combined_df, index=False, header=True):
        ws.append(r)
        
    # Delete Sheet1 if exists
    if "Sheet1" in wb.sheetnames:
        del wb["Sheet1"]
        
    # Save as XLSM
    combined_file = os.path.join(output_dir, f"NF_{year}_{month_num}_todos.xlsm")
    wb.save(combined_file)
    
    print(f"✅ Combined Excel created: {combined_file}")
    print(f"📊 Total rows combined: {len(combined_df)}")

    # === UPDATE RESUMO ===
    print("🔄 Updating ResumoFin25...")
    
    # Calculate totals by Natureza_GRP
    # We need to aggregate by Natureza_GRP
    # Metrics needed:
    # nfs: nunique(NF)
    # linhas: count()
    # itens: 0 (for NF file)
    # vProd: sum(ValorProduto)
    # vICMS: sum(ICMS)
    # vIPI: sum(IPI)
    # vFrete: sum(Frete)
    # Vdesc: sum(Desconto)
    # vNF: sum(TotalNF)
    
    # Group by Natureza_GRP
    grouped = combined_df.groupby("Natureza_GRP")
    
    totals = {}
    for name, group in grouped:
        totals[f"MA_NF_{name}_nfs"] = group["NF"].nunique()
        totals[f"MA_NF_{name}_linhas"] = len(group)
        totals[f"MA_NF_{name}_itens"] = 0 # Not available in NF file
        totals[f"MA_NF_{name}_vProd"] = group["ValorProduto"].sum()
        totals[f"MA_NF_{name}_vICMS"] = group["ICMS"].sum()
        totals[f"MA_NF_{name}_vIPI"] = group["IPI"].sum()
        totals[f"MA_NF_{name}_vFrete"] = group["Frete"].sum()
        totals[f"MA_NF_{name}_Vdesc"] = group["Desconto"].sum()
        totals[f"MA_NF_{name}_vNF"] = group["TotalNF"].sum()

    if not os.path.exists(resumo_file):
        print(f"❌ Resumo file not found: {resumo_file}")
        return

    try:
        wb = load_workbook(resumo_file)
        if "Numbers" not in wb.sheetnames:
             print(f"❌ Sheet 'Numbers' not found in {resumo_file}")
             return
        ws = wb["Numbers"]
        
        # Find column for current month (YYMM)
        target_col_header = f"{year[-2:]}{month_num}" # e.g. 2511
        target_col_idx = None
        
        # Iterate through header row (assuming row 1)
        for cell in ws[1]:
            if str(cell.value) == target_col_header:
                target_col_idx = cell.column
                break
        
        if not target_col_idx:
            print(f"❌ Column '{target_col_header}' not found in Resumo file.")
            return

        # Update values
        # Iterate through rows in column A to find keys
        updates_count = 0
        for row in range(2, ws.max_row + 1):
            key = ws.cell(row=row, column=1).value
            if key in totals:
                ws.cell(row=row, column=target_col_idx).value = totals[key]
                updates_count += 1
        
        wb.save(resumo_file)
        print(f"✅ Resumo updated: {updates_count} values written to column {target_col_header}")

    except Exception as e:
        print(f"❌ Error updating Resumo: {e}")

# === RUN ===
def main(year, month):
    year_str = str(year)
    # Just pass the month number as string, e.g. "11"
    month_str = f"{month:02d}"
    combine_monthly_excels(year_str, month_str)

if __name__ == "__main__":
    import argparse
    from datetime import datetime
    
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", "-y", type=int)
    parser.add_argument("--month", "-m", type=int)
    args = parser.parse_args()
    
    if args.year and args.month:
        main(args.year, args.month)
    else:
        # Default or interactive
        now = datetime.now()
        def_year = now.year
        def_month = now.month - 1 if now.month > 1 else 12
        if def_month == 12: def_year -= 1
        
        print(f"Using default/interactive mode. Default: {def_year}-{def_month}")
        try:
            y = int(input(f"Year [{def_year}]: ") or def_year)
            m = int(input(f"Month [{def_month}]: ") or def_month)
            main(y, m)
        except:
            main(def_year, def_month)
